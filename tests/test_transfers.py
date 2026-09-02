import hashlib
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation


TRANSFER_HEADERS = [
    "No.",
    "Transfer Date",
    "Standart/Low-cost",
    "Asset Tag No. / Inventory Code\n(new standardised system)",
    "Brand-Make/Model/Item Description",
    "Serial/Chassis No.",
    "Holder",
    "Current Project",
    "New Project",
    "New holder",
    "Current Status\n(functionality)",
    "Asset Condition Description",
    "Reason for Asset Transfer",
]

ASSET_HEADERS = [
    "Asset Tag No. / Inventory Code\n(new standardised system)",
    "Asset Classification",
    "Asset Sub Classification",
    "Item Description",
    "Brand / Make ",
    "Model",
    "Serial/ Chassis No.",
    "Quantity",
    "Purchase price",
    "Currency",
    "Current Status\n(functionality)",
]


def create_asset_sheet(workbook, title, stale_tag):
    sheet = workbook.create_sheet(title)
    for column_number, header in enumerate(ASSET_HEADERS, start=1):
        sheet.cell(row=8, column=column_number).value = header
    sheet.cell(row=8, column=25).value = "Last date of transfer"
    sheet.cell(row=9, column=1).value = stale_tag
    sheet.cell(row=9, column=4).value = "Foreign stale asset"
    sheet.cell(row=9, column=25).value = "=IF(A9<>\"\",\"template\",\"\")"
    sheet.column_dimensions["A"].width = 31
    sheet.column_dimensions["B"].hidden = True
    sheet.freeze_panes = "A9"
    return sheet


def create_transfer_workbook(path, *, include_system_id=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transfer log"
    for column_number, header in enumerate(TRANSFER_HEADERS, start=3):
        sheet.cell(row=1, column=column_number).value = header
    if include_system_id:
        sheet.cell(row=1, column=16).value = "System Transfer ID"
        sheet.column_dimensions["P"].hidden = True
    workbook.save(path)


def populate_legacy_transfer_rows(sheet, count):
    for index in range(1, count + 1):
        row_number = index + 1
        values = [
            index,
            "29.08.2026",
            "Standard",
            f"LEGACY-{index:04d}",
            f"Legacy item {index}",
            None,
            "Warehouse",
            "LEGACY-PROJECT",
            "LEGACY-PROJECT",
            "Legacy holder",
            "functional",
            None,
            "Legacy transfer",
        ]
        for column_number, value in enumerate(values, start=3):
            sheet.cell(row=row_number, column=column_number).value = value


def existing_transfer_record(transfer_id=188):
    return {
        "system_transfer_id": transfer_id,
        "source_log_no": None,
        "transfer_date": "2026-08-29",
        "source_asset_type": "Standard",
        "asset_tag_number": "SYN-T2-0001",
        "from_holder_name": "Warehouse",
        "from_project_raw": "SYN-20001",
        "to_project_raw": "SYN-20001",
        "to_holder_name": "Synthetic Tenant #2 Admin",
        "asset_status": "functional",
        "transfer_reason": "Assignment changed in web app",
    }


def transfer_context(app_module, *transfers):
    return {
        "person_lookup": {},
        "project_lookup": {},
        "transfers_by_id": {
            transfer["transfer_id"]: transfer
            for transfer in transfers
            if transfer.get("transfer_id") is not None
        },
        "transfer_signatures": {
            app_module.make_transfer_signature(transfer["asset_id"], transfer)
            for transfer in transfers
        },
        "registration_asset_ids": {
            transfer["asset_id"]
            for transfer in transfers
            if app_module.is_registration_transfer_record(transfer)
        },
    }


def test_normalize_excel_transfer_record_requires_asset_and_date(app_module):
    assert app_module.normalize_excel_transfer_record({"asset_tag_number": "", "transfer_date": "17.04.2026"}, 8) is None
    assert app_module.normalize_excel_transfer_record({"asset_tag_number": "HELP-UKR-0753"}, 8) is None


def test_normalize_excel_transfer_record_maps_transfer_fields(app_module):
    record = app_module.normalize_excel_transfer_record(
        {
            "asset_tag_number": " help-ukr-0753 ",
            "transfer_date": "17.04.2026",
            "source_log_no": "176",
            "from_holder_name": "Supplier",
            "to_holder_name": "Warehouse",
            "transfer_reason": "Registration of the new asset",
            "asset_status": "functional",
        },
        9,
    )

    assert record["asset_tag_number"] == "HELP-UKR-0753"
    assert record["transfer_date"] == "2026-04-17"
    assert record["source_log_no"] == 176
    assert record["transfer_key"] == "registration:HELP-UKR-0753"


def test_transfer_export_roundtrip_uses_hidden_stable_id_and_skips_existing(
    app_module,
    tmp_path,
):
    workbook_path = tmp_path / "roundtrip.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Transfer log"]

    result = app_module.write_transfer_log_records_to_excel_sheet(
        sheet,
        [existing_transfer_record()],
    )
    workbook.save(workbook_path)
    workbook.close()

    reopened = load_workbook(workbook_path)
    assert reopened["Transfer log"]["P1"].value == "System Transfer ID"
    assert reopened["Transfer log"].column_dimensions["P"].hidden is True
    reopened.close()

    imported = app_module.load_excel_transfer_log_rows(str(workbook_path))
    context = transfer_context(
        app_module,
        {
            "transfer_id": 188,
            "asset_id": 1232,
            **existing_transfer_record(),
        },
    )
    preview = app_module.build_transfer_log_preview(
        imported,
        {"SYN-T2-0001": {"asset_id": 1232, "asset_tag_number": "SYN-T2-0001"}},
        context,
    )

    assert result["exported_records"] == 1
    assert imported[0]["system_transfer_id"] == 188
    assert imported[0]["source_log_no"] == 1
    assert preview["summary"]["new_records"] == 0
    assert preview["summary"]["skipped_existing"] == 1


def test_transfer_export_numbers_null_source_logs_after_record_max(app_module, tmp_path):
    workbook_path = tmp_path / "numbered-transfer-log.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    records = [
        {**existing_transfer_record(188), "source_log_no": 160},
        existing_transfer_record(189),
        existing_transfer_record(190),
    ]

    app_module.write_transfer_log_records_to_excel_sheet(
        workbook["Transfer log"],
        records,
    )
    workbook.save(workbook_path)
    workbook.close()

    imported = app_module.load_excel_transfer_log_rows(str(workbook_path))
    assert [row["source_log_no"] for row in imported] == [160, 161, 162]
    assert [row["system_transfer_id"] for row in imported] == [188, 189, 190]
    assert records[1]["source_log_no"] is None
    assert records[2]["source_log_no"] is None


def test_transfer_export_numbers_null_source_logs_after_worksheet_max(app_module, tmp_path):
    workbook_path = tmp_path / "existing-numbered-transfer-log.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Transfer log"]
    populate_legacy_transfer_rows(sheet, 160)

    app_module.write_transfer_log_records_to_excel_sheet(
        sheet,
        [existing_transfer_record(188)],
    )
    workbook.save(workbook_path)
    workbook.close()

    imported = app_module.load_excel_transfer_log_rows(str(workbook_path))
    assert imported[-1]["source_log_no"] == 161
    assert imported[-1]["system_transfer_id"] == 188


def test_transfer_rebuild_numbers_null_source_logs_deterministically(app_module, tmp_path):
    workbook_path = tmp_path / "rebuilt-numbered-transfer-log.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Transfer log"]
    populate_legacy_transfer_rows(sheet, 175)
    records = [
        {**existing_transfer_record(188), "source_log_no": 160},
        existing_transfer_record(189),
        existing_transfer_record(190),
    ]

    app_module.rebuild_transfer_log_records_in_excel_sheet(sheet, records)
    first_numbers = [sheet.cell(row=row, column=3).value for row in range(2, 5)]
    first_ids = [sheet.cell(row=row, column=16).value for row in range(2, 5)]
    app_module.rebuild_transfer_log_records_in_excel_sheet(sheet, records)
    second_numbers = [sheet.cell(row=row, column=3).value for row in range(2, 5)]
    second_ids = [sheet.cell(row=row, column=16).value for row in range(2, 5)]

    assert first_numbers == second_numbers == [160, 161, 162]
    assert first_ids == second_ids == [188, 189, 190]
    assert records[1]["source_log_no"] is None
    assert records[2]["source_log_no"] is None
    workbook.close()


def test_transfer_export_normalizes_only_low_cost_compatibility_label(app_module, tmp_path):
    workbook_path = tmp_path / "transfer-type-labels.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    persisted_low_cost = {
        **existing_transfer_record(188),
        "source_asset_type": "Low-cost item",
    }
    derived_low_cost = {
        **existing_transfer_record(189),
        "source_asset_type": app_module.get_asset_usage_type_label("low_cost"),
    }
    standard = {
        **existing_transfer_record(190),
        "source_asset_type": "Standard asset",
    }

    app_module.write_transfer_log_records_to_excel_sheet(
        workbook["Transfer log"],
        [persisted_low_cost, derived_low_cost, standard],
    )

    sheet = workbook["Transfer log"]
    assert [sheet.cell(row=row, column=5).value for row in range(2, 5)] == [
        "Low-cost",
        "Low-cost",
        "Standard asset",
    ]
    assert [sheet.cell(row=row, column=16).value for row in range(2, 5)] == [188, 189, 190]
    workbook.close()


def test_database_transfer_export_record_includes_exact_transfer_id(app_module, monkeypatch):
    transfer = {
        "transfer_id": 188,
        "asset_id": 1232,
        "transfer_date": "2026-08-29",
        "from_holder_name": "Warehouse",
        "to_holder_name": "Synthetic Tenant #2 Admin",
    }
    monkeypatch.setattr(app_module, "list_asset_transfer_records", lambda request=None: [transfer])
    monkeypatch.setattr(
        app_module,
        "list_asset_records",
        lambda request=None: [{"asset_id": 1232, "asset_tag_number": "SYN-T2-0001"}],
    )
    monkeypatch.setattr(app_module, "list_people", lambda request=None: [])
    monkeypatch.setattr(app_module, "list_projects", lambda request=None: [])
    monkeypatch.setattr(app_module, "list_donors", lambda request=None: [])
    monkeypatch.setattr(app_module, "list_current_assignment_records", lambda request=None: [])
    monkeypatch.setattr(app_module, "list_asset_project_records", lambda request=None: [])
    monkeypatch.setattr(app_module, "list_asset_payment_records", lambda request=None: [])
    captured_request = SimpleNamespace(session={"tenant_id": "tenant-two"})
    project_requests = []
    monkeypatch.setattr(
        app_module,
        "get_asset_transfer_project_rows",
        lambda transfer_ids, request=None: project_requests.append(request) or {},
    )

    records = app_module.build_database_transfer_log_records(request=captured_request)

    assert len(records) == 1
    assert records[0]["system_transfer_id"] == 188
    assert project_requests == [captured_request]


def test_database_transfer_export_does_not_generate_pseudo_registration_rows(
    app_module,
    monkeypatch,
):
    monkeypatch.setattr(app_module, "list_asset_transfer_records", lambda request=None: [])
    monkeypatch.setattr(
        app_module,
        "list_asset_records",
        lambda request=None: [
            {
                "asset_id": 1232,
                "asset_tag_number": "SYN-T2-0001",
                "created_at": "2026-08-29T10:00:00+00:00",
            }
        ],
    )
    monkeypatch.setattr(app_module, "list_people", lambda request=None: [])
    monkeypatch.setattr(app_module, "list_projects", lambda request=None: [])
    monkeypatch.setattr(app_module, "list_donors", lambda request=None: [])
    monkeypatch.setattr(app_module, "list_current_assignment_records", lambda request=None: [])
    monkeypatch.setattr(app_module, "list_asset_project_records", lambda request=None: [])
    monkeypatch.setattr(app_module, "list_asset_payment_records", lambda request=None: [])
    monkeypatch.setattr(
        app_module,
        "get_asset_transfer_project_rows",
        lambda transfer_ids, request=None: {},
    )

    records = app_module.build_database_transfer_log_records(
        request=SimpleNamespace(session={"tenant_id": "tenant-two"})
    )

    assert records == []


def test_same_looking_transfers_with_distinct_system_ids_remain_distinct(
    app_module,
    tmp_path,
):
    workbook_path = tmp_path / "distinct.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    first = existing_transfer_record(188)
    second = existing_transfer_record(189)
    app_module.write_transfer_log_records_to_excel_sheet(
        workbook["Transfer log"],
        [first, second],
    )
    workbook.save(workbook_path)
    workbook.close()

    imported = app_module.load_excel_transfer_log_rows(str(workbook_path))
    context = transfer_context(
        app_module,
        {"transfer_id": 188, "asset_id": 1232, **first},
        {"transfer_id": 189, "asset_id": 1232, **second},
    )
    preview = app_module.build_transfer_log_preview(
        imported,
        {"SYN-T2-0001": {"asset_id": 1232}},
        context,
    )

    assert [row["system_transfer_id"] for row in imported] == [188, 189]
    assert preview["summary"]["new_records"] == 0
    assert preview["summary"]["skipped_existing"] == 2


@pytest.mark.parametrize("system_id", ["abc", "1.5", 0, -1, True])
def test_invalid_system_transfer_id_fails_closed(app_module, system_id):
    with pytest.raises(ValueError, match="invalid System Transfer ID"):
        app_module.normalize_excel_transfer_record(
            {
                "system_transfer_id": system_id,
                "asset_tag_number": "SYN-T2-0001",
                "transfer_date": "29.08.2026",
                "from_holder_name": "Warehouse",
            },
            2,
        )


def test_system_transfer_id_requires_complete_transfer_row(app_module):
    with pytest.raises(ValueError, match="must contain a valid asset tag and transfer date"):
        app_module.normalize_excel_transfer_record(
            {
                "system_transfer_id": 188,
                "asset_tag_number": "SYN-T2-0001",
            },
            2,
        )


def test_legacy_transfer_workbook_without_system_id_still_parses(app_module, tmp_path):
    workbook_path = tmp_path / "legacy.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Transfer log"]
    values = [1, "29.08.2026", "Standard", "SYN-T2-0001", "Monitor", None, "Warehouse"]
    for column_number, value in enumerate(values, start=3):
        sheet.cell(row=2, column=column_number).value = value
    workbook.save(workbook_path)
    workbook.close()

    imported = app_module.load_excel_transfer_log_rows(str(workbook_path))

    assert len(imported) == 1
    assert imported[0]["system_transfer_id"] is None
    assert imported[0]["asset_tag_number"] == "SYN-T2-0001"


def test_cross_tenant_system_transfer_id_fails_closed(app_module):
    record = existing_transfer_record(188)
    current_assets = {"SYN-T2-0001": {"asset_id": 1232}}

    with pytest.raises(app_module.TenantContextError, match="current tenant and asset"):
        app_module.build_transfer_log_preview(
            [record],
            current_assets,
            transfer_context(app_module),
        )


def test_same_tenant_wrong_asset_system_transfer_id_fails_closed(app_module):
    record = existing_transfer_record(188)
    current_assets = {"SYN-T2-0001": {"asset_id": 1232}}

    with pytest.raises(app_module.TenantContextError, match="current tenant and asset"):
        app_module.build_transfer_log_preview(
            [record],
            current_assets,
            transfer_context(
                app_module,
                {"transfer_id": 188, "asset_id": 9999, **record},
            ),
        )


def test_same_looking_other_tenant_transfer_does_not_suppress_active_tenant(app_module):
    record = {**existing_transfer_record(), "system_transfer_id": None}
    other_tenant_signature = app_module.make_transfer_signature(9999, record)
    context = transfer_context(app_module)
    context["other_tenant_signatures"] = {other_tenant_signature}

    preview = app_module.build_transfer_log_preview(
        [record],
        {"SYN-T2-0001": {"asset_id": 1232}},
        context,
    )

    assert preview["summary"]["new_records"] == 1
    assert preview["summary"]["skipped_existing"] == 0


class TransferInsertQuery:
    def __init__(self, database):
        self.database = database
        self.payload = None

    def insert(self, payload):
        self.payload = payload
        return self

    def execute(self):
        self.database.inserts.append(self.payload)
        transfer_id = self.database.returned_ids.pop(0) if self.database.returned_ids else None
        return SimpleNamespace(data=[{"transfer_id": transfer_id}] if transfer_id else [])


class TransferInsertSupabase:
    def __init__(self, returned_ids):
        self.returned_ids = list(returned_ids)
        self.inserts = []

    def table(self, table_name):
        assert table_name == "asset_transfers"
        return TransferInsertQuery(self)


def test_new_semantic_transfer_inserts_once_and_repeat_is_idempotent(
    app_module,
    monkeypatch,
):
    database = TransferInsertSupabase([190])
    monkeypatch.setattr(app_module, "supabase", database)
    monkeypatch.setattr(app_module, "ensure_parent_tenant", lambda *args, **kwargs: None)
    monkeypatch.setattr(app_module, "validate_transfer_person_tenants", lambda *args, **kwargs: None)
    monkeypatch.setattr(app_module, "apply_transfer_project_rows", lambda *args, **kwargs: 0)
    monkeypatch.setattr(app_module, "audit_log_event", lambda **kwargs: True)
    context = transfer_context(app_module)
    record = {
        **existing_transfer_record(),
        "system_transfer_id": None,
        "asset_id": 1232,
    }

    first = app_module.apply_sync_transfer(record, context, SimpleNamespace(session={}))
    second = app_module.apply_sync_transfer(record, context, SimpleNamespace(session={}))

    assert first == 1
    assert second == 0
    assert len(database.inserts) == 1


def test_new_asset_registration_creates_one_durable_tenant_transfer(
    app_module,
    monkeypatch,
):
    database = TransferInsertSupabase([501])
    monkeypatch.setattr(app_module, "supabase", database)
    monkeypatch.setattr(app_module, "ensure_parent_tenant", lambda *args, **kwargs: None)
    monkeypatch.setattr(app_module, "validate_transfer_person_tenants", lambda *args, **kwargs: None)
    monkeypatch.setattr(app_module, "apply_transfer_project_rows", lambda *args, **kwargs: 0)
    audit_calls = []
    monkeypatch.setattr(app_module, "audit_log_event", lambda **kwargs: audit_calls.append(kwargs) or True)
    request = SimpleNamespace(
        session={
            "admin_authenticated": True,
            "tenant_id": "00000000-0000-4000-8000-000000000002",
        }
    )
    context = transfer_context(app_module)
    context.update(
        {
            "assignment_by_asset_id": {},
            "projects_by_asset_id": {},
            "payments_by_asset_id": {},
            "registration_asset_ids": set(),
        }
    )
    asset = {
        "asset_id": 1232,
        "asset_tag_number": "SYN-T2-0001",
        "usage_type": "standard",
        "current_status": "functional",
        "created_at": "2026-08-29T10:00:00+00:00",
    }

    first = app_module.create_asset_registration_transfer(
        asset,
        context,
        request,
        audit_source="Asset creation",
        audit_actor="admin",
    )
    second = app_module.create_asset_registration_transfer(
        asset,
        context,
        request,
        audit_source="Asset creation",
        audit_actor="admin",
    )

    assert first == 1
    assert second == 0
    assert len(database.inserts) == 1
    assert database.inserts[0]["tenant_id"] == "00000000-0000-4000-8000-000000000002"
    assert database.inserts[0]["asset_id"] == 1232
    assert database.inserts[0]["transfer_reason"] == app_module.REGISTRATION_TRANSFER_REASON
    assert context["transfers_by_id"][501]["transfer_id"] == 501
    assert context["registration_asset_ids"] == {1232}
    assert audit_calls[0]["entity_id"] == 501
    assert audit_calls[0]["source"] == "Asset creation"


def test_registration_backfill_is_dry_run_tenant_scoped_and_idempotent(
    app_module,
    monkeypatch,
):
    tenant_id = "00000000-0000-4000-8000-000000000002"
    request = SimpleNamespace(
        session={
            "admin_authenticated": True,
            "tenant_id": tenant_id,
        }
    )
    assets = [
        {
            "asset_id": 1232,
            "asset_tag_number": "SYN-T2-0001",
            "usage_type": "standard",
            "current_status": "functional",
            "created_at": "2026-08-29T10:00:00+00:00",
        },
        {
            "asset_id": 1000,
            "asset_tag_number": "SYN-T2-OLD",
            "created_at": "2026-07-01T10:00:00+00:00",
        },
    ]
    context = {
        "assignment_by_asset_id": {},
        "projects_by_asset_id": {},
        "payments_by_asset_id": {},
        "registration_asset_ids": set(),
    }
    read_requests = []
    create_calls = []
    monkeypatch.setattr(
        app_module,
        "require_active_sync_tenant_id",
        lambda received_request: received_request.session["tenant_id"],
    )
    monkeypatch.setattr(
        app_module,
        "list_asset_records",
        lambda request=None: read_requests.append(request) or assets,
    )
    monkeypatch.setattr(app_module, "build_sync_context", lambda request=None: context)

    def fake_create(asset, sync_context, received_request, **kwargs):
        create_calls.append((asset["asset_id"], received_request, kwargs))
        sync_context["registration_asset_ids"].add(asset["asset_id"])
        return 1

    monkeypatch.setattr(app_module, "create_asset_registration_transfer", fake_create)

    dry_run = app_module.backfill_asset_registration_transfers(request)
    first_apply = app_module.backfill_asset_registration_transfers(request, apply=True)
    second_apply = app_module.backfill_asset_registration_transfers(request, apply=True)

    assert dry_run["tenant_id"] == tenant_id
    assert dry_run["dry_run"] is True
    assert dry_run["candidate_count"] == 1
    assert dry_run["inserted_count"] == 0
    assert dry_run["proposed_inserts"][0]["asset_tag_number"] == "SYN-T2-0001"
    assert first_apply["candidate_count"] == 1
    assert first_apply["inserted_count"] == 1
    assert second_apply["candidate_count"] == 0
    assert second_apply["inserted_count"] == 0
    assert [call[0] for call in create_calls] == [1232]
    assert all(received is request for received in read_requests)
    assert create_calls[0][1] is request


def test_apply_rechecks_system_transfer_id_before_insert(app_module, monkeypatch):
    database = TransferInsertSupabase([190])
    monkeypatch.setattr(app_module, "supabase", database)
    monkeypatch.setattr(app_module, "ensure_parent_tenant", lambda *args, **kwargs: None)
    monkeypatch.setattr(app_module, "validate_transfer_person_tenants", lambda *args, **kwargs: None)
    record = {**existing_transfer_record(), "asset_id": 1232}

    with pytest.raises(app_module.TenantContextError, match="current tenant and asset"):
        app_module.apply_sync_transfer(
            record,
            transfer_context(app_module),
            SimpleNamespace(session={}),
        )

    assert database.inserts == []


def test_empty_transfer_insert_response_is_controlled_failure(app_module, monkeypatch):
    database = TransferInsertSupabase([])
    monkeypatch.setattr(app_module, "supabase", database)
    monkeypatch.setattr(app_module, "ensure_parent_tenant", lambda *args, **kwargs: None)
    monkeypatch.setattr(app_module, "validate_transfer_person_tenants", lambda *args, **kwargs: None)
    monkeypatch.setattr(app_module, "apply_transfer_project_rows", lambda *args, **kwargs: 0)

    with pytest.raises(RuntimeError, match="did not return a valid transfer_id"):
        app_module.apply_sync_transfer(
            {
                **existing_transfer_record(),
                "system_transfer_id": None,
                "asset_id": 1232,
            },
            transfer_context(app_module),
            SimpleNamespace(session={}),
        )


def test_registration_transfer_semantic_fallback_is_idempotent(app_module):
    record = {
        "system_transfer_id": None,
        "asset_tag_number": "SYN-T2-0001",
        "transfer_date": "2026-08-01",
        "from_holder_name": "Supplier",
        "to_holder_name": "Warehouse",
        "transfer_reason": "Registration of the new asset",
    }
    existing = {"transfer_id": 187, "asset_id": 1232, **record}
    preview = app_module.build_transfer_log_preview(
        [record],
        {"SYN-T2-0001": {"asset_id": 1232}},
        transfer_context(app_module, existing),
    )

    assert preview["summary"]["new_records"] == 0
    assert preview["summary"]["skipped_existing"] == 1


def test_generated_export_rebuilds_populated_legacy_transfer_log(
    app_module,
    monkeypatch,
    tmp_path,
):
    source_path = tmp_path / "official_inventory.xlsx"
    export_path = tmp_path / "supabase_inventory_export.xlsx"
    workbook = Workbook()
    standard_sheet = workbook.active
    standard_sheet.title = "Standard Asset List Format"
    standard_sheet.cell(row=8, column=1).value = (
        "Asset Tag No. / Inventory Code\n(new standardised system)"
    )
    transfer_sheet = workbook.create_sheet("Transfer log")
    for column_number, header in enumerate(TRANSFER_HEADERS, start=3):
        transfer_sheet.cell(row=1, column=column_number).value = header
        transfer_sheet.cell(row=1, column=column_number).font = Font(bold=True)
    transfer_sheet.column_dimensions["A"].hidden = True
    transfer_sheet.column_dimensions["F"].width = 24
    transfer_sheet.freeze_panes = "C2"
    transfer_sheet.auto_filter.ref = "A1:O4"
    populate_legacy_transfer_rows(transfer_sheet, 3)
    workbook.save(source_path)
    workbook.close()
    source_hash = hashlib.sha256(source_path.read_bytes()).hexdigest()

    exported_record = existing_transfer_record(188)
    monkeypatch.setattr(app_module, "require_active_sync_tenant_id", lambda request: request.session["tenant_id"])
    monkeypatch.setattr(
        app_module,
        "ensure_sync_storage",
        lambda request: {"export": str(export_path)},
    )
    monkeypatch.setattr(app_module, "ensure_sync_workbook_template", lambda request: str(source_path))
    monkeypatch.setattr(app_module, "build_database_excel_records", lambda usage_type, request=None: [])
    monkeypatch.setattr(
        app_module,
        "build_database_transfer_log_records",
        lambda request=None: [exported_record],
    )

    result = app_module.export_supabase_to_excel(
        SimpleNamespace(session={"tenant_id": "00000000-0000-4000-8000-000000000002"})
    )

    assert hashlib.sha256(source_path.read_bytes()).hexdigest() == source_hash
    source = load_workbook(source_path)
    assert source["Transfer log"]["P1"].value is None
    assert source["Transfer log"]["F2"].value == "LEGACY-0001"
    source.close()

    exported = load_workbook(export_path)
    sheet = exported["Transfer log"]
    assert sheet["P1"].value == "System Transfer ID"
    assert sheet.column_dimensions["P"].hidden is True
    assert sheet.column_dimensions["A"].hidden is True
    assert sheet.column_dimensions["F"].width == 24
    assert sheet.freeze_panes == "C2"
    assert sheet.auto_filter.ref == "A1:O2"
    assert sheet["C1"].font.bold is True
    assert sheet["P2"].value == 188
    assert sheet["F2"].value == "SYN-T2-0001"
    assert all(sheet.cell(row=row, column=6).value is None for row in range(3, sheet.max_row + 1))
    exported.close()
    assert result["transfer_exported_records"] == 1


def test_generated_export_rebuilds_asset_sheets_from_tenant_database_only(
    app_module,
    monkeypatch,
    tmp_path,
):
    source_path = tmp_path / "official_inventory.xlsx"
    export_path = tmp_path / "supabase_inventory_export.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    create_asset_sheet(workbook, "Standard Asset List Format", "SYN-T2-0001")
    create_asset_sheet(workbook, "Low-cost-items", "SYN-T2-LC-0001")
    transfer_sheet = workbook.create_sheet("Transfer log")
    for column_number, header in enumerate(TRANSFER_HEADERS, start=3):
        transfer_sheet.cell(row=1, column=column_number).value = header
    workbook.save(source_path)
    workbook.close()
    source_hash = hashlib.sha256(source_path.read_bytes()).hexdigest()

    request = SimpleNamespace(
        session={
            "admin_authenticated": True,
            "tenant_id": "00000000-0000-4000-8000-000000000001",
        }
    )
    standard_record = {
        "asset_tag_number": "HELP-UKR-0001",
        "usage_type": "standard",
        "item_description": "Tenant one monitor",
        "model": "T1-MONITOR",
        "current_status": "functional",
    }
    low_cost_record = {
        "asset_tag_number": "HELP-UKR-LC-0001",
        "usage_type": "low_cost",
        "item_description": "Tenant one mouse",
        "model": "T1-MOUSE",
        "current_status": "functional",
    }
    registration_transfer = {
        "system_transfer_id": 500,
        "transfer_date": "2026-08-30",
        "source_asset_type": "Standard",
        "asset_tag_number": "HELP-UKR-0001",
        "from_holder_name": "Supplier",
        "to_holder_name": "Warehouse",
        "asset_status": "functional",
        "asset_condition_description": "New",
        "transfer_reason": "Registration of the new asset",
    }
    monkeypatch.setattr(app_module, "require_active_sync_tenant_id", lambda received: received.session["tenant_id"])
    monkeypatch.setattr(app_module, "ensure_sync_storage", lambda received: {"export": str(export_path)})
    monkeypatch.setattr(app_module, "ensure_sync_workbook_template", lambda received: str(source_path))
    monkeypatch.setattr(
        app_module,
        "build_database_excel_records",
        lambda usage_type, request=None: [standard_record] if usage_type == "standard" else [low_cost_record],
    )
    monkeypatch.setattr(
        app_module,
        "build_database_transfer_log_records",
        lambda request=None: [registration_transfer],
    )

    first_result = app_module.export_supabase_to_excel(request)
    first_bytes = export_path.read_bytes()
    second_result = app_module.export_supabase_to_excel(request)

    assert hashlib.sha256(source_path.read_bytes()).hexdigest() == source_hash
    assert first_bytes == export_path.read_bytes()
    assert first_result["exported_records"] == second_result["exported_records"] == 3

    source = load_workbook(source_path, data_only=False)
    assert source["Standard Asset List Format"]["A9"].value == "SYN-T2-0001"
    assert source["Low-cost-items"]["A9"].value == "SYN-T2-LC-0001"
    source.close()

    exported = load_workbook(export_path, data_only=False)
    for sheet_name, expected_tag in [
        ("Standard Asset List Format", "HELP-UKR-0001"),
        ("Low-cost-items", "HELP-UKR-LC-0001"),
    ]:
        sheet = exported[sheet_name]
        assert sheet["A9"].value == expected_tag
        assert sheet.max_row == 9
        assert sheet.column_dimensions["A"].width == 31
        assert sheet.column_dimensions["B"].hidden is True
        assert sheet.freeze_panes == "A9"
        assert sheet["Y9"].data_type == "f"
    exported.close()

    parsed_assets = app_module.load_excel_sync_rows(str(export_path))
    parsed_transfers = app_module.load_excel_transfer_log_rows(str(export_path))
    assert {row["asset_tag_number"] for row in parsed_assets} == {
        "HELP-UKR-0001",
        "HELP-UKR-LC-0001",
    }
    assert all(not row["asset_tag_number"].startswith("SYN-T2") for row in parsed_assets)
    assert [row["system_transfer_id"] for row in parsed_transfers] == [500]

    context = {
        "person_lookup": {},
        "location_lookup": {},
        "project_lookup": {},
        "donor_lookup": {},
        "assignment_by_asset_id": {},
        "projects_by_asset_id": {},
        "payments_by_asset_id": {},
        "transfers_by_id": {500: {"transfer_id": 500, "asset_id": 1}},
        "transfer_signatures": set(),
        "registration_asset_ids": {1},
        "supports_asset_project_purchase_origin": False,
    }
    monkeypatch.setattr(app_module, "build_sync_context", lambda request: context)
    synced_fields = [
        "usage_type",
        "asset_classification",
        "asset_sub_classification",
        "item_description",
        "brand_make",
        "model",
        "serial_chassis_number",
        "quantity",
        "purchase_price",
        "currency",
        "current_status",
    ]
    current_assets = [
        {
            "asset_id": index,
            "asset_tag_number": row["asset_tag_number"],
            **{field_name: row.get(field_name) for field_name in synced_fields},
        }
        for index, row in enumerate(parsed_assets, start=1)
    ]
    preview = app_module.build_sync_preview(
        parsed_assets,
        current_assets,
        parsed_transfers,
        request=request,
    )
    assert preview["summary"]["new_records"] == 0
    assert preview["summary"]["changed_records"] == 0
    assert preview["transfer_log"]["summary"]["new_records"] == 0
    assert preview["transfer_log"]["summary"]["skipped_existing"] == 1


def test_rebuild_175_legacy_rows_keeps_only_authoritative_history(app_module, tmp_path):
    workbook_path = tmp_path / "legacy-175.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Transfer log"]
    populate_legacy_transfer_rows(sheet, 175)
    sheet["G2"] = "=1+1"
    sheet["G2"].comment = Comment("legacy", "legacy")
    sheet["G2"].hyperlink = "https://legacy.invalid"
    validation = DataValidation(type="list", formula1='"Legacy"')
    sheet.add_data_validation(validation)
    validation.add("M2:M176")
    sheet.conditional_formatting.add(
        "M2:M176",
        CellIsRule(operator="equal", formula=['"functional"']),
    )
    records = [existing_transfer_record(188), existing_transfer_record(189)]

    result = app_module.rebuild_transfer_log_records_in_excel_sheet(sheet, records)
    workbook.save(workbook_path)
    workbook.close()

    rebuilt = load_workbook(workbook_path, data_only=False)
    sheet = rebuilt["Transfer log"]
    parsed = app_module.load_excel_transfer_log_rows(str(workbook_path))
    assert result["exported_records"] == 2
    assert [row["system_transfer_id"] for row in parsed] == [188, 189]
    assert [row["asset_tag_number"] for row in parsed] == ["SYN-T2-0001", "SYN-T2-0001"]
    assert all("LEGACY-" not in str(cell.value or "") for row in sheet.iter_rows() for cell in row)
    assert not any(cell.comment or cell.hyperlink for row in sheet.iter_rows() for cell in row)
    assert len(sheet.data_validations.dataValidation) == 0
    assert len(sheet.conditional_formatting) == 0
    rebuilt.close()


def test_rebuilt_export_reimports_as_existing_with_fresh_context(app_module, tmp_path):
    workbook_path = tmp_path / "rebuilt-roundtrip.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Transfer log"]
    populate_legacy_transfer_rows(sheet, 2)
    record = existing_transfer_record(188)
    app_module.rebuild_transfer_log_records_in_excel_sheet(sheet, [record])
    workbook.save(workbook_path)
    workbook.close()

    imported = app_module.load_excel_transfer_log_rows(str(workbook_path))
    fresh_context = transfer_context(
        app_module,
        {"transfer_id": 188, "asset_id": 1232, **record},
    )
    preview = app_module.build_transfer_log_preview(
        imported,
        {"SYN-T2-0001": {"asset_id": 1232}},
        fresh_context,
    )

    assert preview["summary"]["new_records"] == 0
    assert preview["summary"]["skipped_existing"] == 1


def test_repeat_rebuild_keeps_row_count_and_one_system_id_column(app_module, tmp_path):
    workbook_path = tmp_path / "repeat-export.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Transfer log"]
    populate_legacy_transfer_rows(sheet, 4)
    records = [existing_transfer_record(188), existing_transfer_record(189)]

    app_module.rebuild_transfer_log_records_in_excel_sheet(sheet, records)
    app_module.rebuild_transfer_log_records_in_excel_sheet(sheet, records)
    workbook.save(workbook_path)
    workbook.close()

    rebuilt = load_workbook(workbook_path)
    sheet = rebuilt["Transfer log"]
    parsed = app_module.load_excel_transfer_log_rows(str(workbook_path))
    assert len(parsed) == 2
    assert [row["system_transfer_id"] for row in parsed] == [188, 189]
    assert sum(cell.value == "System Transfer ID" for cell in sheet[1]) == 1
    rebuilt.close()


def test_registration_transfer_roundtrips_through_rebuild_and_parser(app_module, tmp_path):
    workbook_path = tmp_path / "registration-roundtrip.xlsx"
    create_transfer_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    record = {
        "system_transfer_id": 187,
        "source_log_no": None,
        "asset_tag_number": "SYN-T2-0001",
        "transfer_date": "2026-08-01",
        "from_holder_name": "Supplier",
        "to_holder_name": "Warehouse",
        "transfer_reason": "Registration of the new asset",
    }
    app_module.rebuild_transfer_log_records_in_excel_sheet(
        workbook["Transfer log"],
        [record],
    )
    workbook.save(workbook_path)
    workbook.close()

    imported = app_module.load_excel_transfer_log_rows(str(workbook_path))
    preview = app_module.build_transfer_log_preview(
        imported,
        {"SYN-T2-0001": {"asset_id": 1232}},
        transfer_context(
            app_module,
            {"transfer_id": 187, "asset_id": 1232, **record},
        ),
    )

    assert imported[0]["transfer_key"] == "registration:SYN-T2-0001"
    assert preview["summary"]["new_records"] == 0
    assert preview["summary"]["skipped_existing"] == 1

