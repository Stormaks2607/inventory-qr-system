from typing import Optional
import base64
import csv
import html
import io
from copy import copy
import secrets
import json
import os
import re
from datetime import datetime
from urllib.parse import urlencode, urlparse
from zoneinfo import ZoneInfo

import httpx
import requests
from dotenv import load_dotenv
from fastapi import Body, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from postgrest.exceptions import APIError
from starlette.middleware.sessions import SessionMiddleware
from fastapi.templating import Jinja2Templates
from itsdangerous import BadSignature, URLSafeSerializer
from supabase import Client, create_client


load_dotenv()

def clean_env_value(name: str) -> Optional[str]:
    value = os.getenv(name)
    if value is None:
        return None
    return value.strip().strip("\"'")


def normalize_audit_limit(value, default: int = 5) -> int:
    try:
        limit = int(value)
    except (TypeError, ValueError):
        return default
    return limit if limit in {5, 10, 20, 50} else default


SUPABASE_URL = clean_env_value("SUPABASE_URL")
SUPABASE_KEY = clean_env_value("SUPABASE_KEY")
BOT_TOKEN = clean_env_value("BOT_TOKEN")
PUBLIC_BASE_URL = "https://inventory-qr-system.onrender.com"
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "change-me")
ADMIN_SESSION_SECRET = os.getenv("ADMIN_SESSION_SECRET", "replace-this-session-secret")
BRANDING_SETTINGS_PATH = os.path.join("private_docs", "company_branding.json")
BRANDING_UPLOAD_DIR = os.path.join("private_docs", "branding")
ALLOWED_LOGO_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
BRANDING_SUPABASE_TABLE = "organization_branding"
DEFAULT_BRANDING_TENANT_KEY = "default"
SYNC_STORAGE_DIR = os.path.join("private_docs", "sync")
SYNC_WORKBOOK_PATH = os.path.join(SYNC_STORAGE_DIR, "official_inventory.xlsx")
SYNC_EXPORT_PATH = os.path.join(SYNC_STORAGE_DIR, "supabase_inventory_export.xlsx")
SYNC_STATE_PATH = os.path.join(SYNC_STORAGE_DIR, "sync_state.json")
EXCEL_SYNC_SHEET_NAME = "Standard Asset List Format"
EXCEL_LOW_COST_SHEET_NAME = "Low-cost-items"
EXCEL_TRANSFER_LOG_SHEET_NAME = "Transfer log"
EXCEL_SYNC_HEADER_ROW = 7
EXCEL_TRANSFER_LOG_HEADER_ROW = 0
PROJECT_NUMBER_PATTERN = r"\b[A-Z]{2,5}-\d{1,5}\b"
EXCEL_SYNC_COLUMN_MAP = {
    "Asset Tag No. / Inventory Code\n(new standardised system)": "asset_tag_number",
    "Previous inventory code\n(if applicable)": "inventory_code_old",
    "Asset Classification": "asset_classification",
    "Asset Sub Classification": "asset_sub_classification",
    "Item Description": "item_description",
    "Brand / Make ": "brand_make",
    "Model": "model",
    "Serial/ Chassis No.": "serial_number",
    "Quantity": "quantity",
    "Location": "location_name",
    "Department ": "department_name",
    "Name of Recipient": "recipient_name",
    "Position of Recipient": "recipient_position",
    "Date (Year) of Purchase": "purchase_date_raw",
    "Purchase price": "purchase_price",
    "Currency": "currency",
    "Purchased to Project No.": "purchased_project_no",
    "Donor ": "purchased_donor_name",
    "Donor": "transferred_donor_name",
    "Donor .1": "transferred_donor_name",
    "Transferred to Project No.": "transferred_project_no",
    "Current Status\n(functionality)": "current_status",
    "Remarks": "remarks",
    "Last date of transfer": "last_transfer_date",
}
ASSET_STATUS_OPTIONS = [
    ("functional", "Функціонуючий / Functional"),
    ("non-functional", "Не функціонуючий / Non-functional"),
    ("lost", "Втрачений / Lost"),
    ("disposed", "Списаний / Disposed"),
]

app = FastAPI(title="Asset API", version="1.0.0")
app.add_middleware(SessionMiddleware, secret_key=ADMIN_SESSION_SECRET)

if not SUPABASE_URL or not SUPABASE_KEY:
    raise RuntimeError("SUPABASE_URL and SUPABASE_KEY must be set.")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
templates = Jinja2Templates(directory="templates")

ASSET_STATUS_SELECT_OPTIONS = [
    ("functional", "Functional"),
    ("non-functional", "Non-functional"),
    ("lost", "Lost"),
    ("disposed", "Disposed"),
]

ASSET_USAGE_TYPE_OPTIONS = [
    ("standard", "Standard asset"),
    ("low_cost", "Low-cost item"),
]

ASSET_USAGE_TYPE_LABELS = dict(ASSET_USAGE_TYPE_OPTIONS)


def infer_asset_usage_type(asset_tag_number: Optional[str]) -> str:
    normalized = normalize_asset_tag(asset_tag_number or "")
    return "low_cost" if "-LC-" in normalized else "standard"


def normalize_asset_usage_type(value: Optional[str], asset_tag_number: Optional[str] = None) -> str:
    normalized = (value or "").strip().lower()
    if normalized in ASSET_USAGE_TYPE_LABELS:
        return normalized
    return infer_asset_usage_type(asset_tag_number)


def get_asset_usage_type_label(value: Optional[str]) -> str:
    return ASSET_USAGE_TYPE_LABELS.get(normalize_asset_usage_type(value), "Standard asset")


class DatabaseConnectionError(RuntimeError):
    pass


def get_supabase_host() -> str:
    if not SUPABASE_URL:
        return "missing SUPABASE_URL"
    parsed = urlparse(SUPABASE_URL)
    return parsed.netloc or parsed.path or SUPABASE_URL


def execute_supabase_query(query, context: str):
    try:
        return query.execute()
    except httpx.ConnectError as exc:
        host = get_supabase_host()
        message = (
            f"Cannot connect to Supabase host '{host}'. "
            "Check SUPABASE_URL in Render environment variables."
        )
        print(f"SUPABASE CONNECTION ERROR ({context}): {message} Original error: {exc}")
        raise DatabaseConnectionError(message) from exc


def get_admin_actor(request: Optional[Request] = None, fallback: str = "Admin") -> str:
    if request is not None:
        username = request.session.get("admin_username")
        if username:
            return str(username)
    return fallback


def stringify_audit_value(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def audit_log_event(
    *,
    entity_type: str,
    action: str,
    entity_id: Optional[int] = None,
    entity_label: Optional[str] = None,
    field_name: Optional[str] = None,
    old_value=None,
    new_value=None,
    summary: Optional[str] = None,
    source: str = "Admin",
    actor: Optional[str] = None,
    request: Optional[Request] = None,
    event_date: Optional[str] = None,
    event_key: Optional[str] = None,
) -> bool:
    payload = {
        "actor": actor or get_admin_actor(request, fallback=source),
        "source": source,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "entity_label": entity_label,
        "action": action,
        "field_name": field_name,
        "old_value": stringify_audit_value(old_value),
        "new_value": stringify_audit_value(new_value),
        "summary": summary,
        "event_date": event_date or datetime.now(ZoneInfo("Europe/Kyiv")).isoformat(),
        "event_key": event_key,
    }
    try:
        supabase.table("audit_log").insert(payload).execute()
        return True
    except Exception as error:
        if isinstance(error, APIError):
            combined = " ".join(part for part in [error.message or "", error.details or ""] if part).lower()
            if "event_key" in combined and ("duplicate" in combined or "unique" in combined):
                return False
            if "event_date" in combined or "event_key" in combined:
                payload.pop("event_date", None)
                payload.pop("event_key", None)
                try:
                    supabase.table("audit_log").insert(payload).execute()
                    return True
                except Exception:
                    return False
        return False


def update_audit_log_event_by_key(event_key: Optional[str], **updates) -> bool:
    if not event_key:
        return False
    payload = {
        key: value
        for key, value in updates.items()
        if value is not None
    }
    if not payload:
        return False
    for field_name in ["old_value", "new_value"]:
        if field_name in payload:
            payload[field_name] = stringify_audit_value(payload[field_name])
    try:
        supabase.table("audit_log").update(payload).eq("event_key", event_key).execute()
        return True
    except Exception:
        return False


def audit_log_field_changes(
    *,
    entity_type: str,
    entity_id: Optional[int],
    entity_label: Optional[str],
    old_record: dict,
    new_record: dict,
    fields: list[str],
    source: str = "Admin",
    request: Optional[Request] = None,
    action: str = "updated",
) -> int:
    logged = 0
    for field_name in fields:
        old_value = old_record.get(field_name)
        new_value = new_record.get(field_name)
        if stringify_audit_value(old_value) == stringify_audit_value(new_value):
            continue
        audit_log_event(
            entity_type=entity_type,
            entity_id=entity_id,
            entity_label=entity_label,
            action=action,
            field_name=field_name,
            old_value=old_value,
            new_value=new_value,
            summary=f"{field_name}: {stringify_audit_value(old_value) or '-'} -> {stringify_audit_value(new_value) or '-'}",
            source=source,
            request=request,
        )
        logged += 1
    return logged


def format_audit_datetime(value) -> str:
    if not value:
        return "-"
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        parsed = parsed.astimezone(ZoneInfo("Europe/Kyiv"))
        return parsed.strftime("%d.%m.%Y %H:%M")
    except Exception:
        return str(value)


def list_recent_audit_events(limit: int = 5) -> list[dict]:
    limit = normalize_audit_limit(limit)
    try:
        response = (
            supabase.table("audit_log")
            .select("*")
            .order("event_date", desc=True)
            .order("created_at", desc=True)
            .limit(limit)
            .execute()
        )
    except Exception as error:
        if isinstance(error, APIError):
            combined = " ".join(part for part in [error.message or "", error.details or ""] if part).lower()
            if "event_date" in combined:
                try:
                    response = (
                        supabase.table("audit_log")
                        .select("*")
                        .order("created_at", desc=True)
                        .limit(limit)
                        .execute()
                    )
                except Exception:
                    return []
            else:
                return []
        else:
            return []

    return enrich_audit_rows(response.data or [])


def normalize_audit_page_size(value, default: int = 100) -> int:
    try:
        page_size = int(value)
    except (TypeError, ValueError):
        return default
    return page_size if page_size in {50, 100, 200} else default


def enrich_audit_rows(rows: list[dict]) -> list[dict]:
    for row in rows:
        row["created_at_display"] = format_audit_datetime(row.get("event_date") or row.get("created_at"))
        row["display_summary"] = row.get("summary") or (
            f"{row.get('field_name')}: {row.get('old_value') or '-'} -> {row.get('new_value') or '-'}"
            if row.get("field_name")
            else row.get("action")
        )
    return rows


def audit_row_matches_query(row: dict, query: str) -> bool:
    if not query:
        return True
    normalized_query = query.strip().casefold()
    haystack = " ".join(
        str(row.get(field_name) or "")
        for field_name in [
            "entity_type",
            "entity_label",
            "action",
            "field_name",
            "old_value",
            "new_value",
            "summary",
            "actor",
            "source",
        ]
    ).casefold()
    return normalized_query in haystack


def list_audit_log_events(
    *,
    q: str = "",
    entity_type: str = "",
    source: str = "",
    page: int = 1,
    page_size: int = 100,
) -> dict:
    page = max(page, 1)
    page_size = normalize_audit_page_size(page_size)
    fetch_limit = max(page * page_size + 1, page_size + 1)

    try:
        query = supabase.table("audit_log").select("*")
        if entity_type:
            query = query.eq("entity_type", entity_type)
        if source:
            query = query.eq("source", source)
        response = (
            query
            .order("event_date", desc=True)
            .order("created_at", desc=True)
            .limit(fetch_limit)
            .execute()
        )
    except Exception as error:
        if isinstance(error, APIError):
            combined = " ".join(part for part in [error.message or "", error.details or ""] if part).lower()
            if "event_date" in combined:
                try:
                    query = supabase.table("audit_log").select("*")
                    if entity_type:
                        query = query.eq("entity_type", entity_type)
                    if source:
                        query = query.eq("source", source)
                    response = query.order("created_at", desc=True).limit(fetch_limit).execute()
                except Exception:
                    return {"rows": [], "page": page, "page_size": page_size, "has_next": False, "total_loaded": 0}
            else:
                return {"rows": [], "page": page, "page_size": page_size, "has_next": False, "total_loaded": 0}
        else:
            return {"rows": [], "page": page, "page_size": page_size, "has_next": False, "total_loaded": 0}

    all_rows = [row for row in (response.data or []) if audit_row_matches_query(row, q)]
    start = (page - 1) * page_size
    end = start + page_size
    page_rows = enrich_audit_rows(all_rows[start:end])
    has_next = len(all_rows) > end

    return {
        "rows": page_rows,
        "page": page,
        "page_size": page_size,
        "has_next": has_next,
        "total_loaded": len(all_rows),
    }


def list_audit_filter_values(field_name: str) -> list[str]:
    try:
        response = (
            supabase.table("audit_log")
            .select(field_name)
            .limit(1000)
            .execute()
        )
    except Exception:
        return []
    values = {
        str(row.get(field_name))
        for row in response.data or []
        if row.get(field_name)
    }
    return sorted(values)


def build_transfer_audit_summary(transfer: dict, project_rows: Optional[list[dict]] = None) -> str:
    asset_label = transfer.get("asset_tag_snapshot") or transfer.get("asset_tag_number") or f"Asset #{transfer.get('asset_id')}"
    from_holder = transfer.get("from_holder_display") or transfer.get("from_holder_name") or "-"
    to_holder = transfer.get("to_holder_display") or transfer.get("to_holder_name") or "-"
    from_project = transfer.get("from_project_display") or transfer.get("from_project_raw") or "-"
    to_project = transfer.get("to_project_display") or transfer.get("to_project_raw") or "-"
    if project_rows is not None:
        from_project = format_transfer_project_rows(project_rows, "from")
        to_project = format_transfer_project_rows(project_rows, "to")

    movement_changed = (
        normalize_sync_match_key(from_holder) != normalize_sync_match_key(to_holder)
        or normalize_sync_match_key(from_project) != normalize_sync_match_key(to_project)
    )
    if movement_changed:
        return f"{asset_label} transferred: {from_holder} / {from_project} -> {to_holder} / {to_project}"

    details = []
    if transfer.get("asset_status"):
        details.append(f"Status: {transfer.get('asset_status')}")
    if transfer.get("asset_condition_description"):
        details.append(f"Condition: {transfer.get('asset_condition_description')}")
    if transfer.get("transfer_reason") or transfer.get("notes"):
        details.append(f"Reason: {transfer.get('transfer_reason') or transfer.get('notes')}")
    if details:
        return f"{asset_label} transfer log update: " + " | ".join(details)
    return f"{asset_label} transfer log update: holder/project unchanged"


def backfill_audit_from_transfer_log() -> dict:
    transfers = list_asset_transfer_records()
    if not transfers:
        return {"created": 0, "available": 0}

    assets_by_id = {row.get("asset_id"): row for row in list_asset_records()}
    people_by_id = {row.get("person_id"): row for row in list_people()}
    transfer_ids = [row.get("transfer_id") for row in transfers if row.get("transfer_id")]
    projects_by_transfer_id = get_asset_transfer_project_rows(transfer_ids)
    created = 0
    updated = 0

    for transfer in sorted(transfers, key=lambda row: (str(row.get("transfer_date") or ""), int(row.get("transfer_id") or 0))):
        transfer_id = transfer.get("transfer_id")
        asset = assets_by_id.get(transfer.get("asset_id")) or {}
        from_person = people_by_id.get(transfer.get("from_person_id")) or {}
        to_person = people_by_id.get(transfer.get("to_person_id")) or {}
        transfer = {
            **transfer,
            "asset_tag_number": asset.get("asset_tag_number"),
            "from_holder_display": get_person_display_name(from_person) if from_person else transfer.get("from_holder_name"),
            "to_holder_display": get_person_display_name(to_person) if to_person else transfer.get("to_holder_name"),
        }
        project_rows = projects_by_transfer_id.get(transfer_id, [])
        summary = build_transfer_audit_summary(transfer, project_rows)
        event_key = f"asset_transfer:{transfer_id}" if transfer_id else None
        was_created = audit_log_event(
            entity_type="Transfer",
            entity_id=transfer_id,
            entity_label=transfer.get("asset_tag_snapshot") or transfer.get("asset_tag_number"),
            action="transferred",
            field_name="assignment",
            old_value=transfer.get("from_holder_display") or transfer.get("from_holder_name"),
            new_value=transfer.get("to_holder_display") or transfer.get("to_holder_name"),
            summary=summary,
            source="Excel Transfer log",
            actor="Excel Transfer log",
            event_date=transfer.get("transfer_date"),
            event_key=event_key,
        )
        if was_created:
            created += 1
        else:
            was_updated = update_audit_log_event_by_key(
                event_key,
                entity_label=transfer.get("asset_tag_snapshot") or transfer.get("asset_tag_number"),
                old_value=transfer.get("from_holder_display") or transfer.get("from_holder_name"),
                new_value=transfer.get("to_holder_display") or transfer.get("to_holder_name"),
                summary=summary,
                event_date=transfer.get("transfer_date"),
            )
            if was_updated:
                updated += 1

    return {"created": created, "updated": updated, "available": len(transfers)}


def get_default_branding_settings() -> dict:
    return {
        "company_name": "Your Company",
        "report_title": "Asset Assignment Statement",
        "report_subtitle": "Official summary of all assets currently assigned to the employee at the time of printing.",
        "report_theme": "classic",
        "primary_color": "#0f6c5c",
        "accent_color": "#f4fbf8",
        "footer_note": "Internal use only",
        "issuer_label": "Issued by",
        "issuer_signature_label": "Name, role, and signature",
        "receiver_label": "Received by employee",
        "receiver_signature_label": "Employee signature",
        "logo_path": "",
    }


def normalize_branding_settings(settings: dict) -> dict:
    defaults = get_default_branding_settings()
    normalized = defaults.copy()
    normalized.update({key: value for key, value in (settings or {}).items() if key in defaults and value is not None})
    if normalized.get("report_theme") not in {"classic", "corporate", "compact", "help_standard"}:
        normalized["report_theme"] = defaults["report_theme"]
    return normalized


def sanitize_tenant_key(value: str) -> str:
    normalized = "".join(
        character.lower() if character.isalnum() else "-"
        for character in (value or "").strip()
    )
    collapsed = "-".join(part for part in normalized.split("-") if part)
    return collapsed or DEFAULT_BRANDING_TENANT_KEY


def get_current_tenant_key(request: Request) -> str:
    return sanitize_tenant_key(request.session.get("admin_tenant_key") or DEFAULT_BRANDING_TENANT_KEY)


def get_legacy_tenant_key(request: Request) -> str:
    return sanitize_tenant_key(request.session.get("admin_username") or DEFAULT_BRANDING_TENANT_KEY)


def get_branding_settings_path(tenant_key: str) -> str:
    safe_tenant_key = sanitize_tenant_key(tenant_key)
    return os.path.join("private_docs", f"company_branding_{safe_tenant_key}.json")


def get_branding_upload_dir(tenant_key: str) -> str:
    safe_tenant_key = sanitize_tenant_key(tenant_key)
    return os.path.join(BRANDING_UPLOAD_DIR, safe_tenant_key)


def ensure_branding_storage() -> None:
    os.makedirs(os.path.dirname(BRANDING_SETTINGS_PATH), exist_ok=True)
    os.makedirs(BRANDING_UPLOAD_DIR, exist_ok=True)


def load_branding_settings_from_supabase(tenant_key: str) -> Optional[dict]:
    try:
        response = (
            supabase.table(BRANDING_SUPABASE_TABLE)
            .select("*")
            .eq("tenant_key", tenant_key)
            .limit(1)
            .execute()
        )
    except Exception:
        return None

    if not response.data:
        return None

    return normalize_branding_settings(response.data[0])


def load_branding_settings_from_file(tenant_key: str) -> dict:
    ensure_branding_storage()
    settings_path = get_branding_settings_path(tenant_key)
    settings = get_default_branding_settings()
    if os.path.exists(settings_path):
        with open(settings_path, "r", encoding="utf-8") as file:
            saved = json.load(file)
        settings.update({key: value for key, value in saved.items() if key in settings})
    elif tenant_key == "default" and os.path.exists(BRANDING_SETTINGS_PATH):
        with open(BRANDING_SETTINGS_PATH, "r", encoding="utf-8") as file:
            saved = json.load(file)
        settings.update({key: value for key, value in saved.items() if key in settings})
    return normalize_branding_settings(settings)


def load_branding_settings(tenant_key: str) -> tuple[dict, str]:
    settings = load_branding_settings_from_supabase(tenant_key)
    if settings is not None:
        return settings, "supabase"
    return load_branding_settings_from_file(tenant_key), "local"


def branding_matches_defaults(settings: dict) -> bool:
    defaults = get_default_branding_settings()
    return all(settings.get(key) == value for key, value in defaults.items())


def resolve_branding_for_request(request: Request) -> tuple[str, dict, str]:
    tenant_key = get_current_tenant_key(request)
    branding, branding_storage = load_branding_settings(tenant_key)
    if not branding_matches_defaults(branding):
        if branding_storage == "local" and save_branding_settings_to_supabase(tenant_key, branding):
            branding_storage = "supabase"
        return tenant_key, branding, branding_storage

    legacy_tenant_key = get_legacy_tenant_key(request)
    if legacy_tenant_key != tenant_key:
        legacy_branding, legacy_storage = load_branding_settings(legacy_tenant_key)
        if not branding_matches_defaults(legacy_branding):
            if save_branding_settings_to_supabase(tenant_key, legacy_branding):
                legacy_storage = "supabase"
            return tenant_key, legacy_branding, legacy_storage

    return tenant_key, branding, branding_storage


def save_branding_settings_to_supabase(tenant_key: str, settings: dict) -> bool:
    payload = {"tenant_key": tenant_key, **normalize_branding_settings(settings)}
    try:
        supabase.table(BRANDING_SUPABASE_TABLE).upsert(payload, on_conflict="tenant_key").execute()
        return True
    except Exception:
        return False


def save_branding_settings_to_file(tenant_key: str, settings: dict) -> None:
    ensure_branding_storage()
    settings_path = get_branding_settings_path(tenant_key)
    with open(settings_path, "w", encoding="utf-8") as file:
        json.dump(normalize_branding_settings(settings), file, ensure_ascii=False, indent=2)


def save_branding_settings(tenant_key: str, settings: dict) -> str:
    if save_branding_settings_to_supabase(tenant_key, settings):
        save_branding_settings_to_file(tenant_key, settings)
        return "supabase"

    save_branding_settings_to_file(tenant_key, settings)
    return "local"


def get_branding_logo_url(settings: dict) -> Optional[str]:
    logo_path = settings.get("logo_path") or ""
    if not logo_path:
        return None
    if logo_path.startswith("data:image/"):
        return logo_path
    return "/admin/branding/logo"


def save_branding_logo(tenant_key: str, logo_file: UploadFile, current_logo_path: str) -> str:
    ensure_branding_storage()
    upload_dir = get_branding_upload_dir(tenant_key)
    os.makedirs(upload_dir, exist_ok=True)
    filename = logo_file.filename or ""
    extension = os.path.splitext(filename)[1].lower()
    if extension not in ALLOWED_LOGO_EXTENSIONS:
        raise ValueError("Logo must be a PNG, JPG, JPEG, or WEBP file.")

    safe_name = f"brand-logo{extension}"
    target_path = os.path.join(upload_dir, safe_name)

    file_bytes = logo_file.file.read()
    if not file_bytes:
        raise ValueError("Uploaded logo file is empty.")

    mime_type = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }.get(extension, "application/octet-stream")
    encoded = base64.b64encode(file_bytes).decode("ascii")
    logo_data_url = f"data:{mime_type};base64,{encoded}"

    if current_logo_path and not current_logo_path.startswith("data:image/") and current_logo_path != target_path and os.path.exists(current_logo_path):
        os.remove(current_logo_path)

    return logo_data_url


def is_admin_authenticated(request: Request) -> bool:
    return request.session.get("admin_authenticated") is True


def normalize_credential(value: str) -> str:
    normalized = (value or "").strip()
    if len(normalized) >= 2 and normalized[0] == normalized[-1] and normalized[0] in {"'", '"'}:
        normalized = normalized[1:-1]
    return normalized


def parse_int_field(value: str) -> Optional[int]:
    normalized = (value or "").strip()
    if not normalized:
        return None
    return int(normalized)


def parse_float_field(value: str) -> Optional[float]:
    normalized = (value or "").strip().replace(",", ".")
    if not normalized:
        return None
    return float(normalized)


def normalize_asset_tag(value: str) -> str:
    return " ".join((value or "").strip().upper().split())


def validate_asset_tag_format(asset_tag_number: str) -> Optional[str]:
    if not asset_tag_number:
        return "Asset tag/Inventory No. is required."

    if not re.fullmatch(r"[A-Z0-9](?:[A-Z0-9/\- ]*[A-Z0-9])?", asset_tag_number):
        return "Asset tag/Inventory No. may only contain letters, digits, spaces, hyphens, and slashes."

    if not any(character.isdigit() for character in asset_tag_number):
        return "Asset tag/Inventory No. must contain at least one numeric part."

    return None


def require_admin(request: Request) -> Optional[RedirectResponse]:
    if is_admin_authenticated(request):
        return None

    login_url = app.url_path_for("admin_login")
    next_path = request.url.path
    if request.url.query:
        next_path = f"{next_path}?{request.url.query}"

    redirect_url = f"{login_url}?next={next_path}"
    return RedirectResponse(url=redirect_url, status_code=303)


def set_flash(request: Request, level: str, message: str) -> None:
    request.session["admin_flash"] = {"level": level, "message": message}


def pop_flash(request: Request) -> Optional[dict]:
    return request.session.pop("admin_flash", None)


def ensure_sync_storage() -> None:
    os.makedirs(SYNC_STORAGE_DIR, exist_ok=True)


def load_sync_state() -> dict:
    ensure_sync_storage()
    if not os.path.exists(SYNC_STATE_PATH):
        return {}
    with open(SYNC_STATE_PATH, "r", encoding="utf-8") as file:
        return json.load(file)


def save_sync_state(state: dict) -> None:
    ensure_sync_storage()
    with open(SYNC_STATE_PATH, "w", encoding="utf-8") as file:
        json.dump(state, file, ensure_ascii=False, indent=2)


def clean_excel_value(value):
    try:
        import pandas as pd  # type: ignore
    except Exception:
        pd = None

    if pd is not None and pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


SYNC_VALUE_ALIASES = {
    "asset_sub_classification": {
        "other...": "other",
        "other…": "other",
    }
}


def parse_sync_float(value) -> Optional[float]:
    value = clean_excel_value(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    normalized = str(value).strip().replace(" ", "")
    if not normalized:
        return None

    if "," in normalized and "." in normalized:
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")

    try:
        return float(normalized)
    except Exception:
        return None


def safe_excel_int(value, default: Optional[int] = None) -> Optional[int]:
    value = clean_excel_value(value)
    if value is None:
        return default
    try:
        return int(float(value))
    except Exception:
        return default


def safe_excel_float(value) -> Optional[float]:
    return parse_sync_float(value)


def normalize_sync_string(value) -> Optional[str]:
    value = clean_excel_value(value)
    if value is None:
        return None
    normalized = " ".join(str(value).split())
    return normalized or None


def normalize_sync_lookup_value(field_name: str, value) -> Optional[str]:
    normalized = normalize_sync_string(value)
    if normalized is None:
        return None

    canonical = normalized.lower().replace("…", "...").rstrip(". ").strip()
    alias_map = SYNC_VALUE_ALIASES.get(field_name, {})
    canonical = alias_map.get(canonical, canonical)
    return canonical or None


def canonicalize_sync_lookup_value(value) -> Optional[str]:
    normalized = normalize_sync_string(value)
    if normalized is None:
        return None
    canonical = normalized.lower().replace("\u2026", "...").replace("вђ¦", "...")
    canonical = re.sub(r"[.]+", "", canonical)
    canonical = " ".join(canonical.split())
    return canonical or None


def normalize_sync_casefold_string(value) -> Optional[str]:
    normalized = normalize_sync_string(value)
    if normalized is None:
        return None
    return normalized.casefold()


def normalize_sync_number(value) -> Optional[float]:
    parsed = parse_sync_float(value)
    if parsed is None:
        return None
    return round(parsed, 2)


def normalize_sync_value(field_name: str, value):
    if field_name == "purchase_price":
        return normalize_sync_number(value)
    if field_name == "quantity":
        return safe_excel_int(value)
    if field_name in {"asset_classification", "asset_sub_classification"}:
        return canonicalize_sync_lookup_value(value)
    if field_name in {"responsible_person", "department", "city", "location_name", "current_status", "currency"}:
        return normalize_sync_casefold_string(value)
    return normalize_sync_string(value)


def sync_values_equal(field_name: str, current_value, excel_value) -> bool:
    if field_name == "purchase_price":
        current_number = parse_sync_float(current_value)
        excel_number = parse_sync_float(excel_value)
        if current_number is None or excel_number is None:
            return current_number == excel_number
        return abs(current_number - excel_number) < 0.01
    return normalize_sync_value(field_name, current_value) == normalize_sync_value(field_name, excel_value)


def normalize_sync_match_key(value) -> Optional[str]:
    normalized = normalize_sync_string(value)
    if normalized is None:
        return None
    normalized = normalized.replace("`", "'").replace("\u2019", "'")
    normalized = normalized.replace("\u2013", "-").replace("\u2014", "-")
    return normalized.casefold()


def parse_excel_sync_date(value) -> Optional[str]:
    value = clean_excel_value(value)
    if value is None:
        return None

    try:
        import pandas as pd  # type: ignore
    except Exception:
        pd = None

    if pd is not None and isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    normalized = str(value).strip()
    for date_format in ("%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(normalized, date_format).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def parse_excel_purchase_period(value) -> Optional[str]:
    parsed_date = parse_excel_sync_date(value)
    if parsed_date:
        return parsed_date

    normalized = normalize_sync_string(value)
    if not normalized:
        return None

    match = re.search(r"\b(\d{1,2})[.\-/](\d{4})\b", normalized)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
        if 1 <= month <= 12:
            return f"{year:04d}-{month:02d}-01"

    match = re.search(r"\b(\d{4})[.\-/](\d{1,2})\b", normalized)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return f"{year:04d}-{month:02d}-01"

    return None


def format_purchase_period(payment_date: Optional[str]) -> Optional[str]:
    if not payment_date:
        return None
    try:
        parsed = datetime.strptime(str(payment_date)[:10], "%Y-%m-%d")
    except ValueError:
        return None
    return parsed.strftime("%m-%Y")


def parse_payment_amount(value: str) -> Optional[float]:
    normalized = value.replace(" ", "").replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return None


def get_excel_payment_records(record: dict) -> list[dict]:
    remarks = normalize_sync_string(record.get("remarks")) or ""
    payments = []
    amount_pattern = r"(?:\d{1,3}(?:[\s\u00a0]\d{3})+|\d+)(?:[,.]\d{1,2})?"
    pattern = re.compile(
        rf"(?P<amount>{amount_pattern})\s*"
        r"(?P<currency>[A-Z]{3})\s*-\s*"
        r"(?P<date>\d{1,2}[./-]\d{1,2}[./-]\d{4})"
        rf"(?P<note>.*?)(?={amount_pattern}\s*[A-Z]{{3}}\s*-\s*\d{{1,2}}[./-]\d{{1,2}}[./-]\d{{4}}|$)",
        re.IGNORECASE,
    )
    for match in pattern.finditer(remarks):
        payment_date = parse_excel_sync_date(match.group("date"))
        payment_amount = parse_payment_amount(match.group("amount"))
        if not payment_date or payment_amount is None:
            continue
        payments.append(
            {
                "payment_date": payment_date,
                "payment_amount": payment_amount,
                "currency": match.group("currency").upper(),
                "payment_status": "paid",
                "notes": (match.group("note") or "").strip() or None,
            }
        )

    if payments:
        return payments

    purchase_date = parse_excel_purchase_period(record.get("purchase_date_raw"))
    if not purchase_date:
        return []

    return [
        {
            "payment_date": purchase_date,
            "payment_amount": record.get("purchase_price"),
            "currency": record.get("currency"),
            "payment_status": "paid",
            "notes": "Imported from Excel purchase period",
        }
    ]


def get_latest_payment_period(payments: list[dict]) -> Optional[str]:
    payment_dates = [payment.get("payment_date") for payment in payments if payment.get("payment_date")]
    if not payment_dates:
        return None
    return format_purchase_period(max(str(date_value)[:10] for date_value in payment_dates))


def normalize_excel_asset_record(row: dict, usage_type: str = "standard", source_sheet: Optional[str] = None) -> Optional[dict]:
    asset_tag = normalize_asset_tag(clean_excel_value(row.get("asset_tag_number")) or "")
    if not asset_tag:
        return None

    remarks = clean_excel_value(row.get("remarks"))

    return {
        "asset_tag_number": asset_tag,
        "usage_type": normalize_asset_usage_type(usage_type, asset_tag),
        "source_sheet": source_sheet,
        "asset_classification": clean_excel_value(row.get("asset_classification")),
        "asset_sub_classification": clean_excel_value(row.get("asset_sub_classification")),
        "item_description": clean_excel_value(row.get("item_description")),
        "brand_make": clean_excel_value(row.get("brand_make")),
        "model": clean_excel_value(row.get("model")),
        "serial_chassis_number": clean_excel_value(row.get("serial_number")),
        "quantity": safe_excel_int(row.get("quantity"), default=1),
        "purchase_price": safe_excel_float(row.get("purchase_price")),
        "purchase_date_raw": clean_excel_value(row.get("purchase_date_raw")),
        "currency": clean_excel_value(row.get("currency")),
        "current_status": clean_excel_value(row.get("current_status")),
        "remarks": remarks,
        "location_name": clean_excel_value(row.get("location_name")),
        "department_name": clean_excel_value(row.get("department_name")),
        "recipient_name": clean_excel_value(row.get("recipient_name")),
        "recipient_position": clean_excel_value(row.get("recipient_position")),
        "purchased_project_no": clean_excel_value(row.get("purchased_project_no")),
        "transferred_project_no": clean_excel_value(row.get("transferred_project_no")),
        "purchased_donor_name": clean_excel_value(row.get("purchased_donor_name") or row.get("donor_name")),
        "transferred_donor_name": clean_excel_value(row.get("transferred_donor_name") or row.get("donor_name")),
        "donor_name": clean_excel_value(row.get("transferred_donor_name") or row.get("purchased_donor_name") or row.get("donor_name")),
        "last_transfer_date": parse_excel_sync_date(row.get("last_transfer_date")),
    }


def resolve_excel_sync_dataframe_columns(columns) -> dict:
    rename_map = {}
    last_project_field = None
    donor_index = 0

    for column in columns:
        normalized = normalize_excel_header(column)
        field_name = None

        if normalized.startswith("donor"):
            donor_index += 1
            if last_project_field == "purchased_project_no":
                field_name = "purchased_donor_name"
            elif last_project_field == "transferred_project_no":
                field_name = "transferred_donor_name"
            else:
                field_name = "purchased_donor_name" if donor_index == 1 else "transferred_donor_name"
        else:
            field_name = resolve_excel_header_field(column)
            if not field_name and "project" in normalized:
                if "transfer" in normalized:
                    field_name = "transferred_project_no"
                elif "purchase" in normalized or normalized in {"project", "project no", "project no.", "project number"}:
                    field_name = "purchased_project_no"

        if field_name in {"purchased_project_no", "transferred_project_no"}:
            last_project_field = field_name

        rename_map[column] = field_name or column

    return rename_map


def load_excel_sync_sheet_rows(file_path: str, sheet_name: str, usage_type: str, required: bool = False) -> list[dict]:
    try:
        import pandas as pd  # type: ignore
    except Exception as exc:
        raise ValueError(f"Excel sync requires pandas/openpyxl support: {exc}") from exc

    try:
        dataframe = pd.read_excel(file_path, sheet_name=sheet_name, header=EXCEL_SYNC_HEADER_ROW)
    except Exception as exc:
        if required:
            raise ValueError(f"Could not read Excel sheet '{sheet_name}': {exc}") from exc
        return []

    dataframe = dataframe.rename(columns=resolve_excel_sync_dataframe_columns(dataframe.columns))
    if "asset_tag_number" not in dataframe.columns:
        if required:
            raise ValueError(f"The workbook sheet '{sheet_name}' does not contain the expected Asset Tag column.")
        return []

    has_recipient_column = "recipient_name" in dataframe.columns
    has_project_column = "purchased_project_no" in dataframe.columns or "transferred_project_no" in dataframe.columns

    dataframe = dataframe[dataframe["asset_tag_number"].notna()].copy()
    records: list[dict] = []
    for _, row in dataframe.iterrows():
        normalized = normalize_excel_asset_record(row.to_dict(), usage_type=usage_type, source_sheet=sheet_name)
        if normalized:
            normalized["_has_recipient_column"] = has_recipient_column
            normalized["_has_project_column"] = has_project_column
            records.append(normalized)
    return records


def load_excel_sync_rows(file_path: str) -> list[dict]:
    standard_rows = load_excel_sync_sheet_rows(file_path, EXCEL_SYNC_SHEET_NAME, "standard", required=True)
    low_cost_rows = load_excel_sync_sheet_rows(file_path, EXCEL_LOW_COST_SHEET_NAME, "low_cost", required=False)
    return [*standard_rows, *low_cost_rows]


def normalize_transfer_header(value) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).casefold()


def resolve_transfer_header_field(value) -> Optional[str]:
    normalized = normalize_transfer_header(value)
    if normalized == "no.":
        return "source_log_no"
    if normalized == "transfer date":
        return "transfer_date"
    if normalized.startswith("standart") or normalized.startswith("standard"):
        return "source_asset_type"
    if "asset tag" in normalized and "inventory" in normalized:
        return "asset_tag_number"
    if normalized.startswith("brand-make"):
        return "description_snapshot"
    if normalized.startswith("serial"):
        return "serial_snapshot"
    if normalized == "holder":
        return "from_holder_name"
    if normalized == "current project":
        return "from_project_raw"
    if normalized == "new project":
        return "to_project_raw"
    if normalized == "new holder":
        return "to_holder_name"
    if normalized.startswith("current status"):
        return "asset_status"
    if normalized == "asset condition description":
        return "asset_condition_description"
    if normalized == "reason for asset transfer":
        return "transfer_reason"
    return None


def normalize_transfer_holder(value) -> Optional[str]:
    normalized = normalize_sync_string(value)
    if not normalized:
        return None
    if normalized.casefold() == "warehouse":
        return "Warehouse"
    return normalized


def make_transfer_key(record: dict) -> str:
    return "|".join(
        [
            normalize_asset_tag(record.get("asset_tag_number") or ""),
            str(record.get("transfer_date") or ""),
            str(record.get("source_log_no") or ""),
            normalize_sync_match_key(record.get("from_holder_name")) or "",
            normalize_sync_match_key(record.get("to_holder_name")) or "",
        ]
    )


def normalize_excel_transfer_record(row: dict, source_row_number: int) -> Optional[dict]:
    asset_tag = normalize_asset_tag(clean_excel_value(row.get("asset_tag_number")) or "")
    transfer_date = parse_excel_sync_date(row.get("transfer_date"))
    if not asset_tag or not transfer_date:
        return None

    record = {
        "transfer_key": "",
        "source_row_number": source_row_number,
        "source_log_no": safe_excel_int(row.get("source_log_no")),
        "transfer_date": transfer_date,
        "source_asset_type": clean_excel_value(row.get("source_asset_type")),
        "asset_tag_number": asset_tag,
        "asset_tag_snapshot": asset_tag,
        "description_snapshot": clean_excel_value(row.get("description_snapshot")),
        "serial_snapshot": clean_excel_value(row.get("serial_snapshot")),
        "from_holder_name": normalize_transfer_holder(row.get("from_holder_name")),
        "from_project_raw": clean_excel_value(row.get("from_project_raw")),
        "to_project_raw": clean_excel_value(row.get("to_project_raw")),
        "to_holder_name": normalize_transfer_holder(row.get("to_holder_name")),
        "asset_status": clean_excel_value(row.get("asset_status")),
        "asset_condition_description": clean_excel_value(row.get("asset_condition_description")),
        "transfer_reason": clean_excel_value(row.get("transfer_reason")),
    }
    if not any(
        record.get(field_name)
        for field_name in [
            "from_holder_name",
            "to_holder_name",
            "from_project_raw",
            "to_project_raw",
            "asset_status",
            "asset_condition_description",
            "transfer_reason",
        ]
    ):
        return None
    record["transfer_key"] = make_transfer_key(record)
    return record


def load_excel_transfer_log_rows(file_path: str) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise ValueError(f"Excel transfer log sync requires openpyxl support: {exc}") from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    if EXCEL_TRANSFER_LOG_SHEET_NAME not in workbook.sheetnames:
        return []

    sheet = workbook[EXCEL_TRANSFER_LOG_SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    header_values = None
    for index, row in enumerate(rows, start=1):
        if index == EXCEL_TRANSFER_LOG_HEADER_ROW + 1:
            header_values = row
            break
    if not header_values:
        return []

    columns = {}
    for index, value in enumerate(header_values):
        field_name = resolve_transfer_header_field(value)
        if field_name:
            columns[field_name] = index

    if "asset_tag_number" not in columns or "transfer_date" not in columns:
        return []

    records = []
    for row_number, values in enumerate(rows, start=EXCEL_TRANSFER_LOG_HEADER_ROW + 2):
        row = {
            field_name: values[column_index] if column_index < len(values) else None
            for field_name, column_index in columns.items()
        }
        if not any(value is not None for value in row.values()):
            continue
        normalized = normalize_excel_transfer_record(row, row_number)
        if normalized:
            records.append(normalized)
    return records


def list_asset_records(batch_size: int = 1000) -> list[dict]:
    rows: list[dict] = []
    start = 0
    while True:
        response = (
            supabase.table("assets")
            .select("*")
            .order("asset_tag_number")
            .range(start, start + batch_size - 1)
            .execute()
        )
        batch = response.data or []
        rows.extend(batch)
        if len(batch) < batch_size:
            break
        start += len(batch)
    return rows


def list_current_assignment_records(batch_size: int = 1000) -> list[dict]:
    rows: list[dict] = []
    start = 0
    while True:
        response = (
            supabase.table("asset_assignments")
            .select("*")
            .is_("return_date", "null")
            .order("assignment_date", desc=True)
            .range(start, start + batch_size - 1)
            .execute()
        )
        batch = response.data or []
        rows.extend(batch)
        if len(batch) < batch_size:
            break
        start += len(batch)
    return rows


def list_asset_project_records(batch_size: int = 1000) -> list[dict]:
    rows: list[dict] = []
    start = 0
    while True:
        response = (
            supabase.table("asset_projects")
            .select("*")
            .order("is_primary", desc=True)
            .order("asset_project_id")
            .range(start, start + batch_size - 1)
            .execute()
        )
        batch = response.data or []
        rows.extend(batch)
        if len(batch) < batch_size:
            break
        start += len(batch)
    return rows


def list_asset_payment_records(batch_size: int = 1000) -> list[dict]:
    rows: list[dict] = []
    start = 0
    while True:
        response = (
            supabase.table("asset_payments")
            .select("*")
            .order("asset_id")
            .order("payment_date")
            .range(start, start + batch_size - 1)
            .execute()
        )
        batch = response.data or []
        rows.extend(batch)
        if len(batch) < batch_size:
            break
        start += len(batch)
    return rows


def list_asset_transfer_records(batch_size: int = 1000) -> list[dict]:
    rows: list[dict] = []
    start = 0
    while True:
        response = (
            supabase.table("asset_transfers")
            .select("*")
            .order("asset_id")
            .order("transfer_date")
            .range(start, start + batch_size - 1)
            .execute()
        )
        batch = response.data or []
        rows.extend(batch)
        if len(batch) < batch_size:
            break
        start += len(batch)
    return rows


_ASSET_PROJECT_PURCHASE_ORIGIN_SUPPORTED: Optional[bool] = None


def asset_project_purchase_origin_supported() -> bool:
    global _ASSET_PROJECT_PURCHASE_ORIGIN_SUPPORTED
    if _ASSET_PROJECT_PURCHASE_ORIGIN_SUPPORTED is not None:
        return _ASSET_PROJECT_PURCHASE_ORIGIN_SUPPORTED

    try:
        supabase.table("asset_projects").select("is_purchase_origin").limit(1).execute()
    except Exception as error:
        if isinstance(error, APIError):
            message = " ".join(part for part in [error.message or "", error.details or ""] if part).lower()
            if "is_purchase_origin" in message and "column" in message:
                _ASSET_PROJECT_PURCHASE_ORIGIN_SUPPORTED = False
                return False
        _ASSET_PROJECT_PURCHASE_ORIGIN_SUPPORTED = False
        return False

    _ASSET_PROJECT_PURCHASE_ORIGIN_SUPPORTED = True
    return True


def build_person_lookup(people: list[dict]) -> dict[str, dict]:
    lookup = {}
    for person in people:
        for field_name in ("name_eng", "name", "full_name"):
            key = normalize_sync_match_key(person.get(field_name))
            if key and key not in lookup:
                lookup[key] = person
    return lookup


def build_location_lookup(locations: list[dict]) -> dict[tuple[Optional[str], Optional[str]], dict]:
    lookup = {}
    for location in locations:
        city_key = normalize_sync_match_key(location.get("city") or location.get("name"))
        department_key = normalize_sync_match_key(location.get("department"))
        if city_key:
            lookup[(city_key, department_key)] = location
            lookup.setdefault((city_key, None), location)
    return lookup


def build_project_lookup(projects: list[dict]) -> dict[str, dict]:
    lookup = {}
    for project in projects:
        for field_name in ("project_number", "project_name", "name"):
            key = normalize_sync_match_key(project.get(field_name))
            if key and key not in lookup:
                lookup[key] = project
    return lookup


def build_donor_lookup(donors: list[dict]) -> dict[str, dict]:
    return {
        key: donor
        for donor in donors
        for key in [normalize_sync_match_key(donor.get("donor_name"))]
        if key
    }


def resolve_excel_person(record: dict, person_lookup: dict[str, dict]) -> Optional[dict]:
    recipient_key = normalize_sync_match_key(record.get("recipient_name"))
    return person_lookup.get(recipient_key) if recipient_key else None


def resolve_excel_location(record: dict, person: Optional[dict], location_lookup: dict[tuple[Optional[str], Optional[str]], dict]) -> Optional[dict]:
    city_key = normalize_sync_match_key(record.get("location_name"))
    department_key = normalize_sync_match_key(record.get("department_name") or (person or {}).get("department"))
    if not city_key:
        return None
    return location_lookup.get((city_key, department_key)) or location_lookup.get((city_key, None))


def get_excel_current_project_number(record: dict) -> Optional[str]:
    return normalize_sync_string(record.get("transferred_project_no") or record.get("purchased_project_no"))


def split_excel_project_numbers(value) -> list[str]:
    normalized = normalize_sync_string(value)
    if not normalized:
        return []
    return re.findall(PROJECT_NUMBER_PATTERN, normalized.upper())


def split_excel_project_percentages(value, expected_count: int) -> list[Optional[float]]:
    normalized = normalize_sync_string(value)
    if not normalized or expected_count <= 0:
        return []

    percentages = [
        float(match.replace(",", "."))
        for match in re.findall(r"(\d+(?:[.,]\d+)?)\s*%", normalized)
    ]
    if len(percentages) == expected_count:
        return percentages
    return []


def get_project_allocations_from_value(value) -> list[dict]:
    project_numbers = split_excel_project_numbers(value)
    if not project_numbers:
        return []

    explicit_percentages = split_excel_project_percentages(value, len(project_numbers))
    fallback_percent = round(100 / len(project_numbers), 2)
    return [
        {
            "project_number": project_number,
            "allocation_percent": explicit_percentages[index] if explicit_percentages else fallback_percent,
        }
        for index, project_number in enumerate(project_numbers)
    ]


def get_excel_purchased_project_allocations(record: dict) -> list[dict]:
    return get_project_allocations_from_value(record.get("purchased_project_no"))


def get_excel_transferred_project_allocations(record: dict) -> list[dict]:
    return get_project_allocations_from_value(record.get("transferred_project_no"))


def get_excel_project_allocations(record: dict) -> list[dict]:
    return get_excel_transferred_project_allocations(record) or get_excel_purchased_project_allocations(record)


def format_project_allocations(allocations: list[dict]) -> Optional[str]:
    if not allocations:
        return None
    if len(allocations) == 1:
        return allocations[0].get("project_number")
    parts = []
    for allocation in allocations:
        allocation_percent = allocation.get("allocation_percent")
        if allocation_percent is None:
            parts.append(str(allocation.get("project_number")))
        else:
            parts.append(f"{allocation.get('project_number')} {allocation_percent:g}%")
    return " / ".join(parts)


def project_allocation_signature(allocations: list[dict]) -> tuple:
    if len(allocations) == 1:
        return ((normalize_sync_match_key(allocations[0].get("project_number")),),)
    return tuple(
        sorted(
            (
                normalize_sync_match_key(allocation.get("project_number")),
                round(float(allocation.get("allocation_percent") or 0), 2),
            )
            for allocation in allocations
        )
    )


def get_current_project_allocations(asset_projects: list[dict]) -> list[dict]:
    return [
        {
            "project_number": row.get("project_number"),
            "allocation_percent": normalize_sync_number(row.get("allocation_percent")),
        }
        for row in get_current_project_rows(asset_projects)
        if row and row.get("project_number")
    ]


def get_current_project_rows(asset_projects: list[dict]) -> list[dict]:
    current_projects = [row for row in asset_projects if row.get("is_current") is True]
    if current_projects:
        return current_projects
    primary_project = select_current_project(asset_projects)
    return [primary_project] if primary_project else []


def get_purchased_project_rows(asset_projects: list[dict]) -> list[dict]:
    if not asset_projects:
        return []
    if any("is_purchase_origin" in row for row in asset_projects):
        purchase_origin_projects = [row for row in asset_projects if row.get("is_purchase_origin") is True]
        return purchase_origin_projects or get_current_project_rows(asset_projects)
    non_current_projects = [row for row in asset_projects if row.get("is_current") is not True]
    return non_current_projects or get_current_project_rows(asset_projects)


def get_project_allocations_from_rows(asset_projects: list[dict]) -> list[dict]:
    return [
        {
            "project_number": row.get("project_number"),
            "allocation_percent": normalize_sync_number(row.get("allocation_percent")),
        }
        for row in asset_projects
        if row and row.get("project_number")
    ]


def select_current_project(asset_projects: list[dict]) -> Optional[dict]:
    current_projects = [row for row in asset_projects if row.get("is_current") is True]
    if current_projects:
        return current_projects[0]
    primary_projects = [row for row in asset_projects if row.get("is_primary") is True]
    if primary_projects:
        return primary_projects[0]
    return asset_projects[0] if asset_projects else None


def build_sync_context() -> dict:
    people = list_people()
    locations = list_locations()
    projects = list_projects()
    donors = list_donors()
    assignments = list_current_assignment_records()
    asset_projects = list_asset_project_records()
    asset_payments = list_asset_payment_records()
    asset_transfers = list_asset_transfer_records()

    people_by_id = {row.get("person_id"): row for row in people}
    locations_by_id = {row.get("location_id"): row for row in locations}
    projects_by_id = {row.get("project_id"): row for row in projects}
    donors_by_id = {row.get("donor_id"): row for row in donors}

    assignment_by_asset_id = {}
    for assignment in assignments:
        asset_id = assignment.get("asset_id")
        if asset_id not in assignment_by_asset_id:
            person = people_by_id.get(assignment.get("person_id")) or {}
            location = locations_by_id.get(assignment.get("location_id")) or {}
            assignment_by_asset_id[asset_id] = {
                **assignment,
                "responsible_person": get_person_display_name(person) if person else None,
                "department": person.get("department") or location.get("department"),
                "city": location.get("city") or location.get("name"),
            }

    projects_by_asset_id: dict[int, list[dict]] = {}
    for asset_project in asset_projects:
        project = projects_by_id.get(asset_project.get("project_id")) or {}
        donor = donors_by_id.get(asset_project.get("donor_id")) or {}
        projects_by_asset_id.setdefault(asset_project.get("asset_id"), []).append(
            {
                **asset_project,
                "project_number": project.get("project_number"),
                "project_name": project.get("project_name") or project.get("name"),
                "donor_name": donor.get("donor_name"),
            }
        )

    payments_by_asset_id: dict[int, list[dict]] = {}
    for payment in asset_payments:
        payments_by_asset_id.setdefault(payment.get("asset_id"), []).append(payment)

    transfer_signatures = set()
    for transfer in asset_transfers:
        transfer_signatures.add(
            "|".join(
                [
                    str(transfer.get("asset_id") or ""),
                    str(transfer.get("transfer_date") or ""),
                    str(transfer.get("source_log_no") or ""),
                    normalize_sync_match_key(transfer.get("from_holder_name")) or "",
                    normalize_sync_match_key(transfer.get("to_holder_name")) or "",
                ]
            )
        )

    return {
        "person_lookup": build_person_lookup(people),
        "location_lookup": build_location_lookup(locations),
        "project_lookup": build_project_lookup(projects),
        "donor_lookup": build_donor_lookup(donors),
        "assignment_by_asset_id": assignment_by_asset_id,
        "projects_by_asset_id": projects_by_asset_id,
        "payments_by_asset_id": payments_by_asset_id,
        "transfer_signatures": transfer_signatures,
        "supports_asset_project_purchase_origin": asset_project_purchase_origin_supported(),
    }


def make_transfer_signature(asset_id: int, record: dict) -> str:
    return "|".join(
        [
            str(asset_id or ""),
            str(record.get("transfer_date") or ""),
            str(record.get("source_log_no") or ""),
            normalize_sync_match_key(record.get("from_holder_name")) or "",
            normalize_sync_match_key(record.get("to_holder_name")) or "",
        ]
    )


def transfer_project_allocations(value) -> list[dict]:
    allocations = get_project_allocations_from_value(value)
    if allocations:
        return allocations
    raw = normalize_sync_string(value)
    if not raw:
        return []
    return [{"project_number": raw, "allocation_percent": None}]


def build_transfer_log_preview(transfer_records: list[dict], current_by_tag: dict[str, dict], sync_context: dict) -> dict:
    new_transfer_records = []
    skipped_existing = 0
    unmatched_assets = 0

    for record in transfer_records:
        warnings = []
        asset = current_by_tag.get(normalize_asset_tag(record.get("asset_tag_number") or ""))
        if not asset:
            unmatched_assets += 1
            warnings.append(f"Asset not found: {record.get('asset_tag_number')}")

        from_holder = record.get("from_holder_name")
        to_holder = record.get("to_holder_name")
        from_person = None
        to_person = None
        if from_holder and from_holder.casefold() != "warehouse":
            from_person = sync_context["person_lookup"].get(normalize_sync_match_key(from_holder))
            if not from_person:
                warnings.append(f"Holder not found in People: {from_holder}")
        if to_holder and to_holder.casefold() != "warehouse":
            to_person = sync_context["person_lookup"].get(normalize_sync_match_key(to_holder))
            if not to_person:
                warnings.append(f"New holder not found in People: {to_holder}")

        for direction, field_name in [("Current project", "from_project_raw"), ("New project", "to_project_raw")]:
            for allocation in transfer_project_allocations(record.get(field_name)):
                project_number = allocation.get("project_number")
                if project_number and not sync_context["project_lookup"].get(normalize_sync_match_key(project_number)):
                    warnings.append(f"{direction} not found in Projects: {project_number}")

        if asset:
            signature = make_transfer_signature(asset.get("asset_id"), record)
            if signature in sync_context.get("transfer_signatures", set()):
                skipped_existing += 1
                continue

        new_transfer_records.append(
            {
                **record,
                "asset_id": asset.get("asset_id") if asset else None,
                "from_person_id": from_person.get("person_id") if from_person else None,
                "to_person_id": to_person.get("person_id") if to_person else None,
                "warnings": warnings,
                "can_apply": bool(asset),
            }
        )

    return {
        "summary": {
            "excel_rows": len(transfer_records),
            "new_records": len(new_transfer_records),
            "skipped_existing": skipped_existing,
            "unmatched_assets": unmatched_assets,
        },
        "new_records": new_transfer_records,
    }


def build_sync_preview(excel_records: list[dict], current_assets: list[dict], transfer_records: Optional[list[dict]] = None) -> dict:
    sync_context = build_sync_context()
    current_by_tag = {
        normalize_asset_tag(asset.get("asset_tag_number") or ""): asset
        for asset in current_assets
        if asset.get("asset_tag_number")
    }
    synced_fields = [
        "usage_type",
        "asset_classification",
        "asset_sub_classification",
        "item_description",
        "brand_make",
        "model",
        "serial_chassis_number",
        "quantity",
        "purchase_price",
        "currency",
        "current_status",
    ]

    new_records: list[dict] = []
    changed_records: list[dict] = []
    unchanged_count = 0
    standard_count = 0
    low_cost_count = 0

    for record in excel_records:
        if normalize_asset_usage_type(record.get("usage_type"), record.get("asset_tag_number")) == "low_cost":
            low_cost_count += 1
        else:
            standard_count += 1
        asset_tag = record["asset_tag_number"]
        current_asset = current_by_tag.get(asset_tag)
        if not current_asset:
            new_records.append(record)
            continue

        changed_fields = []
        current_values = {}
        excel_values = {}
        warnings = []
        for field_name in synced_fields:
            current_value = current_asset.get(field_name)
            excel_value = record.get(field_name)
            if not sync_values_equal(field_name, current_value, excel_value):
                changed_fields.append(field_name)
                if field_name == "usage_type":
                    current_values[field_name] = get_asset_usage_type_label(current_value)
                    excel_values[field_name] = get_asset_usage_type_label(excel_value)
                else:
                    current_values[field_name] = current_value
                    excel_values[field_name] = excel_value

        if record.get("_has_recipient_column"):
            current_assignment = sync_context["assignment_by_asset_id"].get(current_asset.get("asset_id")) or {}
            excel_person = resolve_excel_person(record, sync_context["person_lookup"])
            excel_recipient = normalize_sync_string(record.get("recipient_name"))
            current_person_name = current_assignment.get("responsible_person")

            if not sync_values_equal("responsible_person", current_person_name, excel_recipient):
                changed_fields.append("responsible_person")
                current_values["responsible_person"] = current_person_name
                excel_values["responsible_person"] = excel_recipient
                if excel_recipient and not excel_person:
                    warnings.append(f"Responsible person not found in People: {excel_recipient}")

            excel_location = resolve_excel_location(record, excel_person, sync_context["location_lookup"])
            if excel_recipient and excel_person and not excel_location:
                warnings.append(
                    "Location not found for assignment: "
                    f"{record.get('location_name') or '-'} / "
                    f"{record.get('department_name') or excel_person.get('department') or '-'}"
                )

        if record.get("_has_project_column"):
            current_asset_projects = sync_context["projects_by_asset_id"].get(current_asset.get("asset_id"), [])
            current_purchased_allocations = get_project_allocations_from_rows(get_purchased_project_rows(current_asset_projects))
            current_transferred_allocations = get_project_allocations_from_rows(get_current_project_rows(current_asset_projects))
            excel_purchased_allocations = get_excel_purchased_project_allocations(record)
            excel_transferred_allocations = get_excel_transferred_project_allocations(record)

            if excel_purchased_allocations and project_allocation_signature(current_purchased_allocations) != project_allocation_signature(excel_purchased_allocations):
                changed_fields.append("purchased_project_number")
                current_values["purchased_project_number"] = format_project_allocations(current_purchased_allocations)
                excel_values["purchased_project_number"] = format_project_allocations(excel_purchased_allocations)
                for allocation in excel_purchased_allocations:
                    project_number = allocation.get("project_number")
                    project_key = normalize_sync_match_key(project_number)
                    if project_number and not sync_context["project_lookup"].get(project_key):
                        warnings.append(f"Purchased project not found: {project_number}")

            if excel_transferred_allocations and project_allocation_signature(current_transferred_allocations) != project_allocation_signature(excel_transferred_allocations):
                changed_fields.append("transferred_project_number")
                current_values["transferred_project_number"] = format_project_allocations(current_transferred_allocations)
                excel_values["transferred_project_number"] = format_project_allocations(excel_transferred_allocations)
                for allocation in excel_transferred_allocations:
                    project_number = allocation.get("project_number")
                    project_key = normalize_sync_match_key(project_number)
                    if project_number and not sync_context["project_lookup"].get(project_key):
                        warnings.append(f"Transferred project not found: {project_number}")

        excel_payments = get_excel_payment_records(record)
        excel_payment_period = get_latest_payment_period(excel_payments)
        current_payments = sync_context["payments_by_asset_id"].get(current_asset.get("asset_id"), [])
        current_payment_period = get_latest_payment_period(current_payments)
        if excel_payment_period and current_payment_period != excel_payment_period:
            changed_fields.append("purchase_date_raw")
            current_values["purchase_date_raw"] = current_payment_period
            excel_values["purchase_date_raw"] = excel_payment_period

        if changed_fields:
            changed_records.append(
                {
                    "asset_id": current_asset.get("asset_id"),
                    "asset_tag_number": asset_tag,
                    "changed_fields": changed_fields,
                    "current": current_values,
                    "excel": excel_values,
                    "record": record,
                    "warnings": warnings,
                }
            )
        else:
            unchanged_count += 1

    return {
        "summary": {
            "excel_rows": len(excel_records),
            "standard_rows": standard_count,
            "low_cost_rows": low_cost_count,
            "new_records": len(new_records),
            "changed_records": len(changed_records),
            "unchanged_records": unchanged_count,
        },
        "new_records": new_records,
        "changed_records": changed_records,
        "transfer_log": build_transfer_log_preview(transfer_records or [], current_by_tag, sync_context),
    }


def apply_sync_assignment(asset_id: int, record: dict, sync_context: dict) -> int:
    excel_recipient = normalize_sync_string(record.get("recipient_name"))
    assignment_date = record.get("last_transfer_date") or datetime.now(ZoneInfo("Europe/Kyiv")).strftime("%Y-%m-%d")

    if not excel_recipient:
        close_current_assignments(asset_id, assignment_date)
        return 1

    person = resolve_excel_person(record, sync_context["person_lookup"])
    if not person:
        return 0

    location = resolve_excel_location(record, person, sync_context["location_lookup"])
    if not location:
        return 0

    existing = sync_context.get("assignment_by_asset_id", {}).get(asset_id) or {}
    if (
        existing
        and existing.get("person_id") == person.get("person_id")
        and existing.get("location_id") == location.get("location_id")
    ):
        return 0

    notes_parts = []
    if record.get("recipient_position"):
        notes_parts.append(f"Position from Excel: {record.get('recipient_position')}")
    if record.get("current_status"):
        notes_parts.append(f"Asset status: {record.get('current_status')}")
    if record.get("remarks"):
        notes_parts.append(f"Remarks: {record.get('remarks')}")

    close_current_assignments(asset_id, assignment_date)
    supabase.table("asset_assignments").insert(
        {
            "asset_id": asset_id,
            "person_id": person.get("person_id"),
            "location_id": location.get("location_id"),
            "assignment_date": assignment_date,
            "return_date": None,
            "status": record.get("current_status"),
            "notes": " | ".join(notes_parts) if notes_parts else None,
        }
    ).execute()
    return 1


def apply_sync_project(asset_id: int, record: dict, sync_context: dict) -> int:
    purchased_allocations = get_excel_purchased_project_allocations(record)
    transferred_allocations = get_excel_transferred_project_allocations(record)
    if not purchased_allocations and not transferred_allocations:
        return 0

    purchased_donor = None
    purchased_donor_key = normalize_sync_match_key(record.get("purchased_donor_name") or record.get("donor_name"))
    if purchased_donor_key:
        purchased_donor = sync_context["donor_lookup"].get(purchased_donor_key)

    transferred_donor = None
    transferred_donor_key = normalize_sync_match_key(record.get("transferred_donor_name") or record.get("donor_name"))
    if transferred_donor_key:
        transferred_donor = sync_context["donor_lookup"].get(transferred_donor_key)

    project_rows = sync_context.get("projects_by_asset_id", {}).get(asset_id, [])
    existing_by_project_id = {row.get("project_id"): row for row in project_rows}
    supports_purchase_origin = sync_context.get("supports_asset_project_purchase_origin") is True

    def resolve_allocations(allocations: list[dict]) -> Optional[list[dict]]:
        resolved = []
        for allocation in allocations:
            project_number = allocation.get("project_number")
            project = sync_context["project_lookup"].get(normalize_sync_match_key(project_number))
            if not project:
                return None
            resolved.append({**allocation, "project": project})
        return resolved

    resolved_purchased = resolve_allocations(purchased_allocations)
    if resolved_purchased is None:
        return 0

    resolved_transferred = resolve_allocations(transferred_allocations)
    if resolved_transferred is None:
        return 0

    reset_payload = {"is_current": False, "is_primary": False}
    if supports_purchase_origin:
        reset_payload["is_purchase_origin"] = False
    supabase.table("asset_projects").update(reset_payload).eq("asset_id", asset_id).execute()

    applied = 0
    has_transferred_project = bool(resolved_transferred)
    rows_by_project_id = {}

    def merge_project_payload(
        allocation: dict,
        donor: Optional[dict],
        is_purchase_origin: bool,
        is_current: bool,
        is_primary: bool,
        transfer_date: Optional[str],
        transfer_reason: Optional[str],
    ) -> None:
        project = allocation["project"]
        project_id = project.get("project_id")
        payload = rows_by_project_id.setdefault(
            project_id,
            {
                "project_id": project_id,
                "donor_id": donor.get("donor_id") if donor else None,
                "allocation_percent": allocation.get("allocation_percent"),
                "is_current": False,
                "is_primary": False,
                "transfer_date": None,
                "transfer_reason": None,
                "condition_at_transfer": record.get("current_status"),
            },
        )
        if donor:
            payload["donor_id"] = donor.get("donor_id")
        payload["allocation_percent"] = allocation.get("allocation_percent")
        payload["is_current"] = payload["is_current"] or is_current
        payload["is_primary"] = payload["is_primary"] or is_primary
        if transfer_date:
            payload["transfer_date"] = transfer_date
        if transfer_reason:
            payload["transfer_reason"] = transfer_reason
        if supports_purchase_origin:
            payload["is_purchase_origin"] = payload.get("is_purchase_origin", False) or is_purchase_origin

    for index, allocation in enumerate(resolved_purchased or []):
        merge_project_payload(
            allocation=allocation,
            donor=purchased_donor,
            is_purchase_origin=True,
            is_current=not has_transferred_project,
            is_primary=not has_transferred_project and index == 0,
            transfer_date=None,
            transfer_reason=None,
        )
    for index, allocation in enumerate(resolved_transferred or []):
        merge_project_payload(
            allocation=allocation,
            donor=transferred_donor,
            is_purchase_origin=False,
            is_current=True,
            is_primary=index == 0,
            transfer_date=record.get("last_transfer_date"),
            transfer_reason="Excel synchronization",
        )

    for project_id, payload in rows_by_project_id.items():
        existing = existing_by_project_id.get(project_id)

        if existing:
            supabase.table("asset_projects").update(payload).eq("asset_project_id", existing["asset_project_id"]).execute()
        else:
            supabase.table("asset_projects").insert({"asset_id": asset_id, **payload}).execute()
        applied += 1

    return applied


def apply_sync_payments(asset_id: int, record: dict) -> int:
    payments = get_excel_payment_records(record)
    if not payments:
        return 0

    supabase.table("asset_payments").delete().eq("asset_id", asset_id).execute()
    payloads = []
    for index, payment in enumerate(payments, start=1):
        payloads.append(
            {
                "asset_id": asset_id,
                "payment_number": index,
                "payment_date": payment.get("payment_date"),
                "payment_amount": payment.get("payment_amount") or 0,
                "currency": payment.get("currency") or record.get("currency") or "EUR",
                "payment_status": payment.get("payment_status") or "paid",
                "notes": payment.get("notes"),
            }
        )

    supabase.table("asset_payments").insert(payloads).execute()
    return len(payloads)


def apply_transfer_project_rows(transfer_id: int, record: dict, sync_context: dict) -> int:
    payloads = []
    for direction, field_name in [("from", "from_project_raw"), ("to", "to_project_raw")]:
        for allocation in transfer_project_allocations(record.get(field_name)):
            project_number = allocation.get("project_number")
            project = sync_context["project_lookup"].get(normalize_sync_match_key(project_number))
            payloads.append(
                {
                    "transfer_id": transfer_id,
                    "direction": direction,
                    "project_id": project.get("project_id") if project else None,
                    "project_number_raw": project_number,
                    "allocation_percent": allocation.get("allocation_percent"),
                }
            )

    if not payloads:
        return 0

    supabase.table("asset_transfer_projects").delete().eq("transfer_id", transfer_id).execute()
    supabase.table("asset_transfer_projects").insert(payloads).execute()
    return len(payloads)


def apply_sync_transfer(record: dict, sync_context: dict) -> int:
    asset_id = record.get("asset_id")
    if not asset_id:
        return 0

    payload = {
        "asset_id": asset_id,
        "from_person_id": record.get("from_person_id"),
        "to_person_id": record.get("to_person_id"),
        "transfer_date": record.get("transfer_date"),
        "transfer_reason": record.get("transfer_reason"),
        "notes": record.get("transfer_reason"),
        "source_log_no": record.get("source_log_no"),
        "source_row_number": record.get("source_row_number"),
        "source_asset_type": record.get("source_asset_type"),
        "asset_tag_snapshot": record.get("asset_tag_snapshot") or record.get("asset_tag_number"),
        "from_holder_name": record.get("from_holder_name"),
        "to_holder_name": record.get("to_holder_name"),
        "from_project_raw": record.get("from_project_raw"),
        "to_project_raw": record.get("to_project_raw"),
        "asset_status": record.get("asset_status"),
        "asset_condition_description": record.get("asset_condition_description"),
    }

    signature = make_transfer_signature(asset_id, record)
    if signature in sync_context.get("transfer_signatures", set()):
        return 0

    response = supabase.table("asset_transfers").insert(payload).execute()
    transfer = (response.data or [{}])[0]
    transfer_id = transfer.get("transfer_id")
    if transfer_id:
        apply_transfer_project_rows(transfer_id, record, sync_context)
        sync_context.setdefault("transfer_signatures", set()).add(signature)
        audit_log_event(
            entity_type="Transfer",
            entity_id=transfer_id,
            entity_label=record.get("asset_tag_number"),
            action="transferred",
            field_name="assignment",
            old_value=record.get("from_holder_name"),
            new_value=record.get("to_holder_name"),
            summary=build_transfer_audit_summary(record),
            source="Excel Transfer log",
            actor="Excel Transfer log",
            event_date=record.get("transfer_date"),
            event_key=f"asset_transfer:{transfer_id}",
        )
    return 1


def apply_sync_preview(preview: dict) -> dict:
    inserted = 0
    updated = 0
    assignment_updated = 0
    project_updated = 0
    payment_updated = 0
    transfer_updated = 0
    skipped_relationships = 0
    sync_context = build_sync_context()
    asset_fields = {
        "usage_type",
        "asset_classification",
        "asset_sub_classification",
        "item_description",
        "brand_make",
        "model",
        "serial_chassis_number",
        "quantity",
        "purchase_price",
        "currency",
        "current_status",
        "remarks",
    }

    for record in preview.get("new_records", []):
        insert_record = {field_name: record.get(field_name) for field_name in asset_fields}
        insert_record["asset_tag_number"] = record.get("asset_tag_number")
        insert_record["inventory_code"] = insert_record.get("asset_tag_number")
        insert_response = supabase.table("assets").insert(insert_record).execute()
        inserted_asset = (insert_response.data or [{}])[0]
        asset_id = inserted_asset.get("asset_id")
        inserted += 1
        audit_log_event(
            entity_type="Asset",
            entity_id=asset_id,
            entity_label=record.get("asset_tag_number"),
            action="created",
            summary=f"Created asset from Excel: {record.get('asset_tag_number')}",
            source="Excel import",
            actor="Excel import",
        )
        if asset_id:
            if record.get("_has_recipient_column") and normalize_sync_string(record.get("recipient_name")):
                assignment_updated += apply_sync_assignment(asset_id, record, sync_context)
            if record.get("_has_project_column") and (
                get_excel_purchased_project_allocations(record) or get_excel_transferred_project_allocations(record)
            ):
                project_updated += apply_sync_project(asset_id, record, sync_context)
            payment_updated += apply_sync_payments(asset_id, record)

    for item in preview.get("changed_records", []):
        asset_id = item.get("asset_id")
        record = item.get("record") or {}
        if not asset_id:
            continue
        update_data = {
            field_name: record.get(field_name)
            for field_name in item.get("changed_fields", [])
            if field_name in asset_fields
        }
        if update_data:
            supabase.table("assets").update(update_data).eq("asset_id", asset_id).execute()
            updated += 1
            audit_log_event(
                entity_type="Asset",
                entity_id=asset_id,
                entity_label=item.get("asset_tag_number"),
                action="updated",
                summary=f"Excel updated fields: {', '.join(update_data.keys())}",
                source="Excel import",
                actor="Excel import",
            )
        if "responsible_person" in item.get("changed_fields", []):
            applied = apply_sync_assignment(asset_id, record, sync_context)
            assignment_updated += applied
            skipped_relationships += 0 if applied else 1
        if any(field_name in item.get("changed_fields", []) for field_name in ["project_number", "purchased_project_number", "transferred_project_number"]):
            applied = apply_sync_project(asset_id, record, sync_context)
            project_updated += applied
            skipped_relationships += 0 if applied else 1
        if "purchase_date_raw" in item.get("changed_fields", []) or "remarks" in item.get("changed_fields", []):
            applied = apply_sync_payments(asset_id, record)
            payment_updated += applied
            skipped_relationships += 0 if applied else 1

    for record in preview.get("transfer_log", {}).get("new_records", []):
        applied = apply_sync_transfer(record, sync_context)
        transfer_updated += applied
        skipped_relationships += 0 if applied else 1

    return {
        "inserted": inserted,
        "updated": updated,
        "assignment_updated": assignment_updated,
        "project_updated": project_updated,
        "payment_updated": payment_updated,
        "transfer_updated": transfer_updated,
        "skipped_relationships": skipped_relationships,
    }


def normalize_excel_header(value) -> str:
    return " ".join(str(value or "").split()).casefold()


def resolve_excel_header_field(header_value) -> Optional[str]:
    normalized = normalize_excel_header(header_value)
    expected_headers = {
        normalize_excel_header(excel_header): field_name
        for excel_header, field_name in EXCEL_SYNC_COLUMN_MAP.items()
    }
    if normalized in expected_headers:
        return expected_headers[normalized]

    if "asset tag" in normalized and ("inventory" in normalized or "code" in normalized):
        return "asset_tag_number"
    if normalized.startswith("previous inventory"):
        return "inventory_code_old"
    if normalized == "asset classification":
        return "asset_classification"
    if normalized == "asset sub classification":
        return "asset_sub_classification"
    if normalized == "item description":
        return "item_description"
    if normalized.startswith("brand"):
        return "brand_make"
    if normalized == "model":
        return "model"
    if normalized.startswith("serial"):
        return "serial_number"
    if normalized == "quantity":
        return "quantity"
    if normalized == "location":
        return "location_name"
    if normalized == "department":
        return "department_name"
    if normalized == "name of recipient":
        return "recipient_name"
    if normalized == "position of recipient":
        return "recipient_position"
    if normalized.startswith("date") and "purchase" in normalized:
        return "purchase_date_raw"
    if normalized == "purchase price":
        return "purchase_price"
    if normalized == "currency":
        return "currency"
    if normalized.startswith("purchased to proj"):
        return "purchased_project_no"
    if normalized == "donor":
        return "donor_name"
    if normalized.startswith("transferred to proj"):
        return "transferred_project_no"
    if normalized.startswith("current status"):
        return "current_status"
    if normalized == "remarks":
        return "remarks"
    if normalized.startswith("last date"):
        return "last_transfer_date"
    return None


def get_excel_header_columns(sheet) -> dict[str, list[int]]:
    header_row_number = EXCEL_SYNC_HEADER_ROW + 1
    columns = {}
    donor_column_count = 0

    for cell in sheet[header_row_number]:
        if normalize_excel_header(cell.value) == "donor":
            donor_column_count += 1
            field_name = "purchased_donor_name" if donor_column_count == 1 else "transferred_donor_name"
        else:
            field_name = resolve_excel_header_field(cell.value)
        if field_name:
            columns.setdefault(field_name, []).append(cell.column)

    return columns


def get_transfer_header_columns(sheet) -> dict[str, int]:
    header_row_number = EXCEL_TRANSFER_LOG_HEADER_ROW + 1
    columns = {}
    for cell in sheet[header_row_number]:
        field_name = resolve_transfer_header_field(cell.value)
        if field_name and field_name not in columns:
            columns[field_name] = cell.column
    return columns


def get_export_project_numbers(asset_projects: list[dict], mode: str) -> str:
    if mode == "purchased":
        allocations = get_project_allocations_from_rows(get_purchased_project_rows(asset_projects))
    else:
        allocations = get_project_allocations_from_rows(get_current_project_rows(asset_projects))
    project_numbers = [allocation.get("project_number") for allocation in allocations if allocation.get("project_number")]
    return "/".join(project_numbers)


def get_export_donor_name(asset_projects: list[dict], mode: str) -> Optional[str]:
    if mode == "purchased":
        selected_projects = get_purchased_project_rows(asset_projects)
    else:
        selected_projects = get_current_project_rows(asset_projects)
    for row in selected_projects:
        if row.get("donor_name"):
            return row.get("donor_name")
    return None


def get_export_transfer_date(asset_projects: list[dict]) -> Optional[str]:
    current_projects = [row for row in asset_projects if row.get("is_current") is True]
    selected_projects = current_projects or asset_projects
    dates = [row.get("transfer_date") for row in selected_projects if row.get("transfer_date")]
    return max(dates) if dates else None


def build_database_excel_records(usage_type: Optional[str] = None) -> list[dict]:
    assets = list_asset_records()
    sync_context = build_sync_context()
    records = []

    for asset in assets:
        asset_usage_type = normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number"))
        if usage_type and asset_usage_type != usage_type:
            continue
        asset_id = asset.get("asset_id")
        assignment = sync_context["assignment_by_asset_id"].get(asset_id) or {}
        asset_projects = sync_context["projects_by_asset_id"].get(asset_id, [])
        asset_payments = sync_context["payments_by_asset_id"].get(asset_id, [])
        purchased_project_numbers = get_export_project_numbers(asset_projects, "purchased")
        transferred_project_numbers = get_export_project_numbers(asset_projects, "transferred")

        records.append(
            {
                "asset_tag_number": asset.get("asset_tag_number"),
                "usage_type": asset_usage_type,
                "inventory_code_old": asset.get("inventory_code"),
                "asset_classification": asset.get("asset_classification"),
                "asset_sub_classification": asset.get("asset_sub_classification"),
                "item_description": asset.get("item_description"),
                "brand_make": asset.get("brand_make"),
                "model": asset.get("model"),
                "serial_number": asset.get("serial_chassis_number") or asset.get("serial_number"),
                "quantity": asset.get("quantity"),
                "location_name": assignment.get("city") or assignment.get("location_name"),
                "department_name": assignment.get("department"),
                "recipient_name": assignment.get("responsible_person"),
                "recipient_position": None,
                "purchase_date_raw": get_latest_payment_period(asset_payments),
                "purchase_price": asset.get("purchase_price"),
                "currency": asset.get("currency"),
                "purchased_project_no": purchased_project_numbers,
                "purchased_donor_name": get_export_donor_name(asset_projects, "purchased"),
                "transferred_project_no": transferred_project_numbers,
                "transferred_donor_name": get_export_donor_name(asset_projects, "transferred"),
                "current_status": assignment.get("status") or asset.get("current_status"),
                "remarks": asset.get("remarks"),
                "last_transfer_date": get_export_transfer_date(asset_projects) or assignment.get("assignment_date"),
            }
        )

    return records


def build_database_transfer_log_records() -> list[dict]:
    transfers = list_asset_transfer_records()
    if not transfers:
        return []

    assets_by_id = {row.get("asset_id"): row for row in list_asset_records()}
    people_by_id = {row.get("person_id"): row for row in list_people()}
    transfer_ids = [row.get("transfer_id") for row in transfers if row.get("transfer_id")]
    projects_by_transfer_id = get_asset_transfer_project_rows(transfer_ids)
    records = []

    for index, transfer in enumerate(
        sorted(transfers, key=lambda row: (str(row.get("transfer_date") or ""), int(row.get("transfer_id") or 0))),
        start=1,
    ):
        asset = assets_by_id.get(transfer.get("asset_id")) or {}
        from_person = people_by_id.get(transfer.get("from_person_id")) or {}
        to_person = people_by_id.get(transfer.get("to_person_id")) or {}
        project_rows = projects_by_transfer_id.get(transfer.get("transfer_id"), [])
        from_project = transfer.get("from_project_raw") or (format_transfer_project_rows(project_rows, "from") if project_rows else None)
        to_project = transfer.get("to_project_raw") or (format_transfer_project_rows(project_rows, "to") if project_rows else None)
        record = {
            "source_log_no": transfer.get("source_log_no"),
            "transfer_date": transfer.get("transfer_date"),
            "source_asset_type": transfer.get("source_asset_type") or get_asset_usage_type_label(asset.get("usage_type")),
            "asset_tag_number": transfer.get("asset_tag_snapshot") or asset.get("asset_tag_number"),
            "asset_tag_snapshot": transfer.get("asset_tag_snapshot") or asset.get("asset_tag_number"),
            "description_snapshot": transfer.get("description_snapshot") or asset.get("item_description"),
            "serial_snapshot": transfer.get("serial_snapshot") or asset.get("serial_chassis_number") or asset.get("serial_number"),
            "from_holder_name": get_person_display_name(from_person) if from_person else transfer.get("from_holder_name"),
            "from_project_raw": None if from_project == "-" else from_project,
            "to_project_raw": None if to_project == "-" else to_project,
            "to_holder_name": get_person_display_name(to_person) if to_person else transfer.get("to_holder_name"),
            "asset_status": transfer.get("asset_status"),
            "asset_condition_description": transfer.get("asset_condition_description"),
            "transfer_reason": transfer.get("transfer_reason") or transfer.get("notes"),
        }
        record["transfer_key"] = make_transfer_key(record)
        records.append(record)
    return records


def copy_excel_row_style(sheet, source_row: int, target_row: int) -> None:
    for column_index in range(1, sheet.max_column + 1):
        source_cell = sheet.cell(row=source_row, column=column_index)
        target_cell = sheet.cell(row=target_row, column=column_index)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def get_first_excel_column(columns: dict[str, list[int]], field_name: str) -> Optional[int]:
    column_numbers = columns.get(field_name) or []
    return column_numbers[0] if column_numbers else None


def write_excel_record(sheet, row_number: int, columns: dict[str, list[int]], record: dict) -> int:
    written = 0
    for field_name, column_numbers in columns.items():
        if field_name not in record:
            continue
        for column_number in column_numbers:
            sheet.cell(row=row_number, column=column_number).value = record.get(field_name)
            written += 1
    return written


def expand_excel_data_ranges(sheet, header_row_number: int, last_data_row: int) -> None:
    try:
        from openpyxl.utils import get_column_letter, range_boundaries  # type: ignore
    except Exception:
        return

    for table in sheet.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_row <= header_row_number <= max_row and last_data_row > max_row:
            table.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{last_data_row}"
            )

    auto_filter_ref = getattr(sheet.auto_filter, "ref", None)
    if auto_filter_ref:
        min_col, min_row, max_col, max_row = range_boundaries(auto_filter_ref)
        if min_row <= header_row_number <= max_row and last_data_row > max_row:
            sheet.auto_filter.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{last_data_row}"
            )


def write_database_records_to_excel_sheet(sheet, records: list[dict]) -> dict:
    columns = get_excel_header_columns(sheet)
    if "asset_tag_number" not in columns:
        raise ValueError(f"The workbook sheet '{sheet.title}' does not contain the expected Asset Tag column.")

    header_row_number = EXCEL_SYNC_HEADER_ROW + 1
    data_start_row = EXCEL_SYNC_HEADER_ROW + 2
    asset_tag_column = get_first_excel_column(columns, "asset_tag_number")
    if not asset_tag_column:
        raise ValueError(f"The workbook sheet '{sheet.title}' does not contain the expected Asset Tag column.")

    row_by_tag = {}
    for row_number in range(data_start_row, sheet.max_row + 1):
        asset_tag = normalize_asset_tag(sheet.cell(row=row_number, column=asset_tag_column).value or "")
        if asset_tag and asset_tag not in row_by_tag:
            row_by_tag[asset_tag] = row_number

    updated_rows = 0
    appended_rows = 0
    written_cells = 0
    last_data_row = max(row_by_tag.values(), default=data_start_row - 1)
    template_row = max(data_start_row, last_data_row)

    for record in records:
        asset_tag = normalize_asset_tag(record.get("asset_tag_number") or "")
        if not asset_tag:
            continue
        row_number = row_by_tag.get(asset_tag)
        if row_number:
            updated_rows += 1
        else:
            row_number = last_data_row + 1
            copy_excel_row_style(sheet, template_row, row_number)
            row_by_tag[asset_tag] = row_number
            last_data_row = row_number
            template_row = row_number
            appended_rows += 1
        written_cells += write_excel_record(sheet, row_number, columns, record)

    expand_excel_data_ranges(sheet, header_row_number, last_data_row)

    return {
        "updated_rows": updated_rows,
        "appended_rows": appended_rows,
        "written_cells": written_cells,
        "exported_records": len(records),
    }


def write_transfer_log_records_to_excel_sheet(sheet, records: list[dict]) -> dict:
    columns = get_transfer_header_columns(sheet)
    if "asset_tag_number" not in columns or "transfer_date" not in columns:
        raise ValueError(f"The workbook sheet '{sheet.title}' does not contain the expected Transfer log headers.")

    header_row_number = EXCEL_TRANSFER_LOG_HEADER_ROW + 1
    data_start_row = EXCEL_TRANSFER_LOG_HEADER_ROW + 2
    row_by_key = {}
    max_source_log_no = 0

    for row_number in range(data_start_row, sheet.max_row + 1):
        row = {
            field_name: sheet.cell(row=row_number, column=column_number).value
            for field_name, column_number in columns.items()
        }
        normalized = normalize_excel_transfer_record(row, row_number)
        if normalized:
            row_by_key.setdefault(normalized.get("transfer_key"), row_number)
            max_source_log_no = max(max_source_log_no, int(normalized.get("source_log_no") or 0))

    last_data_row = max(row_by_key.values(), default=data_start_row - 1)
    template_row = max(data_start_row, last_data_row)
    updated_rows = 0
    appended_rows = 0
    written_cells = 0

    for record in records:
        transfer_key = record.get("transfer_key") or make_transfer_key(record)
        row_number = row_by_key.get(transfer_key)
        if row_number:
            updated_rows += 1
        else:
            row_number = last_data_row + 1
            copy_excel_row_style(sheet, template_row, row_number)
            row_by_key[transfer_key] = row_number
            last_data_row = row_number
            template_row = row_number
            appended_rows += 1
            if not record.get("source_log_no"):
                max_source_log_no += 1
                record["source_log_no"] = max_source_log_no

        for field_name, column_number in columns.items():
            if field_name in record:
                sheet.cell(row=row_number, column=column_number).value = record.get(field_name)
                written_cells += 1

    expand_excel_data_ranges(sheet, header_row_number, last_data_row)
    return {
        "updated_rows": updated_rows,
        "appended_rows": appended_rows,
        "written_cells": written_cells,
        "exported_records": len(records),
    }


def export_supabase_to_excel() -> dict:
    if not os.path.exists(SYNC_WORKBOOK_PATH):
        raise ValueError("No official workbook is available. Upload an Excel file first.")

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise ValueError(f"Excel export requires openpyxl support: {exc}") from exc

    workbook = load_workbook(SYNC_WORKBOOK_PATH)
    if EXCEL_SYNC_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"The workbook does not contain the expected sheet '{EXCEL_SYNC_SHEET_NAME}'.")

    standard_result = write_database_records_to_excel_sheet(
        workbook[EXCEL_SYNC_SHEET_NAME],
        build_database_excel_records("standard"),
    )
    low_cost_result = {
        "updated_rows": 0,
        "appended_rows": 0,
        "written_cells": 0,
        "exported_records": 0,
    }
    if EXCEL_LOW_COST_SHEET_NAME in workbook.sheetnames:
        low_cost_result = write_database_records_to_excel_sheet(
            workbook[EXCEL_LOW_COST_SHEET_NAME],
            build_database_excel_records("low_cost"),
        )
    transfer_result = {
        "updated_rows": 0,
        "appended_rows": 0,
        "written_cells": 0,
        "exported_records": 0,
    }
    if EXCEL_TRANSFER_LOG_SHEET_NAME in workbook.sheetnames:
        transfer_result = write_transfer_log_records_to_excel_sheet(
            workbook[EXCEL_TRANSFER_LOG_SHEET_NAME],
            build_database_transfer_log_records(),
        )

    ensure_sync_storage()
    workbook.save(SYNC_EXPORT_PATH)

    return {
        "path": SYNC_EXPORT_PATH,
        "updated_rows": standard_result["updated_rows"] + low_cost_result["updated_rows"] + transfer_result["updated_rows"],
        "appended_rows": standard_result["appended_rows"] + low_cost_result["appended_rows"] + transfer_result["appended_rows"],
        "written_cells": standard_result["written_cells"] + low_cost_result["written_cells"] + transfer_result["written_cells"],
        "exported_records": standard_result["exported_records"] + low_cost_result["exported_records"] + transfer_result["exported_records"],
        "standard_exported_records": standard_result["exported_records"],
        "low_cost_exported_records": low_cost_result["exported_records"],
        "transfer_exported_records": transfer_result["exported_records"],
        "exported_at": datetime.now(ZoneInfo("Europe/Kyiv")).strftime("%d.%m.%Y %H:%M"),
    }


def filter_sync_preview(preview: dict, selected_new_assets: list[str], selected_changed_assets: list[str], selected_transfers: Optional[list[str]] = None) -> dict:
    selected_new_tags = {normalize_asset_tag(value) for value in selected_new_assets if value}
    selected_changed_tags = {normalize_asset_tag(value) for value in selected_changed_assets if value}
    selected_transfer_keys = {value for value in (selected_transfers or []) if value}

    filtered_new_records = [
        record
        for record in preview.get("new_records", [])
        if normalize_asset_tag(record.get("asset_tag_number") or "") in selected_new_tags
    ]
    filtered_changed_records = [
        item
        for item in preview.get("changed_records", [])
        if normalize_asset_tag(item.get("asset_tag_number") or "") in selected_changed_tags
    ]
    transfer_log = preview.get("transfer_log") or {}
    filtered_transfer_records = [
        record
        for record in transfer_log.get("new_records", [])
        if record.get("transfer_key") in selected_transfer_keys
    ]

    summary = preview.get("summary", {})
    transfer_summary = transfer_log.get("summary", {})
    unchanged_records = summary.get("unchanged_records", 0)
    excel_rows = summary.get("excel_rows", unchanged_records + len(filtered_new_records) + len(filtered_changed_records))

    return {
        "summary": {
            "excel_rows": excel_rows,
            "standard_rows": summary.get("standard_rows", 0),
            "low_cost_rows": summary.get("low_cost_rows", 0),
            "new_records": len(filtered_new_records),
            "changed_records": len(filtered_changed_records),
            "unchanged_records": unchanged_records,
        },
        "new_records": filtered_new_records,
        "changed_records": filtered_changed_records,
        "transfer_log": {
            "summary": {
                **transfer_summary,
                "new_records": len(filtered_transfer_records),
            },
            "new_records": filtered_transfer_records,
        },
    }


def get_current_assignment(asset_id: int) -> Optional[dict]:
    assignment_response = (
        supabase.table("asset_assignments")
        .select("*")
        .eq("asset_id", asset_id)
        .is_("return_date", "null")
        .order("assignment_date", desc=True)
        .limit(1)
        .execute()
    )

    if not assignment_response.data:
        return None

    assignment = assignment_response.data[0]
    person = None
    location = None

    person_id = assignment.get("person_id")
    if person_id:
        person_response = (
            supabase.table("persons")
            .select("*")
            .eq("person_id", person_id)
            .limit(1)
            .execute()
        )
        if person_response.data:
            person = person_response.data[0]

    location_id = assignment.get("location_id")
    if location_id:
        location_response = (
            supabase.table("locations")
            .select("*")
            .eq("location_id", location_id)
            .limit(1)
            .execute()
        )
        if location_response.data:
            location = location_response.data[0]

    return {
        "assignment_id": assignment.get("assignment_id"),
        "person_id": assignment.get("person_id"),
        "location_id": assignment.get("location_id"),
        "assignment_date": assignment.get("assignment_date"),
        "return_date": assignment.get("return_date"),
        "status": assignment.get("status"),
        "notes": assignment.get("notes"),
        "handover_condition": assignment.get("handover_condition"),
        "responsible_person": (
            person.get("name_eng")
            or person.get("name")
            or person.get("full_name")
            if person
            else None
        ),
        "department": (
            (person.get("department") if person else None)
            or (location.get("department") if location else None)
        ),
        "city": location.get("city") if location else None,
        "location_name": location.get("name") if location else None,
    }


def enrich_assignment(assignment: dict) -> dict:
    person = None
    location = None

    person_id = assignment.get("person_id")
    if person_id:
        person_response = (
            supabase.table("persons")
            .select("*")
            .eq("person_id", person_id)
            .limit(1)
            .execute()
        )
        if person_response.data:
            person = person_response.data[0]

    location_id = assignment.get("location_id")
    if location_id:
        location_response = (
            supabase.table("locations")
            .select("*")
            .eq("location_id", location_id)
            .limit(1)
            .execute()
        )
        if location_response.data:
            location = location_response.data[0]

    return {
        "assignment_id": assignment.get("assignment_id"),
        "person_id": assignment.get("person_id"),
        "location_id": assignment.get("location_id"),
        "assignment_date": assignment.get("assignment_date"),
        "return_date": assignment.get("return_date"),
        "status": assignment.get("status"),
        "notes": assignment.get("notes"),
        "handover_condition": assignment.get("handover_condition"),
        "responsible_person": (
            person.get("name_eng")
            or person.get("name")
            or person.get("full_name")
            if person
            else None
        ),
        "department": (
            (person.get("department") if person else None)
            or (location.get("department") if location else None)
        ),
        "city": location.get("city") if location else None,
        "location_name": location.get("name") if location else None,
    }


def get_asset_by_tag(asset_tag: str) -> Optional[dict]:
    query = (
        supabase.table("assets")
        .select("*")
        .eq("asset_tag_number", asset_tag)
        .limit(1)
    )
    response = execute_supabase_query(query, "get_asset_by_tag")
    if not response.data:
        return None

    asset = response.data[0]
    asset["usage_type"] = normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number"))
    asset["usage_type_label"] = get_asset_usage_type_label(asset.get("usage_type"))
    asset["current_assignment"] = get_current_assignment(asset["asset_id"])
    return asset


def get_asset_by_id(asset_id: int) -> Optional[dict]:
    response = (
        supabase.table("assets")
        .select("*")
        .eq("asset_id", asset_id)
        .limit(1)
        .execute()
    )
    if not response.data:
        return None

    asset = response.data[0]
    asset["usage_type"] = normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number"))
    asset["usage_type_label"] = get_asset_usage_type_label(asset.get("usage_type"))
    asset["current_assignment"] = get_current_assignment(asset["asset_id"])
    asset["effective_status"] = get_effective_status(asset)
    return asset


def get_assignment_history(asset_id: int, limit: int = 20) -> list[dict]:
    response = (
        supabase.table("asset_assignments")
        .select("*")
        .eq("asset_id", asset_id)
        .order("assignment_date", desc=True)
        .limit(limit)
        .execute()
    )

    assignments = response.data or []
    return [enrich_assignment(assignment) for assignment in assignments]


def list_people() -> list[dict]:
    response = (
        supabase.table("persons")
        .select("*")
        .order("name_eng")
        .execute()
    )
    return response.data or []


def list_projects() -> list[dict]:
    response = (
        supabase.table("projects")
        .select("*")
        .order("project_number")
        .execute()
    )
    return response.data or []


def list_donors() -> list[dict]:
    response = (
        supabase.table("donors")
        .select("*")
        .order("donor_name")
        .execute()
    )
    return response.data or []


def get_next_numeric_id(table_name: str, id_column: str) -> int:
    response = (
        supabase.table(table_name)
        .select(id_column)
        .order(id_column, desc=True)
        .limit(1)
        .execute()
    )
    if not response.data:
        return 1
    current_max = response.data[0].get(id_column) or 0
    return int(current_max) + 1


def get_project_by_id(project_id: int) -> Optional[dict]:
    response = (
        supabase.table("projects")
        .select("*")
        .eq("project_id", project_id)
        .limit(1)
        .execute()
    )
    if not response.data:
        return None
    return response.data[0]


def get_donor_by_id(donor_id: int) -> Optional[dict]:
    response = (
        supabase.table("donors")
        .select("*")
        .eq("donor_id", donor_id)
        .limit(1)
        .execute()
    )
    if not response.data:
        return None
    return response.data[0]


def get_project_form_values(project: Optional[dict] = None) -> dict:
    project = project or {}
    return {
        "project_number": project.get("project_number") or "",
        "project_name": project.get("project_name") or project.get("name") or "",
        "start_date": project.get("start_date") or "",
        "end_date": project.get("end_date") or "",
        "status": project.get("status") or "",
    }


def get_donor_form_values(donor: Optional[dict] = None) -> dict:
    donor = donor or {}
    return {
        "donor_name": donor.get("donor_name") or "",
        "contact_person": donor.get("contact_person") or "",
        "contact_email": donor.get("contact_email") or "",
    }


def describe_reference_data_error(error: Exception, entity_label: str, unique_field: str) -> str:
    if not isinstance(error, APIError):
        return f"{entity_label} could not be saved due to an unexpected database error."

    message = error.message or "Database error"
    details = error.details or ""
    combined = " ".join(part for part in [message, details] if part).lower()

    if unique_field in combined and ("duplicate" in combined or "unique" in combined):
        field_label = "Project number" if unique_field == "project_number" else "Donor name"
        return f"{field_label} already exists."

    if entity_label == "Project" and "projects_pkey" in combined:
        return "Project could not be saved because the database project_id sequence is out of sync."

    if entity_label == "Donor" and "donors_pkey" in combined:
        return "Donor could not be saved because the database donor_id sequence is out of sync."

    return f"{entity_label} could not be saved: {message}"


def safe_parse_percentage(value: str) -> Optional[float]:
    normalized = (value or "").strip().replace(",", ".")
    if not normalized:
        return None
    parsed = float(normalized)
    if parsed < 0:
        raise ValueError("Allocation percent cannot be negative.")
    return parsed


def describe_asset_project_error(error: Exception) -> str:
    if not isinstance(error, APIError):
        return "Project funding could not be saved due to an unexpected database error."

    message = error.message or "Database error"
    details = error.details or ""
    combined = " ".join(part for part in [message, details] if part).lower()

    if "allocation_percent" in combined and "column" in combined:
        return "The database schema is missing allocation_percent for asset project funding."
    if "allocation_amount" in combined and "column" in combined:
        return "The database schema is missing allocation_amount for asset project funding."
    if "funding_note" in combined and "column" in combined:
        return "The database schema is missing funding_note for asset project funding."
    if "is_purchase_origin" in combined and "column" in combined:
        return "The database schema is missing is_purchase_origin for asset project funding."

    return f"Project funding could not be saved: {message}"


def get_asset_projects(asset_id: int) -> list[dict]:
    response = (
        supabase.table("asset_projects")
        .select("*")
        .eq("asset_id", asset_id)
        .order("is_primary", desc=True)
        .order("asset_project_id")
        .execute()
    )
    rows = response.data or []
    projects_by_id = {row.get("project_id"): row for row in list_projects()}
    donors_by_id = {row.get("donor_id"): row for row in list_donors()}

    enriched = []
    for row in rows:
        project = projects_by_id.get(row.get("project_id")) or {}
        donor = donors_by_id.get(row.get("donor_id")) or {}
        enriched.append(
            {
                **row,
                "project_number": project.get("project_number") or f"Project #{row.get('project_id')}",
                "project_name": project.get("project_name") or project.get("name") or "",
                "donor_name": donor.get("donor_name") or "",
            }
        )
    return enriched


def get_asset_project_total_percent(asset_id: int, exclude_asset_project_id: Optional[int] = None) -> float:
    total = 0.0
    for row in get_asset_projects(asset_id):
        if exclude_asset_project_id is not None and row.get("asset_project_id") == exclude_asset_project_id:
            continue
        if row.get("is_current") is False and row.get("is_purchase_origin") is True:
            continue
        value = row.get("allocation_percent")
        if value is not None:
            try:
                total += float(value)
            except Exception:
                continue
    return total


def get_asset_project_form_context(asset_id: int) -> dict:
    asset_projects = get_asset_projects(asset_id)
    current_project_rows = get_current_project_rows(asset_projects)
    purchased_project_rows = get_purchased_project_rows(asset_projects)
    allocated_percent = sum(float(row.get("allocation_percent") or 0) for row in current_project_rows)
    return {
        "asset_projects": asset_projects,
        "projects": list_projects(),
        "donors": list_donors(),
        "asset_project_summary": {
            "allocation_count": len(asset_projects),
            "purchase_origin_count": len(purchased_project_rows),
            "current_count": len(current_project_rows),
            "allocated_percent": round(allocated_percent, 2),
            "is_complete": abs(allocated_percent - 100.0) < 0.001 if current_project_rows else False,
        },
        "asset_project_roles": {
            "purchased": purchased_project_rows,
            "current": current_project_rows,
        },
    }


def get_asset_transfer_project_rows(transfer_ids: list[int]) -> dict[int, list[dict]]:
    if not transfer_ids:
        return {}

    try:
        response = (
            supabase.table("asset_transfer_projects")
            .select("*")
            .in_("transfer_id", transfer_ids)
            .order("direction")
            .order("transfer_project_id")
            .execute()
        )
    except Exception:
        return {}
    rows = response.data or []
    projects_by_id = {row.get("project_id"): row for row in list_projects()}
    grouped: dict[int, list[dict]] = {}
    for row in rows:
        project = projects_by_id.get(row.get("project_id")) or {}
        grouped.setdefault(row.get("transfer_id"), []).append(
            {
                **row,
                "project_number": project.get("project_number") or row.get("project_number_raw"),
                "project_name": project.get("project_name") or project.get("name") or "",
            }
        )
    return grouped


def format_transfer_project_rows(rows: list[dict], direction: str) -> str:
    selected = [row for row in rows if row.get("direction") == direction]
    if not selected:
        return "-"
    parts = []
    for row in selected:
        project_number = row.get("project_number") or row.get("project_number_raw") or "-"
        allocation_percent = row.get("allocation_percent")
        if allocation_percent is None:
            parts.append(project_number)
        else:
            try:
                parts.append(f"{project_number} {float(allocation_percent):g}%")
            except Exception:
                parts.append(project_number)
    return " / ".join(parts)


def create_asset_transfer_from_assignment_change(
    *,
    asset: dict,
    from_assignment: Optional[dict],
    to_person_id: Optional[int],
    transfer_date: str,
    transfer_reason: str,
    status: Optional[str] = None,
    condition: Optional[str] = None,
) -> Optional[int]:
    asset_id = asset.get("asset_id")
    if not asset_id or not transfer_date:
        return None

    from_assignment = from_assignment or {}
    from_person_id = from_assignment.get("person_id")
    from_holder_name = from_assignment.get("responsible_person")
    if not from_holder_name and from_assignment.get("location_id"):
        from_holder_name = "Warehouse"

    to_person = get_person_by_id(to_person_id) if to_person_id else None
    to_holder_name = get_person_display_name(to_person) if to_person else "Warehouse"
    current_project_rows = get_current_project_rows(get_asset_projects(asset_id))
    current_project_raw = format_project_allocations(get_project_allocations_from_rows(current_project_rows))

    payload = {
        "asset_id": asset_id,
        "from_person_id": from_person_id,
        "to_person_id": to_person_id,
        "transfer_date": transfer_date,
        "transfer_reason": transfer_reason,
        "notes": transfer_reason,
        "source_asset_type": get_asset_usage_type_label(asset.get("usage_type")),
        "asset_tag_snapshot": asset.get("asset_tag_number"),
        "description_snapshot": asset.get("item_description"),
        "serial_snapshot": asset.get("serial_chassis_number") or asset.get("serial_number"),
        "from_holder_name": from_holder_name,
        "to_holder_name": to_holder_name,
        "from_project_raw": current_project_raw,
        "to_project_raw": current_project_raw,
        "asset_status": status or from_assignment.get("status") or asset.get("current_status"),
        "asset_condition_description": condition or from_assignment.get("handover_condition"),
    }

    try:
        response = supabase.table("asset_transfers").insert(payload).execute()
    except Exception:
        return None

    transfer = (response.data or [{}])[0]
    transfer_id = transfer.get("transfer_id")
    if transfer_id and current_project_rows:
        project_payloads = []
        for direction in ["from", "to"]:
            for row in current_project_rows:
                project_payloads.append(
                    {
                        "transfer_id": transfer_id,
                        "direction": direction,
                        "project_id": row.get("project_id"),
                        "project_number_raw": row.get("project_number"),
                        "allocation_percent": row.get("allocation_percent"),
                    }
                )
        try:
            supabase.table("asset_transfer_projects").insert(project_payloads).execute()
        except Exception:
            pass

    return transfer_id


def get_asset_transfer_history(asset_id: int, limit: int = 50) -> list[dict]:
    try:
        response = (
            supabase.table("asset_transfers")
            .select("*")
            .eq("asset_id", asset_id)
            .order("transfer_date", desc=True)
            .order("transfer_id", desc=True)
            .limit(limit)
            .execute()
        )
    except Exception:
        return []
    transfers = response.data or []
    transfer_ids = [row.get("transfer_id") for row in transfers if row.get("transfer_id")]
    projects_by_transfer_id = get_asset_transfer_project_rows(transfer_ids)

    people_by_id = {row.get("person_id"): row for row in list_people()}
    for transfer in transfers:
        from_person = people_by_id.get(transfer.get("from_person_id")) or {}
        to_person = people_by_id.get(transfer.get("to_person_id")) or {}
        project_rows = projects_by_transfer_id.get(transfer.get("transfer_id"), [])
        transfer["from_holder_display"] = (
            get_person_display_name(from_person)
            if from_person
            else transfer.get("from_holder_name") or "-"
        )
        transfer["to_holder_display"] = (
            get_person_display_name(to_person)
            if to_person
            else transfer.get("to_holder_name") or "-"
        )
        transfer["from_project_display"] = format_transfer_project_rows(project_rows, "from") if project_rows else (transfer.get("from_project_raw") or "-")
        transfer["to_project_display"] = format_transfer_project_rows(project_rows, "to") if project_rows else (transfer.get("to_project_raw") or "-")
    return transfers


def get_asset_transfer_context(asset_id: int) -> dict:
    transfers = get_asset_transfer_history(asset_id)
    return {
        "asset_transfers": transfers,
        "asset_transfer_summary": {
            "transfer_count": len(transfers),
            "latest_transfer_date": transfers[0].get("transfer_date") if transfers else None,
        },
    }


def get_asset_payments(asset_id: int) -> list[dict]:
    response = (
        supabase.table("asset_payments")
        .select("*")
        .eq("asset_id", asset_id)
        .order("payment_date")
        .order("payment_number")
        .execute()
    )
    rows = response.data or []
    for row in rows:
        row["purchase_period"] = format_purchase_period(row.get("payment_date"))
    return rows


def get_asset_payment_context(asset_id: int) -> dict:
    payments = get_asset_payments(asset_id)
    return {
        "asset_payments": payments,
        "asset_payment_summary": {
            "payment_count": len(payments),
            "purchase_period": get_latest_payment_period(payments),
        },
    }


def get_person_display_name(person: dict) -> str:
    return (
        person.get("name_eng")
        or person.get("name")
        or person.get("full_name")
        or f"Person #{person.get('person_id')}"
    )


def get_person_report_name(person: dict) -> str:
    local_name = (person.get("name") or person.get("full_name") or "").strip()
    english_name = (person.get("name_eng") or "").strip()

    if local_name and english_name and local_name.casefold() != english_name.casefold():
        return f"{local_name} / {english_name}"

    return english_name or local_name or f"Person #{person.get('person_id')}"


def is_person_active(person: dict) -> bool:
    return person.get("is_active") is not False


def get_person_status_label(person: dict) -> str:
    return "Active" if is_person_active(person) else "Inactive"


def get_person_by_id(person_id: int) -> Optional[dict]:
    response = (
        supabase.table("persons")
        .select("*")
        .eq("person_id", person_id)
        .limit(1)
        .execute()
    )
    if not response.data:
        return None
    return response.data[0]


def find_warehouse_person(people: Optional[list[dict]] = None) -> Optional[dict]:
    people = people if people is not None else list_people()
    for person in people:
        names = [
            person.get("name"),
            person.get("name_eng"),
            person.get("full_name"),
        ]
        if any((name or "").strip().casefold() == "warehouse" for name in names):
            return person
    return None


def get_location_display_name(location: dict) -> str:
    parts = [
        location.get("city"),
        location.get("department"),
        location.get("office"),
        location.get("building"),
        location.get("room"),
    ]
    return " / ".join(str(part) for part in parts if part) or f"Location #{location.get('location_id')}"


def get_default_offboarding_location_id(asset: dict, locations: list[dict]) -> Optional[int]:
    current_assignment = asset.get("current_assignment") or {}
    if current_assignment.get("location_id"):
        return current_assignment.get("location_id")

    for location in locations:
        if (
            (location.get("city") or "").strip().casefold() == "kyiv"
            and (location.get("department") or "").strip().casefold() == "administration"
        ):
            return location.get("location_id")

    return locations[0].get("location_id") if locations else None


def get_offboarding_target_people(person_id: int) -> list[dict]:
    people = list_people()
    return [
        {
            **person,
            "display_name": get_person_display_name(person),
            "status_label": get_person_status_label(person),
        }
        for person in people
        if person.get("person_id") != person_id and is_person_active(person)
    ]


def build_person_offboarding_context(person: dict) -> dict:
    assigned_assets = get_assets_for_person(person["person_id"])
    locations = [
        {**location, "display_name": get_location_display_name(location)}
        for location in list_locations()
    ]
    people = get_offboarding_target_people(person["person_id"])
    warehouse_person = find_warehouse_person(people)
    warehouse_person_id = warehouse_person.get("person_id") if warehouse_person else None

    rows = []
    for asset in assigned_assets:
        assignment = asset.get("current_assignment") or {}
        rows.append(
            {
                "asset": asset,
                "assignment": assignment,
                "default_person_id": warehouse_person_id or "",
                "default_location_id": get_default_offboarding_location_id(asset, locations) or "",
            }
        )

    return {
        "assigned_assets": assigned_assets,
        "offboarding_rows": rows,
        "target_people": people,
        "locations": locations,
        "warehouse_person": warehouse_person,
        "offboarding_date": datetime.now(ZoneInfo("Europe/Kyiv")).date().isoformat(),
    }


def update_person_inactive(person_id: int, offboarding_date: str, note: str) -> None:
    supabase.table("persons").update({"is_active": False}).eq("person_id", person_id).execute()

    try:
        supabase.table("persons").update(
            {
                "offboarded_at": offboarding_date,
                "offboarding_note": note or None,
            }
        ).eq("person_id", person_id).execute()
    except Exception as error:
        if isinstance(error, APIError):
            combined = " ".join(part for part in [error.message or "", error.details or ""] if part).lower()
            if "offboarded_at" in combined or "offboarding_note" in combined:
                return
        raise


def apply_person_offboarding(person: dict, form) -> dict:
    person_id = person["person_id"]
    offboarding_date = str(form.get("offboarding_date") or "").strip()
    offboarding_note = str(form.get("offboarding_note") or "").strip()
    if not offboarding_date:
        raise ValueError("Offboarding date is required.")

    assigned_assets = get_assets_for_person(person_id)
    selected_asset_ids = {
        int(value)
        for value in form.getlist("asset_id")
        if str(value).strip().isdigit()
    }
    if not selected_asset_ids:
        selected_asset_ids = {asset["asset_id"] for asset in assigned_assets}

    target_people = {person["person_id"]: person for person in get_offboarding_target_people(person_id)}
    locations = {location["location_id"]: location for location in list_locations()}
    reassigned = 0
    skipped = 0

    for asset in assigned_assets:
        asset_id = asset.get("asset_id")
        if asset_id not in selected_asset_ids:
            skipped += 1
            continue

        person_value = str(form.get(f"target_person_id_{asset_id}") or "").strip()
        location_value = str(form.get(f"target_location_id_{asset_id}") or "").strip()
        target_person_id = int(person_value) if person_value.isdigit() else None
        target_location_id = int(location_value) if location_value.isdigit() else None

        if target_person_id == person_id:
            raise ValueError("An offboarded employee cannot remain responsible for assigned assets.")
        if target_person_id and target_person_id not in target_people:
            raise ValueError("Selected responsible person is not available for offboarding.")
        if target_location_id and target_location_id not in locations:
            raise ValueError("Selected location is not available for offboarding.")
        if target_person_id and not target_location_id:
            raise ValueError("Location is required when a responsible person is selected.")

        close_current_assignments(asset_id, offboarding_date)

        if target_person_id or target_location_id:
            current_assignment = asset.get("current_assignment") or {}
            status = current_assignment.get("status") or asset.get("current_status")
            notes = "Reassigned during employee offboarding"
            if offboarding_note:
                notes = f"{notes}: {offboarding_note}"
            supabase.table("asset_assignments").insert(
                {
                    "asset_id": asset_id,
                    "person_id": target_person_id,
                    "location_id": target_location_id,
                    "assignment_date": offboarding_date,
                    "return_date": None,
                    "status": status,
                    "notes": notes,
                    "handover_condition": current_assignment.get("handover_condition"),
                }
            ).execute()
            transfer_id = create_asset_transfer_from_assignment_change(
                asset=asset,
                from_assignment=asset.get("current_assignment"),
                to_person_id=target_person_id,
                transfer_date=offboarding_date,
                transfer_reason="Employee offboarding",
                status=(asset.get("current_assignment") or {}).get("status") or asset.get("current_status"),
                condition=(asset.get("current_assignment") or {}).get("handover_condition"),
            )
            target_person = target_people.get(target_person_id) if target_person_id else None
            target_label = get_person_display_name(target_person) if target_person else "No responsible person"
            audit_log_event(
                entity_type="Transfer",
                entity_id=transfer_id or asset_id,
                entity_label=asset.get("asset_tag_number"),
                action="transferred",
                field_name="responsible_person",
                old_value=get_person_display_name(person),
                new_value=target_label,
                summary=f"{asset.get('asset_tag_number')} reassigned during offboarding: {get_person_display_name(person)} -> {target_label}",
                source="Offboarding",
                actor="Offboarding",
                event_date=offboarding_date,
                event_key=f"asset_transfer:{transfer_id}" if transfer_id else None,
            )
        reassigned += 1

    update_person_inactive(person_id, offboarding_date, offboarding_note)
    audit_log_event(
        entity_type="Employee",
        entity_id=person_id,
        entity_label=get_person_display_name(person),
        action="offboarded",
        field_name="is_active",
        old_value=True,
        new_value=False,
        summary=f"Employee offboarded: {get_person_display_name(person)}",
        source="Offboarding",
        actor="Offboarding",
    )
    return {"reassigned": reassigned, "skipped": skipped}


PERSON_FIELD_LABELS = {
    "person_id": "Person ID",
    "name": "Employee name (local)",
    "name_eng": "Employee name (English)",
    "full_name": "Full name",
    "position": "Position",
    "department": "Department",
    "is_active": "Active employee",
    "offboarded_at": "Offboarded at",
    "offboarding_note": "Offboarding note",
    "email": "Email",
    "phone": "Phone",
    "telegram": "Telegram",
    "notes": "Notes",
    "comment": "Comment",
}

PERSON_FIELD_ORDER = [
    "person_id",
    "name",
    "name_eng",
    "full_name",
    "position",
    "department",
    "is_active",
    "offboarded_at",
    "offboarding_note",
    "email",
    "phone",
    "telegram",
    "notes",
    "comment",
]

READONLY_PERSON_FIELDS = {"person_id"}


def format_person_field_label(field_name: str) -> str:
    return PERSON_FIELD_LABELS.get(field_name, field_name.replace("_", " ").title())


def get_person_field_names(person: dict) -> list[str]:
    keys = list(person.keys())
    ordered = [field for field in PERSON_FIELD_ORDER if field in keys]
    remaining = sorted(field for field in keys if field not in ordered)
    return ordered + remaining


def build_person_field_rows(person: dict) -> list[dict]:
    rows = []
    for field_name in get_person_field_names(person):
        rows.append(
            {
                "name": field_name,
                "label": format_person_field_label(field_name),
                "value": person.get(field_name),
                "readonly": field_name in READONLY_PERSON_FIELDS,
            }
        )
    return rows


def build_person_edit_fields(person: dict) -> list[dict]:
    fields = []
    for row in build_person_field_rows(person):
        if row["readonly"]:
            continue
        value = row["value"]
        if row["name"] == "is_active":
            fields.append(
                {
                    "name": row["name"],
                    "label": row["label"],
                    "value": "false" if value is False else "true",
                    "multiline": False,
                    "input_type": "text",
                    "control": "select",
                    "options": [
                        {"value": "true", "label": "Active (TRUE)"},
                        {"value": "false", "label": "Inactive (FALSE)"},
                    ],
                }
            )
            continue

        as_text = "" if value is None else str(value)
        multiline = (
            "\n" in as_text
            or row["name"] in {"notes", "comment"}
            or len(as_text) > 120
        )
        input_type = "email" if row["name"] == "email" else "text"
        if row["name"] == "offboarded_at":
            input_type = "date"
        fields.append(
            {
                "name": row["name"],
                "label": row["label"],
                "value": as_text,
                "multiline": multiline,
                "input_type": input_type,
                "control": "input",
            }
        )
    return fields


def get_person_form_values(values: Optional[dict] = None) -> dict:
    values = values or {}
    return {
        "name": values.get("name", ""),
        "name_eng": values.get("name_eng", ""),
        "department": values.get("department", ""),
    }


def find_existing_person(person_form: dict) -> Optional[dict]:
    local_name = (person_form.get("name") or "").strip()
    english_name = (person_form.get("name_eng") or "").strip()

    candidates = []
    if local_name:
        candidates.append(("name", local_name))
    if english_name:
        candidates.append(("name_eng", english_name))
    if local_name:
        candidates.append(("name_eng", local_name))
    if english_name:
        candidates.append(("name", english_name))

    for field_name, value in candidates:
        try:
            response = (
                supabase.table("persons")
                .select("*")
                .eq(field_name, value)
                .limit(1)
                .execute()
            )
        except Exception:
            continue
        if response.data:
            return response.data[0]

    def normalize_person_lookup(value: Optional[str]) -> str:
        return " ".join((value or "").strip().casefold().split())

    normalized_candidates = [normalize_person_lookup(local_name), normalize_person_lookup(english_name)]
    normalized_candidates = [value for value in normalized_candidates if value]

    if normalized_candidates:
        try:
            for person in list_people():
                person_names = [
                    normalize_person_lookup(person.get("name")),
                    normalize_person_lookup(person.get("name_eng")),
                    normalize_person_lookup(person.get("full_name")),
                ]
                if any(candidate and candidate in person_names for candidate in normalized_candidates):
                    return person
        except Exception:
            pass

    return None


def get_next_person_id() -> int:
    response = (
        supabase.table("persons")
        .select("person_id")
        .order("person_id", desc=True)
        .limit(1)
        .execute()
    )
    rows = response.data or []
    if not rows:
        return 1
    return int(rows[0].get("person_id") or 0) + 1


def is_person_id_sequence_conflict(error: Exception) -> bool:
    if not isinstance(error, APIError):
        return False
    message = (error.message or "").lower()
    details = (error.details or "").lower()
    combined = f"{message} {details}"
    return "person_id" in combined and "already exists" in combined


def describe_person_create_error(error: Exception) -> str:
    if not isinstance(error, APIError):
        return "Employee could not be created due to an unexpected database error."

    message = error.message or "Database error"
    details = error.details or ""
    combined = " ".join(part for part in [message, details] if part).lower()

    if is_person_id_sequence_conflict(error):
        return "The employee ID sequence in the database is out of sync. The application could not assign a new person_id automatically."
    if "duplicate" in combined or "unique" in combined:
        if details:
            return f"An employee with the same unique data already exists. Database details: {details}"
        return f"An employee with the same unique data already exists. Database message: {message}"
    if "name" in combined and "null value" in combined:
        return "The database requires an employee name."

    return f"Employee could not be created: {message}"


def list_locations() -> list[dict]:
    response = (
        supabase.table("locations")
        .select("*")
        .order("city")
        .execute()
    )
    return response.data or []


def close_current_assignments(asset_id: int, return_date: Optional[str]) -> None:
    current_assignments = (
        supabase.table("asset_assignments")
        .select("assignment_id")
        .eq("asset_id", asset_id)
        .is_("return_date", "null")
        .execute()
    )

    for row in current_assignments.data or []:
        (
            supabase.table("asset_assignments")
            .update({"return_date": return_date})
            .eq("assignment_id", row["assignment_id"])
            .execute()
        )


def get_assignment_form_context(asset: dict) -> dict:
    current_assignment = asset.get("current_assignment") or {}
    return {
        "people": list_people(),
        "locations": list_locations(),
        "assignment_form": {
            "person_id": current_assignment.get("person_id") or "",
            "location_id": current_assignment.get("location_id") or "",
            "assignment_date": current_assignment.get("assignment_date") or "",
            "status": current_assignment.get("status") or asset.get("current_status") or "",
            "notes": current_assignment.get("notes") or "",
            "handover_condition": current_assignment.get("handover_condition") or "",
        },
    }


def describe_assignment_update_error(error: Exception) -> str:
    if not isinstance(error, APIError):
        return "Assignment could not be saved due to an unexpected database error."

    message = error.message or "Database error"
    details = error.details or ""
    combined = " ".join(part for part in [message, details] if part).lower()

    if "handover_condition" in combined and "column" in combined:
        return "The database schema is missing the handover_condition column in asset_assignments."

    return f"Assignment could not be saved: {message}"


def get_asset_serial_value(asset: dict) -> str:
    return asset.get("serial_chassis_number") or asset.get("serial_number") or ""


def get_asset_form_values(asset: Optional[dict] = None) -> dict:
    asset = asset or {}
    current_status = asset.get("current_status") or ""
    standard_status_values = {value for value, _ in ASSET_STATUS_SELECT_OPTIONS}
    return {
        "asset_tag_number": asset.get("asset_tag_number") or suggest_next_asset_tag(),
        "usage_type": normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number")),
        "item_description": asset.get("item_description") or "",
        "brand_make": asset.get("brand_make") or "",
        "model": asset.get("model") or "",
        "asset_classification": asset.get("asset_classification") or "",
        "asset_sub_classification": asset.get("asset_sub_classification") or "",
        "quantity": asset.get("quantity") or "",
        "purchase_price": asset.get("purchase_price") or "",
        "currency": asset.get("currency") or "",
        "serial_number": get_asset_serial_value(asset),
        "current_status": current_status,
        "current_status_select": current_status if current_status in standard_status_values else ("__custom__" if current_status else ""),
        "current_status_custom": current_status if current_status and current_status not in standard_status_values else "",
        "remarks": asset.get("remarks") or "",
        "payment_date": "",
        "payment_amount": "",
        "payment_currency": asset.get("currency") or "",
        "payment_status": "paid",
        "payment_notes": "",
    }


def parse_payment_date_field(value: str) -> Optional[str]:
    normalized = (value or "").strip()
    if not normalized:
        return None
    if re.fullmatch(r"\d{4}-\d{2}", normalized):
        normalized = f"{normalized}-01"
    for date_format in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(normalized, date_format).strftime("%Y-%m-%d")
        except ValueError:
            pass
    raise ValueError("Payment date must be a valid date.")


def get_next_payment_number(asset_id: int) -> int:
    response = (
        supabase.table("asset_payments")
        .select("payment_number")
        .eq("asset_id", asset_id)
        .order("payment_number", desc=True)
        .limit(1)
        .execute()
    )
    if not response.data:
        return 1
    return int(response.data[0].get("payment_number") or 0) + 1


def build_asset_payment_payload(
    asset_id: int,
    payment_number: Optional[int],
    payment_date: str,
    payment_amount: str,
    currency: str,
    payment_status: str,
    notes: str,
    fallback_amount: Optional[float] = None,
) -> dict:
    parsed_date = parse_payment_date_field(payment_date)
    parsed_amount = parse_float_field(payment_amount)
    if parsed_amount is None:
        parsed_amount = fallback_amount
    if not parsed_date:
        raise ValueError("Payment date is required.")
    if parsed_amount is None:
        raise ValueError("Payment amount is required.")

    return {
        "asset_id": asset_id,
        "payment_number": payment_number if payment_number is not None else get_next_payment_number(asset_id),
        "payment_date": parsed_date,
        "payment_amount": parsed_amount,
        "currency": currency.strip() or "EUR",
        "payment_status": payment_status.strip() or "paid",
        "notes": notes.strip() or None,
    }


def describe_asset_payment_error(error: Exception) -> str:
    if isinstance(error, ValueError):
        return str(error)
    if not isinstance(error, APIError):
        return "Payment could not be saved due to an unexpected database error."
    return f"Payment could not be saved: {error.message or 'Database error'}"


def list_lookup_values(table_name: str, column_name: str, fallback: Optional[list[str]] = None) -> list[str]:
    fallback = fallback or []
    try:
        response = (
            supabase.table(table_name)
            .select(column_name)
            .order(column_name)
            .execute()
        )
    except Exception:
        return fallback

    values = []
    for row in response.data or []:
        value = (row.get(column_name) or "").strip()
        if value and value not in values:
            values.append(value)
    return values or fallback


def list_distinct_asset_field_values(field_name: str) -> list[str]:
    try:
        response = supabase.table("assets").select(field_name).execute()
    except Exception:
        return []

    values = []
    for row in response.data or []:
        value = (row.get(field_name) or "").strip()
        if value and value not in values:
            values.append(value)
    return sorted(values)


def canonical_option_key(value: str) -> str:
    translation = str.maketrans({
        "0": "o",
        "1": "i",
        "3": "e",
        "4": "i",
        "5": "s",
        "7": "t",
    })
    return "".join(character for character in value.lower().translate(translation) if character.isalnum())


def option_quality_score(value: str) -> tuple[int, int, int]:
    digit_count = sum(character.isdigit() for character in value)
    alpha_count = sum(character.isalpha() for character in value)
    return (-digit_count, alpha_count, -len(value))


def merge_preferred_options(primary: list[str], secondary: list[str]) -> list[str]:
    merged: dict[str, str] = {}
    for value in [*primary, *secondary]:
        normalized = (value or "").strip()
        if not normalized:
            continue
        key = canonical_option_key(normalized) or normalized.casefold()
        existing = merged.get(key)
        if existing is None or option_quality_score(normalized) > option_quality_score(existing):
            merged[key] = normalized
    return sorted(merged.values())


def get_asset_create_options() -> dict:
    asset_classifications = list_distinct_asset_field_values("asset_classification")
    asset_sub_classifications = list_distinct_asset_field_values("asset_sub_classification")
    classifications = merge_preferred_options(
        list_lookup_values("asset_classifications", "classification_name"),
        asset_classifications,
    )
    sub_classifications = merge_preferred_options(
        list_lookup_values("asset_sub_classifications", "sub_classification_name"),
        asset_sub_classifications,
    )

    currencies = []
    for table_name, column_name in [
        ("currencies", "currency_code"),
        ("currencies", "code"),
        ("asset_currencies", "currency_code"),
        ("asset_currencies", "code"),
        ("currency", "currency_code"),
        ("currency", "code"),
    ]:
        currencies = list_lookup_values(table_name, column_name)
        if currencies:
            break

    if not currencies:
        currencies = list_distinct_asset_field_values("currency")

    return {
        "status_options": ASSET_STATUS_SELECT_OPTIONS,
        "usage_type_options": ASSET_USAGE_TYPE_OPTIONS,
        "classification_options": classifications,
        "sub_classification_options": sub_classifications,
        "currency_options": currencies,
    }


def asset_tag_exists(asset_tag_number: str) -> bool:
    try:
        response = (
            supabase.table("assets")
            .select("asset_id")
            .eq("asset_tag_number", asset_tag_number)
            .limit(1)
            .execute()
        )
    except Exception:
        return False

    return bool(response.data)


def describe_asset_create_error(error: Exception) -> str:
    if not isinstance(error, APIError):
        return "Asset could not be created due to an unexpected database error."

    message = error.message or "Database error"
    details = error.details or ""
    combined = " ".join(part for part in [message, details] if part).lower()

    if "asset_tag_number" in combined and ("duplicate" in combined or "unique" in combined):
        return "Asset tag/Inventory No. already exists."

    if "inventory_code" in combined and "null value" in combined:
        return "Asset could not be created because Inventory Code was not mapped before saving."

    if "inventory_code" in combined and ("duplicate" in combined or "unique" in combined):
        return "Inventory Code already exists."

    if "serial_chassis_number" in combined and ("duplicate" in combined or "unique" in combined):
        return "Serial number already exists."

    if "usage_type" in combined and "column" in combined:
        return "The database schema is missing usage_type for asset type."

    return f"Asset could not be created: {message}"


def get_asset_tag_standard() -> dict:
    try:
        response = supabase.table("assets").select("asset_tag_number").execute()
    except Exception:
        return {"prefix": "", "width": 0, "example": "", "suggested_next": ""}

    sequence_map: dict[str, dict[str, int]] = {}
    for row in response.data or []:
        asset_tag = normalize_asset_tag(row.get("asset_tag_number") or "")
        match = re.match(r"^(.*?)(\d+)$", asset_tag)
        if not match:
            continue

        prefix = match.group(1)
        number = int(match.group(2))
        width = len(match.group(2))
        sequence = sequence_map.setdefault(prefix, {"count": 0, "max_number": 0, "width": width})
        sequence["count"] += 1
        if number > sequence["max_number"]:
            sequence["max_number"] = number
            sequence["width"] = width

    if not sequence_map:
        return {"prefix": "", "width": 0, "example": "", "suggested_next": ""}

    best_prefix, best_sequence = max(
        sequence_map.items(),
        key=lambda item: (item[1]["count"], item[1]["max_number"], item[0]),
    )
    next_number = best_sequence["max_number"] + 1
    width = best_sequence["width"]
    return {
        "prefix": best_prefix,
        "width": width,
        "example": f"{best_prefix}{'0' * max(width - 1, 0)}1" if width else "",
        "suggested_next": f"{best_prefix}{str(next_number).zfill(width)}" if width else "",
    }


def suggest_next_asset_tag() -> str:
    return get_asset_tag_standard().get("suggested_next") or ""


def get_asset_tag_warning(asset_tag_number: str, standard: Optional[dict] = None) -> str:
    asset_tag_number = normalize_asset_tag(asset_tag_number)
    standard = standard or get_asset_tag_standard()
    prefix = standard.get("prefix") or ""
    width = standard.get("width") or 0

    if not asset_tag_number or not prefix or not width:
        return ""

    match = re.match(r"^(.*?)(\d+)$", asset_tag_number)
    if not match:
        return "This inventory number does not follow the numbering pattern currently used in the database."

    current_prefix = match.group(1)
    current_width = len(match.group(2))
    if current_prefix != prefix or current_width != width:
        return "This inventory number differs from the numbering pattern currently used in the database."

    return ""


def get_effective_status(asset: dict) -> str:
    assignment = asset.get("current_assignment") or {}
    return assignment.get("status") or asset.get("current_status") or "-"


def list_assets(limit: Optional[int] = None, batch_size: int = 500) -> list[dict]:
    assets: list[dict] = []
    start = 0

    while True:
        query = (
            supabase.table("assets")
            .select("*")
            .order("asset_tag_number")
            .range(start, start + batch_size - 1)
        )
        if limit is not None:
            remaining = limit - len(assets)
            if remaining <= 0:
                break
            query = query.limit(min(batch_size, remaining))

        response = execute_supabase_query(query, "list_assets")
        batch = response.data or []
        if not batch:
            break

        assets.extend(batch)

        if limit is not None and len(assets) >= limit:
            assets = assets[:limit]
            break

        if len(batch) < batch_size:
            break

        start += len(batch)

    for asset in assets:
        asset["usage_type"] = normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number"))
        asset["usage_type_label"] = get_asset_usage_type_label(asset.get("usage_type"))
        asset["current_assignment"] = get_current_assignment(asset["asset_id"])
        asset["effective_status"] = get_effective_status(asset)
    return assets


def search_people_with_assets(query: str = "", show_all: bool = False, status: str = "active") -> list[dict]:
    assets = list_assets()
    people = list_people()
    assignments_by_person: dict[int, list[dict]] = {}

    for asset in assets:
        assignment = asset.get("current_assignment") or {}
        person_id = assignment.get("person_id")
        if person_id:
            assignments_by_person.setdefault(person_id, []).append(asset)

    rows: list[dict] = []
    normalized_query = query.strip().lower()
    normalized_status = (status or "active").strip().lower()

    for person in people:
        person_active = is_person_active(person)
        if normalized_status == "active" and not person_active:
            continue
        if normalized_status == "inactive" and person_active:
            continue

        person_id = person.get("person_id")
        assigned_assets = assignments_by_person.get(person_id, [])
        assigned_standard_assets = [
            asset
            for asset in assigned_assets
            if normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number")) == "standard"
        ]
        assigned_low_cost_assets = [
            asset
            for asset in assigned_assets
            if normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number")) == "low_cost"
        ]
        row = {
            "person_id": person_id,
            "display_name": get_person_display_name(person),
            "department": person.get("department") or "-",
            "assigned_count": len(assigned_assets),
            "standard_count": len(assigned_standard_assets),
            "low_cost_count": len(assigned_low_cost_assets),
            "is_active": person_active,
            "status_label": get_person_status_label(person),
            "assets": assigned_assets,
        }

        if not show_all and row["assigned_count"] == 0:
            continue

        if normalized_query:
            haystack = " ".join(
                [
                    str(row["display_name"]),
                    str(person.get("department") or ""),
                    str(person.get("name") or ""),
                    str(person.get("name_eng") or ""),
                    str(person.get("full_name") or ""),
                ]
            ).lower()
            if normalized_query not in haystack:
                continue

        rows.append(row)

    rows.sort(key=lambda item: (-item["assigned_count"], item["display_name"]))
    return rows


def get_assets_for_person(person_id: int) -> list[dict]:
    assignments_response = (
        supabase.table("asset_assignments")
        .select("*")
        .eq("person_id", person_id)
        .is_("return_date", "null")
        .execute()
    )
    assignments = assignments_response.data or []
    if not assignments:
        return []

    assignments_by_asset_id = {
        assignment.get("asset_id"): assignment
        for assignment in assignments
        if assignment.get("asset_id")
    }
    asset_ids = list(assignments_by_asset_id.keys())
    assets_response = (
        supabase.table("assets")
        .select("*")
        .in_("asset_id", asset_ids)
        .execute()
    )

    assets = assets_response.data or []
    for asset in assets:
        asset["usage_type"] = normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number"))
        asset["usage_type_label"] = get_asset_usage_type_label(asset.get("usage_type"))
        assignment = assignments_by_asset_id.get(asset.get("asset_id")) or {}
        asset["current_assignment"] = enrich_assignment(assignment)
        asset["effective_status"] = get_effective_status(asset)

    assets.sort(key=lambda item: item.get("asset_tag_number") or "")
    return assets


def asset_matches_query(asset: dict, query: str) -> bool:
    if not query:
        return True

    assignment = asset.get("current_assignment") or {}
    haystack = [
        asset.get("asset_tag_number"),
        asset.get("item_description"),
        asset.get("brand_make"),
        asset.get("model"),
        asset.get("asset_classification"),
        asset.get("asset_sub_classification"),
        asset.get("usage_type_label"),
        asset.get("effective_status"),
        assignment.get("responsible_person"),
        assignment.get("department"),
        assignment.get("city"),
    ]

    normalized_query = query.lower()
    return any(
        normalized_query in str(value).lower()
        for value in haystack
        if value
    )


def field_contains(value: Optional[str], query: str) -> bool:
    normalized_query = (query or "").strip().lower()
    if not normalized_query:
        return True
    return normalized_query in str(value or "").lower()


def asset_matches_filters(asset: dict, filters: dict[str, str]) -> bool:
    assignment = asset.get("current_assignment") or {}
    return (
        field_contains(asset.get("asset_tag_number"), filters.get("asset_tag", ""))
        and field_contains(asset.get("usage_type_label"), filters.get("usage_type", ""))
        and field_contains(asset.get("item_description"), filters.get("description", ""))
        and field_contains(asset.get("effective_status"), filters.get("status", ""))
        and field_contains(assignment.get("responsible_person"), filters.get("person", ""))
        and field_contains(assignment.get("department"), filters.get("department", ""))
        and field_contains(assignment.get("city"), filters.get("city", ""))
    )


def build_asset_summary(assets: list[dict]) -> dict:
    assigned_count = 0
    unassigned_count = 0
    lost_count = 0
    city_map: dict[str, int] = {}
    department_map: dict[str, int] = {}
    person_map: dict[str, int] = {}

    for asset in assets:
        assignment = asset.get("current_assignment") or {}
        status = (asset.get("effective_status") or "").upper()
        city = assignment.get("city") or "Unknown"
        department = assignment.get("department") or "Unknown"
        person = assignment.get("responsible_person") or "Unassigned"

        if assignment:
            assigned_count += 1
        else:
            unassigned_count += 1

        if status == "LOST":
            lost_count += 1

        city_map[city] = city_map.get(city, 0) + 1
        department_map[department] = department_map.get(department, 0) + 1
        person_map[person] = person_map.get(person, 0) + 1

    def top_items(data: dict[str, int], limit: int = 8) -> list[dict]:
        return [
            {"label": key, "count": value}
            for key, value in sorted(data.items(), key=lambda item: (-item[1], item[0]))[:limit]
        ]

    return {
        "total_assets": len(assets),
        "assigned_assets": assigned_count,
        "unassigned_assets": unassigned_count,
        "lost_assets": lost_count,
        "top_cities": top_items(city_map),
        "top_departments": top_items(department_map),
        "top_people": top_items(person_map),
    }


def summarize_assets_by_field(assets: list[dict], field_name: str, empty_label: str) -> list[dict]:
    counts: dict[str, int] = {}

    for asset in assets:
        assignment = asset.get("current_assignment") or {}
        label = assignment.get(field_name) or empty_label
        counts[label] = counts.get(label, 0) + 1

    return [
        {"label": key, "count": value}
        for key, value in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    ]


def build_assignment_audit_rows(assets: list[dict]) -> list[dict]:
    rows: list[dict] = []

    for asset in assets:
        assignment = asset.get("current_assignment") or {}
        rows.append(
            {
                "asset_tag_number": asset.get("asset_tag_number") or "",
                "item_description": asset.get("item_description") or "",
                "brand_make": asset.get("brand_make") or "",
                "model": asset.get("model") or "",
                "effective_status": asset.get("effective_status") or "",
                "current_status": asset.get("current_status") or "",
                "responsible_person": assignment.get("responsible_person") or "",
                "department": assignment.get("department") or "",
                "city": assignment.get("city") or "",
                "location_name": assignment.get("location_name") or "",
                "assignment_date": assignment.get("assignment_date") or "",
                "assignment_status": assignment.get("status") or "",
                "assignment_notes": assignment.get("notes") or "",
            }
        )

    return rows


def csv_response(filename: str, fieldnames: list[str], rows: list[dict]) -> StreamingResponse:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    buffer.seek(0)

    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(iter([buffer.getvalue()]), media_type="text/csv; charset=utf-8", headers=headers)


def format_asset_message(asset: dict) -> str:
    assignment = asset.get("current_assignment") or {}
    effective_status = assignment.get("status") or asset.get("current_status") or "-"
    assignment_date = assignment.get("assignment_date") or "-"

    return (
        f"📦 Asset: {asset.get('asset_tag_number', '-')}\n"
        f"📝 Description: {asset.get('item_description', '-')}\n"
        f"🏷️ Brand: {asset.get('brand_make', '-')}\n"
        f"📐 Model: {asset.get('model', '-')}\n"
        f"📂 Classification: {asset.get('asset_classification', '-')} / "
        f"{asset.get('asset_sub_classification', '-')}\n"
        f"📊 Status: {effective_status}\n"
        f"💰 Price: {asset.get('purchase_price', '-')} {asset.get('currency', '')}\n"
        f"🔢 Qty: {asset.get('quantity', '-')}\n"
        f"👤 Responsible person: {assignment.get('responsible_person') or '-'}\n"
        f"🏢 Department: {assignment.get('department') or '-'}\n"
        f"📍 City: {assignment.get('city') or '-'}\n"
        f"📅 Assignment date: {assignment_date}"
    )


def normalize_phone_number(value: Optional[str]) -> str:
    digits = re.sub(r"\D+", "", value or "")
    if not digits:
        return ""
    if digits.startswith("00"):
        digits = digits[2:]
    if digits.startswith("0") and len(digits) == 10:
        digits = "38" + digits
    return digits


def get_person_phone_values(person: dict) -> list[str]:
    return [
        person.get("mobile_phone"),
        person.get("phone"),
        person.get("telegram"),
    ]


def find_person_by_phone(phone_number: str) -> Optional[dict]:
    normalized_phone = normalize_phone_number(phone_number)
    if not normalized_phone:
        return None

    for person in list_people():
        for value in get_person_phone_values(person):
            candidate = normalize_phone_number(value)
            if candidate and candidate == normalized_phone:
                return person
    return None


def find_person_by_telegram_user_id(telegram_user_id) -> Optional[dict]:
    if telegram_user_id is None:
        return None
    telegram_user_id = str(telegram_user_id)
    try:
        response = (
            supabase.table("persons")
            .select("*")
            .eq("messenger_type", "telegram")
            .eq("messenger_id", telegram_user_id)
            .limit(1)
            .execute()
        )
    except Exception:
        return None
    return (response.data or [None])[0]


def save_person_telegram_identity(person_id: int, telegram_user: dict, phone_number: Optional[str] = None) -> None:
    payload = {
        "messenger_type": "telegram",
        "messenger_id": str(telegram_user.get("id")),
    }
    username = telegram_user.get("username")
    if username:
        payload["username"] = username
    if phone_number:
        payload["mobile_phone"] = phone_number
    try:
        supabase.table("persons").update(payload).eq("person_id", person_id).execute()
    except Exception:
        return


def get_authorized_telegram_person(message: dict) -> Optional[dict]:
    telegram_user = message.get("from") or {}
    person = find_person_by_telegram_user_id(telegram_user.get("id"))
    if person and is_person_active(person):
        return person
    return None


def format_person_assets_message(person: dict) -> str:
    assets = get_assets_for_person(person["person_id"])
    display_name = get_person_display_name(person)
    if not assets:
        return f"{display_name}, no active assets are currently assigned to you."

    lines = [
        f"{display_name}, assets assigned to you: {len(assets)}",
        "",
    ]
    for index, asset in enumerate(assets, start=1):
        if index > 1:
            lines.append("-----")
        brand_model = " / ".join(
            value
            for value in [asset.get("brand_make"), asset.get("model")]
            if value
        )
        lines.append(
            f"{index}. {asset.get('asset_tag_number') or '-'} | "
            f"{asset.get('item_description') or '-'}"
            f"{f' | {brand_model}' if brand_model else ''}"
        )
    return "\n".join(lines)


def get_telegram_asset_list_serializer() -> URLSafeSerializer:
    return URLSafeSerializer(ADMIN_SESSION_SECRET, salt="telegram-asset-list")


def create_telegram_asset_list_token(person: dict) -> str:
    return get_telegram_asset_list_serializer().dumps(
        {
            "person_id": person.get("person_id"),
            "messenger_id": person.get("messenger_id"),
        }
    )


def load_telegram_asset_list_person(token: str) -> Optional[dict]:
    try:
        data = get_telegram_asset_list_serializer().loads(token)
    except BadSignature:
        return None
    person_id = data.get("person_id")
    if not person_id:
        return None
    person = get_person_by_id(int(person_id))
    if not person or not is_person_active(person):
        return None
    expected_messenger_id = data.get("messenger_id")
    if expected_messenger_id and str(person.get("messenger_id") or "") != str(expected_messenger_id):
        return None
    return person


def split_person_assets_by_usage(person: dict) -> tuple[list[dict], list[dict], list[dict]]:
    assigned_assets = get_assets_for_person(person["person_id"])
    assigned_standard_assets = [
        asset
        for asset in assigned_assets
        if normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number")) == "standard"
    ]
    assigned_low_cost_assets = [
        asset
        for asset in assigned_assets
        if normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number")) == "low_cost"
    ]
    return assigned_assets, assigned_standard_assets, assigned_low_cost_assets


def build_telegram_asset_report_pdf(
    person: dict,
    assigned_assets: list[dict],
    assigned_standard_assets: list[dict],
    assigned_low_cost_assets: list[dict],
    branding: dict,
) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise HTTPException(status_code=503, detail="PDF generation is not available yet") from exc

    def plain_text(value) -> str:
        return str(value if value not in (None, "") else "-")

    def pdf_text(value) -> str:
        return html.escape(plain_text(value), quote=False)

    def paragraph(value) -> Paragraph:
        return Paragraph(pdf_text(value), body_style)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Asset List - {get_person_display_name(person)}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "AssetReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#102033"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "AssetReportSubtitle",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#657186"),
    )
    body_style = ParagraphStyle(
        "AssetReportBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        textColor=colors.HexColor("#102033"),
    )
    small_style = ParagraphStyle(
        "AssetReportSmall",
        parent=body_style,
        fontSize=6.5,
        leading=8,
        textColor=colors.HexColor("#657186"),
    )
    right_style = ParagraphStyle("AssetReportRight", parent=small_style, alignment=TA_RIGHT)

    printed_at = datetime.now(ZoneInfo("Europe/Kyiv")).strftime("%d.%m.%Y")
    display_name = get_person_display_name(person)
    story = [
        Table(
            [
                [
                    Paragraph(pdf_text(branding.get("company_name") or "Your Company"), small_style),
                    Paragraph(f"Printed on<br/><b>{printed_at}</b>", right_style),
                ],
                [
                    Paragraph(pdf_text(branding.get("report_title") or "Asset Assignment Report"), title_style),
                    "",
                ],
                [
                    Paragraph(pdf_text(branding.get("report_subtitle") or "Current asset assignment register."), subtitle_style),
                    "",
                ],
            ],
            colWidths=[230 * mm, 40 * mm],
        ),
        Spacer(1, 5 * mm),
        Table(
            [
                ["Employee name", plain_text(display_name), "Department", plain_text(person.get("department"))],
                ["Person ID", plain_text(person.get("person_id")), "Report type", "Current assignment register"],
                ["Total assigned assets", plain_text(len(assigned_assets)), "Standard / Low-cost", f"{len(assigned_standard_assets)} / {len(assigned_low_cost_assets)}"],
            ],
            colWidths=[38 * mm, 96 * mm, 38 * mm, 98 * mm],
            style=[
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#657186")),
                ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#657186")),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
                ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#dfe7ef")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dfe7ef")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ],
        ),
        Spacer(1, 6 * mm),
    ]

    def add_asset_table(title: str, assets: list[dict], is_low_cost: bool = False):
        story.append(Paragraph(title, ParagraphStyle("SectionTitle" + title, parent=styles["Heading2"], fontSize=12, leading=14)))
        headers = ["Inventory code" if is_low_cost else "Asset tag", "Description", "Status", "City", "Department" if is_low_cost else "Location", "Assigned"]
        rows = [[Paragraph(header, small_style) for header in headers]]
        if not assets:
            rows.append([Paragraph("No active items are currently assigned.", body_style), "", "", "", "", ""])
        for asset in assets:
            assignment = asset.get("current_assignment") or {}
            brand_model = " / ".join(value for value in [asset.get("brand_make"), asset.get("model")] if value)
            description = pdf_text(asset.get("item_description"))
            if brand_model:
                description = f"{description}<br/><font color='#657186'>{brand_model}</font>"
            rows.append(
                [
                    paragraph(asset.get("asset_tag_number")),
                    Paragraph(description, body_style),
                    paragraph(asset.get("effective_status")),
                    paragraph(assignment.get("city")),
                    paragraph(assignment.get("department") if is_low_cost else assignment.get("location_name") or assignment.get("department")),
                    paragraph(assignment.get("assignment_date")),
                ]
            )

        table = Table(
            rows,
            colWidths=[29 * mm, 139 * mm, 24 * mm, 25 * mm, 35 * mm, 18 * mm],
            repeatRows=1,
            style=TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#edf4f1")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#657186")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dfe7ef")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            ),
        )
        story.extend([table, Spacer(1, 5 * mm)])

    add_asset_table("Assigned Standard Assets", assigned_standard_assets)
    add_asset_table("Assigned Low-cost Items", assigned_low_cost_assets, is_low_cost=True)

    story.append(
        Table(
            [
                [
                    Paragraph(f"{pdf_text(branding.get('issuer_label') or 'Issued by')}<br/><br/><br/>____________________________<br/>{pdf_text(branding.get('issuer_signature_label') or 'Signature')}", body_style),
                    Paragraph(f"{pdf_text(branding.get('receiver_label') or 'Received by')}<br/><br/><br/>____________________________<br/>{pdf_text(branding.get('receiver_signature_label') or display_name + ' signature')}", body_style),
                ]
            ],
            colWidths=[132 * mm, 132 * mm],
            style=[
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#dfe7ef")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dfe7ef")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ],
        )
    )
    if branding.get("footer_note"):
        story.extend([Spacer(1, 4 * mm), Paragraph(pdf_text(branding.get("footer_note")), small_style)])

    doc.build(story)
    return buffer.getvalue()


def get_telegram_asset_list_url(person: dict) -> str:
    token = create_telegram_asset_list_token(person)
    return f"{PUBLIC_BASE_URL}/telegram/assets/{token}"


def send_telegram_auth_prompt(chat_id: int) -> None:
    send_telegram_message(
        chat_id,
        "Please authorize first: tap 'Share phone number'. The number must match your employee profile.",
        reply_markup=AUTH_KEYBOARD,
    )


def send_telegram_message(chat_id: int, text: str, reply_markup: Optional[dict] = None) -> None:
    payload = {"chat_id": chat_id, "text": text}
    if reply_markup is not None:
        payload["reply_markup"] = reply_markup

    requests.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
        json=payload,
        timeout=15,
    )


def send_telegram_long_message(chat_id: int, text: str, reply_markup: Optional[dict] = None, max_length: int = 3500) -> None:
    lines = text.splitlines()
    chunks = []
    current = ""
    for line in lines:
        addition = line if not current else f"\n{line}"
        if len(current) + len(addition) > max_length:
            if current:
                chunks.append(current)
            current = line
        else:
            current += addition
    if current:
        chunks.append(current)
    if not chunks:
        chunks = [text]

    for index, chunk in enumerate(chunks):
        send_telegram_message(chat_id, chunk, reply_markup=reply_markup if index == len(chunks) - 1 else None)


MAIN_KEYBOARD = {
    "keyboard": [
        [
            {"text": "My assets"},
        ],
        [
            {"text": "Scan QR", "web_app": {"url": f"{PUBLIC_BASE_URL}/miniapp"}},
            {"text": "Enter code"},
        ],
        [{"text": "Help"}],
    ],
    "resize_keyboard": True,
}

AUTH_KEYBOARD = {
    "keyboard": [
        [{"text": "Share phone number", "request_contact": True}],
        [{"text": "Help"}],
    ],
    "resize_keyboard": True,
    "one_time_keyboard": True,
}


@app.get("/")
def root():
    return {"status": "ok", "message": "Inventory system is running"}


@app.get("/asset/{asset_tag}")
def read_asset(asset_tag: str):
    asset = get_asset_by_tag(asset_tag.strip())
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return asset


@app.get("/view/{asset_tag}", response_class=HTMLResponse)
def view_asset(request: Request, asset_tag: str):
    asset = get_asset_by_tag(asset_tag.strip())
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    return templates.TemplateResponse(
        request=request,
        name="asset.html",
        context={"asset": asset},
    )


@app.get("/miniapp", response_class=HTMLResponse)
def miniapp(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="miniapp.html",
        context={},
    )


@app.get("/telegram/assets/{token}", response_class=HTMLResponse)
def telegram_person_asset_list(request: Request, token: str):
    person = load_telegram_asset_list_person(token)
    if not person:
        raise HTTPException(status_code=404, detail="Asset list not found")

    assigned_assets, assigned_standard_assets, assigned_low_cost_assets = split_person_assets_by_usage(person)
    _, branding, branding_storage = resolve_branding_for_request(request)

    return templates.TemplateResponse(
        request=request,
        name="telegram_asset_list.html",
        context={
            "person": person,
            "display_name": get_person_display_name(person),
            "report_display_name": get_person_report_name(person),
            "assigned_assets": assigned_assets,
            "assigned_standard_assets": assigned_standard_assets,
            "assigned_low_cost_assets": assigned_low_cost_assets,
            "printed_at": datetime.now(ZoneInfo("Europe/Kyiv")).strftime("%d.%m.%Y"),
            "branding": branding,
            "branding_storage": branding_storage,
            "branding_logo_url": get_branding_logo_url(branding),
            "pdf_url": f"/telegram/assets/{token}/pdf",
            "page_title": f"Asset List - {get_person_display_name(person)}",
        },
    )


@app.get("/telegram/assets/{token}/pdf")
def telegram_person_asset_list_pdf(request: Request, token: str):
    person = load_telegram_asset_list_person(token)
    if not person:
        raise HTTPException(status_code=404, detail="Asset list not found")

    assigned_assets, assigned_standard_assets, assigned_low_cost_assets = split_person_assets_by_usage(person)
    _, branding, _ = resolve_branding_for_request(request)
    pdf_bytes = build_telegram_asset_report_pdf(
        person,
        assigned_assets,
        assigned_standard_assets,
        assigned_low_cost_assets,
        branding,
    )
    filename_name = re.sub(r"[^A-Za-z0-9_-]+", "_", get_person_display_name(person)).strip("_") or "employee"
    headers = {
        "Content-Disposition": f'attachment; filename="asset_list_{filename_name}.pdf"',
        "Cache-Control": "no-store",
    }
    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)


@app.get("/admin/login", response_class=HTMLResponse, name="admin_login")
def admin_login(request: Request, next: str = "/admin"):
    if is_admin_authenticated(request):
        return RedirectResponse(url=next or "/admin", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="admin_login.html",
        context={
            "error": None,
            "next_url": next or "/admin",
        },
    )


@app.post("/admin/login", response_class=HTMLResponse)
def admin_login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    next: str = Form("/admin"),
):
    input_username = normalize_credential(username)
    input_password = normalize_credential(password)
    expected_username = normalize_credential(ADMIN_USERNAME)
    expected_password = normalize_credential(ADMIN_PASSWORD)

    valid_username = secrets.compare_digest(input_username, expected_username)
    valid_password = secrets.compare_digest(input_password, expected_password)

    if not (valid_username and valid_password):
        return templates.TemplateResponse(
            request=request,
            name="admin_login.html",
            context={
                "error": "Invalid username or password.",
                "next_url": next or "/admin",
            },
            status_code=401,
        )

    request.session["admin_authenticated"] = True
    request.session["admin_username"] = input_username
    request.session["admin_tenant_key"] = DEFAULT_BRANDING_TENANT_KEY
    return RedirectResponse(url=next or "/admin", status_code=303)


@app.get("/admin/logout")
def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/admin/login", status_code=303)


@app.get("/admin", response_class=HTMLResponse)
def admin_dashboard(request: Request, changes_limit: int = 5):
    redirect = require_admin(request)
    if redirect:
        return redirect

    database_error = ""
    changes_limit = normalize_audit_limit(changes_limit)
    try:
        assets = list_assets()
        summary = build_asset_summary(assets)
    except DatabaseConnectionError as exc:
        summary = build_asset_summary([])
        database_error = str(exc)

    return templates.TemplateResponse(
        request=request,
        name="admin_dashboard.html",
        context={
            "summary": summary,
            "recent_changes": list_recent_audit_events(changes_limit),
            "changes_limit": changes_limit,
            "changes_limit_options": [5, 10, 20, 50],
            "database_error": database_error,
            "flash": pop_flash(request),
            "active_page": "dashboard",
            "page_title": "Admin Dashboard",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.post("/admin/audit/backfill-transfer-log")
def admin_audit_backfill_transfer_log(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    try:
        result = backfill_audit_from_transfer_log()
    except Exception as error:
        set_flash(request, "error", f"Transfer log backfill could not be completed: {error}")
        return RedirectResponse(url="/admin", status_code=303)

    set_flash(
        request,
        "success",
        f"Transfer log backfill completed: {result['created']} audit records created, "
        f"{result.get('updated', 0)} refreshed from {result['available']} transfer records.",
    )
    return RedirectResponse(url="/admin", status_code=303)


@app.get("/admin/audit-log", response_class=HTMLResponse)
def admin_audit_log(
    request: Request,
    q: str = "",
    entity_type: str = "",
    source: str = "",
    page: int = 1,
    page_size: int = 100,
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    page = max(page, 1)
    page_size = normalize_audit_page_size(page_size)
    result = list_audit_log_events(
        q=q.strip(),
        entity_type=entity_type.strip(),
        source=source.strip(),
        page=page,
        page_size=page_size,
    )

    query_params = {
        "q": q.strip(),
        "entity_type": entity_type.strip(),
        "source": source.strip(),
        "page_size": page_size,
    }
    base_query_string = urlencode({key: value for key, value in query_params.items() if value not in ("", None)})
    previous_page_url = f"/admin/audit-log?{base_query_string}&page={page - 1}" if page > 1 else ""
    next_page_url = f"/admin/audit-log?{base_query_string}&page={page + 1}" if result["has_next"] else ""

    return templates.TemplateResponse(
        request=request,
        name="admin_audit_log.html",
        context={
            "audit_rows": result["rows"],
            "query": q.strip(),
            "entity_type": entity_type.strip(),
            "source": source.strip(),
            "page": result["page"],
            "page_size": result["page_size"],
            "has_next": result["has_next"],
            "total_loaded": result["total_loaded"],
            "page_size_options": [50, 100, 200],
            "entity_type_options": list_audit_filter_values("entity_type"),
            "source_options": list_audit_filter_values("source"),
            "query_params": query_params,
            "previous_page_url": previous_page_url,
            "next_page_url": next_page_url,
            "flash": pop_flash(request),
            "active_page": "dashboard",
            "page_title": "Audit Log",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.get("/admin/assets", response_class=HTMLResponse)
def admin_assets(
    request: Request,
    q: str = "",
    usage_type: str = "",
    asset_tag: str = "",
    description: str = "",
    status: str = "",
    person: str = "",
    department: str = "",
    city: str = "",
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    assets = list_assets()
    filters = {
        "usage_type": usage_type.strip(),
        "asset_tag": asset_tag.strip(),
        "description": description.strip(),
        "status": status.strip(),
        "person": person.strip(),
        "department": department.strip(),
        "city": city.strip(),
    }
    filtered_assets = [
        asset
        for asset in assets
        if asset_matches_query(asset, q.strip()) and asset_matches_filters(asset, filters)
    ]

    return templates.TemplateResponse(
        request=request,
        name="admin_assets.html",
        context={
            "assets": filtered_assets,
            "query": q,
            "filters": filters,
            "total_found": len(filtered_assets),
            "flash": pop_flash(request),
            "active_page": "assets",
            "page_title": "Admin Assets",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.get("/admin/assets/new", response_class=HTMLResponse)
def admin_asset_new(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    asset_tag_standard = get_asset_tag_standard()
    return templates.TemplateResponse(
        request=request,
        name="admin_asset_create.html",
        context={
            "asset_form": get_asset_form_values(),
            **get_asset_create_options(),
            "asset_tag_standard": asset_tag_standard,
            "asset_tag_warning": "",
            "flash": pop_flash(request),
            "active_page": "assets",
            "page_title": "New Asset",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.post("/admin/assets/new")
def admin_asset_create(
    request: Request,
    asset_tag_number: str = Form(""),
    usage_type: str = Form("standard"),
    item_description: str = Form(""),
    brand_make: str = Form(""),
    model: str = Form(""),
    asset_classification: str = Form(""),
    asset_sub_classification: str = Form(""),
    quantity: str = Form(""),
    purchase_price: str = Form(""),
    currency: str = Form(""),
    serial_number: str = Form(""),
    current_status: str = Form(""),
    current_status_custom: str = Form(""),
    remarks: str = Form(""),
    payment_date: str = Form(""),
    payment_amount: str = Form(""),
    payment_currency: str = Form(""),
    payment_status: str = Form("paid"),
    payment_notes: str = Form(""),
    confirm_nonstandard_asset_tag: str = Form(""),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    resolved_status = current_status_custom.strip() if current_status == "__custom__" else current_status.strip()
    asset_tag_standard = get_asset_tag_standard()
    asset_form = {
        "asset_tag_number": normalize_asset_tag(asset_tag_number),
        "usage_type": normalize_asset_usage_type(usage_type, asset_tag_number),
        "item_description": item_description.strip(),
        "brand_make": brand_make.strip(),
        "model": model.strip(),
        "asset_classification": asset_classification.strip(),
        "asset_sub_classification": asset_sub_classification.strip(),
        "quantity": quantity.strip(),
        "purchase_price": purchase_price.strip(),
        "currency": currency.strip(),
        "serial_number": serial_number.strip(),
        "current_status": resolved_status,
        "current_status_select": current_status.strip(),
        "current_status_custom": current_status_custom.strip(),
        "remarks": remarks.strip(),
        "payment_date": payment_date.strip(),
        "payment_amount": payment_amount.strip(),
        "payment_currency": payment_currency.strip() or currency.strip(),
        "payment_status": payment_status.strip() or "paid",
        "payment_notes": payment_notes.strip(),
    }

    format_error = validate_asset_tag_format(asset_form["asset_tag_number"])
    if format_error:
        return templates.TemplateResponse(
            request=request,
            name="admin_asset_create.html",
            context={
                "asset_form": asset_form,
                **get_asset_create_options(),
                "asset_tag_standard": asset_tag_standard,
                "asset_tag_warning": "",
                "flash": {"level": "error", "message": format_error},
                "active_page": "assets",
                "page_title": "New Asset",
                "admin_username": request.session.get("admin_username"),
            },
            status_code=400,
        )

    if current_status == "__custom__" and not asset_form["current_status_custom"]:
        return templates.TemplateResponse(
            request=request,
            name="admin_asset_create.html",
            context={
                "asset_form": asset_form,
                **get_asset_create_options(),
                "asset_tag_standard": asset_tag_standard,
                "asset_tag_warning": "",
                "flash": {"level": "error", "message": "Enter a custom asset status or choose one from the list."},
                "active_page": "assets",
                "page_title": "New Asset",
                "admin_username": request.session.get("admin_username"),
            },
            status_code=400,
        )

    asset_tag_warning = get_asset_tag_warning(asset_form["asset_tag_number"], asset_tag_standard)
    if asset_tag_warning and confirm_nonstandard_asset_tag != "yes":
        return templates.TemplateResponse(
            request=request,
            name="admin_asset_create.html",
            context={
                "asset_form": asset_form,
                **get_asset_create_options(),
                "asset_tag_standard": asset_tag_standard,
                "asset_tag_warning": asset_tag_warning,
                "flash": {"level": "error", "message": "Please confirm saving this non-standard inventory number."},
                "active_page": "assets",
                "page_title": "New Asset",
                "admin_username": request.session.get("admin_username"),
            },
            status_code=400,
        )

    if asset_tag_exists(asset_form["asset_tag_number"]):
        return templates.TemplateResponse(
            request=request,
            name="admin_asset_create.html",
            context={
                "asset_form": asset_form,
                **get_asset_create_options(),
                "asset_tag_standard": asset_tag_standard,
                "asset_tag_warning": asset_tag_warning,
                "flash": {"level": "error", "message": f"Asset tag/Inventory No. '{asset_form['asset_tag_number']}' already exists."},
                "active_page": "assets",
                "page_title": "New Asset",
                "admin_username": request.session.get("admin_username"),
            },
            status_code=400,
        )

    try:
        insert_data = {
            "asset_tag_number": asset_form["asset_tag_number"],
            "inventory_code": asset_form["asset_tag_number"],
            "usage_type": asset_form["usage_type"],
            "item_description": asset_form["item_description"] or None,
            "brand_make": asset_form["brand_make"] or None,
            "model": asset_form["model"] or None,
            "asset_classification": asset_form["asset_classification"] or None,
            "asset_sub_classification": asset_form["asset_sub_classification"] or None,
            "quantity": parse_int_field(asset_form["quantity"]),
            "purchase_price": parse_float_field(asset_form["purchase_price"]),
            "currency": asset_form["currency"] or None,
            "serial_chassis_number": asset_form["serial_number"] or None,
            "current_status": asset_form["current_status"] or None,
            "remarks": asset_form["remarks"] or None,
        }
    except ValueError:
        return templates.TemplateResponse(
            request=request,
            name="admin_asset_create.html",
            context={
                "asset_form": asset_form,
                **get_asset_create_options(),
                "asset_tag_standard": asset_tag_standard,
                "asset_tag_warning": asset_tag_warning,
                "flash": {"level": "error", "message": "Quantity must be an integer and purchase price must be a number."},
                "active_page": "assets",
                "page_title": "New Asset",
                "admin_username": request.session.get("admin_username"),
            },
            status_code=400,
        )

    if asset_form["payment_date"] or asset_form["payment_amount"]:
        try:
            build_asset_payment_payload(
                asset_id=0,
                payment_number=1,
                payment_date=asset_form["payment_date"],
                payment_amount=asset_form["payment_amount"],
                currency=asset_form["payment_currency"],
                payment_status=asset_form["payment_status"],
                notes=asset_form["payment_notes"],
                fallback_amount=insert_data.get("purchase_price"),
            )
        except Exception as error:
            return templates.TemplateResponse(
                request=request,
                name="admin_asset_create.html",
                context={
                    "asset_form": asset_form,
                    **get_asset_create_options(),
                    "asset_tag_standard": asset_tag_standard,
                    "asset_tag_warning": asset_tag_warning,
                    "flash": {"level": "error", "message": describe_asset_payment_error(error)},
                    "active_page": "assets",
                    "page_title": "New Asset",
                    "admin_username": request.session.get("admin_username"),
                },
                status_code=400,
            )

    try:
        response = (
            supabase.table("assets")
            .insert(insert_data)
            .execute()
        )
    except Exception as exc:
        return templates.TemplateResponse(
            request=request,
            name="admin_asset_create.html",
            context={
                "asset_form": asset_form,
                **get_asset_create_options(),
                "asset_tag_standard": asset_tag_standard,
                "asset_tag_warning": asset_tag_warning,
                "flash": {"level": "error", "message": describe_asset_create_error(exc)},
                "active_page": "assets",
                "page_title": "New Asset",
                "admin_username": request.session.get("admin_username"),
            },
            status_code=400,
        )

    created_asset = (response.data or [{}])[0]
    created_asset_id = created_asset.get("asset_id")
    if not created_asset_id:
        created_asset = get_asset_by_tag(asset_form["asset_tag_number"]) or {}
        created_asset_id = created_asset.get("asset_id")

    if not created_asset_id:
        set_flash(request, "success", f"Asset {asset_form['asset_tag_number']} was created.")
        return RedirectResponse(url="/admin/assets", status_code=303)

    if asset_form["payment_date"] or asset_form["payment_amount"]:
        try:
            payment_payload = build_asset_payment_payload(
                asset_id=created_asset_id,
                payment_number=1,
                payment_date=asset_form["payment_date"],
                payment_amount=asset_form["payment_amount"],
                currency=asset_form["payment_currency"],
                payment_status=asset_form["payment_status"],
                notes=asset_form["payment_notes"],
                fallback_amount=parse_float_field(asset_form["purchase_price"]),
            )
            supabase.table("asset_payments").insert(payment_payload).execute()
            audit_log_event(
                entity_type="Payment",
                entity_id=created_asset_id,
                entity_label=asset_form["asset_tag_number"],
                action="created",
                summary=f"Added initial payment for {asset_form['asset_tag_number']}: {payment_payload.get('payment_amount') or '-'} {payment_payload.get('currency') or ''}",
                request=request,
            )
        except Exception as error:
            set_flash(request, "error", f"Asset was created, but payment was not saved: {describe_asset_payment_error(error)}")
            return RedirectResponse(url=f"/admin/assets/{created_asset_id}", status_code=303)

    audit_log_event(
        entity_type="Asset",
        entity_id=created_asset_id,
        entity_label=asset_form["asset_tag_number"],
        action="created",
        summary=f"Created asset: {asset_form['asset_tag_number']}",
        request=request,
    )
    set_flash(request, "success", f"Asset {asset_form['asset_tag_number']} was created.")
    return RedirectResponse(url=f"/admin/assets/{created_asset_id}", status_code=303)


@app.get("/admin/people", response_class=HTMLResponse)
def admin_people(request: Request, q: str = "", show_all: bool = False, status: str = "active"):
    redirect = require_admin(request)
    if redirect:
        return redirect

    status = status if status in {"active", "inactive", "all"} else "active"
    people = search_people_with_assets(q, show_all=show_all, status=status)

    return templates.TemplateResponse(
        request=request,
        name="admin_people.html",
        context={
            "people": people,
            "query": q,
            "show_all": show_all,
            "status": status,
            "flash": pop_flash(request),
            "active_page": "people",
            "page_title": "People",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.get("/admin/people/new", response_class=HTMLResponse)
def admin_person_new(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    return templates.TemplateResponse(
        request=request,
        name="admin_person_create.html",
        context={
            "person_form": get_person_form_values(),
            "flash": pop_flash(request),
            "active_page": "people",
            "page_title": "New Employee",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.post("/admin/people/new")
def admin_person_create(
    request: Request,
    name: str = Form(""),
    name_eng: str = Form(""),
    department: str = Form(""),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    person_form = get_person_form_values(
        {
            "name": name.strip(),
            "name_eng": name_eng.strip(),
            "department": department.strip(),
        }
    )

    if not person_form["name"] and not person_form["name_eng"]:
        return templates.TemplateResponse(
            request=request,
            name="admin_person_create.html",
            context={
                "person_form": person_form,
                "flash": {"level": "error", "message": "Enter at least one employee name: local or English."},
                "active_page": "people",
                "page_title": "New Employee",
                "admin_username": request.session.get("admin_username"),
            },
            status_code=400,
        )

    insert_data = {
        "name": person_form["name"] or None,
        "name_eng": person_form["name_eng"] or None,
        "department": person_form["department"] or None,
    }

    try:
        response = supabase.table("persons").insert(insert_data).execute()
    except Exception as exc:
        if is_person_id_sequence_conflict(exc):
            retry_data = dict(insert_data)
            try:
                retry_data["person_id"] = get_next_person_id()
                response = supabase.table("persons").insert(retry_data).execute()
            except Exception as retry_exc:
                exc = retry_exc
            else:
                exc = None

        if exc is None:
            created_person = (response.data or [{}])[0]
            created_person_id = created_person.get("person_id")
            created_name = person_form["name_eng"] or person_form["name"]

            set_flash(request, "success", f"Employee {created_name} was created.")
            audit_log_event(
                entity_type="Employee",
                entity_id=created_person_id,
                entity_label=created_name,
                action="created",
                summary=f"Created employee: {created_name}",
                request=request,
            )
            if created_person_id:
                return RedirectResponse(url=f"/admin/people/{created_person_id}", status_code=303)
            return RedirectResponse(url="/admin/people?show_all=true", status_code=303)

        if isinstance(exc, APIError):
            message = (exc.message or "").lower()
            details = (exc.details or "").lower()
            combined = f"{message} {details}"
            if "duplicate" in combined or "unique" in combined:
                existing_person = find_existing_person(person_form)
                if existing_person and existing_person.get("person_id"):
                    existing_name = get_person_display_name(existing_person)
                    set_flash(request, "success", f"Employee {existing_name} already exists. Opened the existing record.")
                    return RedirectResponse(url=f"/admin/people/{existing_person['person_id']}", status_code=303)

        return templates.TemplateResponse(
            request=request,
            name="admin_person_create.html",
            context={
                "person_form": person_form,
                "flash": {"level": "error", "message": describe_person_create_error(exc)},
                "active_page": "people",
                "page_title": "New Employee",
                "admin_username": request.session.get("admin_username"),
            },
            status_code=400,
        )

    created_person = (response.data or [{}])[0]
    created_person_id = created_person.get("person_id")
    created_name = person_form["name_eng"] or person_form["name"]

    set_flash(request, "success", f"Employee {created_name} was created.")
    audit_log_event(
        entity_type="Employee",
        entity_id=created_person_id,
        entity_label=created_name,
        action="created",
        summary=f"Created employee: {created_name}",
        request=request,
    )
    if created_person_id:
        return RedirectResponse(url=f"/admin/people/{created_person_id}", status_code=303)
    return RedirectResponse(url="/admin/people?show_all=true", status_code=303)


@app.get("/admin/people/{person_id}", response_class=HTMLResponse)
def admin_person_detail(request: Request, person_id: int):
    redirect = require_admin(request)
    if redirect:
        return redirect

    person = get_person_by_id(person_id)
    if not person:
        raise HTTPException(status_code=404, detail="Person not found")

    assigned_assets = get_assets_for_person(person_id)
    assigned_standard_assets = [
        asset
        for asset in assigned_assets
        if normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number")) == "standard"
    ]
    assigned_low_cost_assets = [
        asset
        for asset in assigned_assets
        if normalize_asset_usage_type(asset.get("usage_type"), asset.get("asset_tag_number")) == "low_cost"
    ]
    display_name = get_person_display_name(person)
    report_display_name = get_person_report_name(person)
    printed_at = datetime.now(ZoneInfo("Europe/Kyiv")).strftime("%d.%m.%Y")
    tenant_key, branding, branding_storage = resolve_branding_for_request(request)

    return templates.TemplateResponse(
        request=request,
        name="admin_person_detail.html",
        context={
            "person": person,
            "person_field_rows": build_person_field_rows(person),
            "display_name": display_name,
            "report_display_name": report_display_name,
            "assigned_assets": assigned_assets,
            "assigned_standard_assets": assigned_standard_assets,
            "assigned_low_cost_assets": assigned_low_cost_assets,
            "printed_at": printed_at,
            "branding": branding,
            "branding_storage": branding_storage,
            "tenant_key": tenant_key,
            "branding_logo_url": get_branding_logo_url(branding),
            "flash": pop_flash(request),
            "active_page": "people",
            "page_title": display_name,
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.get("/admin/people/{person_id}/offboard", response_class=HTMLResponse)
def admin_person_offboard(request: Request, person_id: int):
    redirect = require_admin(request)
    if redirect:
        return redirect

    person = get_person_by_id(person_id)
    if not person:
        raise HTTPException(status_code=404, detail="Person not found")

    context = build_person_offboarding_context(person)
    return templates.TemplateResponse(
        request=request,
        name="admin_person_offboard.html",
        context={
            "person": person,
            "display_name": get_person_display_name(person),
            "flash": pop_flash(request),
            "active_page": "people",
            "page_title": f"Offboard {get_person_display_name(person)}",
            "admin_username": request.session.get("admin_username"),
            **context,
        },
    )


@app.post("/admin/people/{person_id}/offboard")
async def admin_person_offboard_submit(request: Request, person_id: int):
    redirect = require_admin(request)
    if redirect:
        return redirect

    person = get_person_by_id(person_id)
    if not person:
        raise HTTPException(status_code=404, detail="Person not found")

    form = await request.form()
    try:
        result = apply_person_offboarding(person, form)
    except Exception as error:
        context = build_person_offboarding_context(person)
        return templates.TemplateResponse(
            request=request,
            name="admin_person_offboard.html",
            context={
                "person": person,
                "display_name": get_person_display_name(person),
                "flash": {"level": "error", "message": f"Offboarding could not be applied: {error}"},
                "active_page": "people",
                "page_title": f"Offboard {get_person_display_name(person)}",
                "admin_username": request.session.get("admin_username"),
                **context,
            },
            status_code=400,
        )

    set_flash(
        request,
        "success",
        f"Employee {get_person_display_name(person)} was marked inactive. "
        f"Reassigned assets: {result['reassigned']}.",
    )
    return RedirectResponse(url=f"/admin/people/{person_id}", status_code=303)


@app.get("/admin/people/{person_id}/edit", response_class=HTMLResponse)
def admin_person_edit(request: Request, person_id: int):
    redirect = require_admin(request)
    if redirect:
        return redirect

    person = get_person_by_id(person_id)
    if not person:
        raise HTTPException(status_code=404, detail="Person not found")

    return templates.TemplateResponse(
        request=request,
        name="admin_person_edit.html",
        context={
            "person": person,
            "display_name": get_person_display_name(person),
            "person_field_rows": build_person_field_rows(person),
            "person_edit_fields": build_person_edit_fields(person),
            "flash": pop_flash(request),
            "active_page": "people",
            "page_title": f"Edit {get_person_display_name(person)}",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.post("/admin/people/{person_id}/edit")
async def admin_person_edit_submit(request: Request, person_id: int):
    redirect = require_admin(request)
    if redirect:
        return redirect

    person = get_person_by_id(person_id)
    if not person:
        raise HTTPException(status_code=404, detail="Person not found")

    form = await request.form()
    update_data = {}
    for field in build_person_edit_fields(person):
        raw_value = form.get(field["name"], "")
        normalized = str(raw_value).strip() if raw_value is not None else ""
        if field["name"] == "is_active":
            update_data[field["name"]] = normalized.casefold() == "true"
            continue
        update_data[field["name"]] = normalized or None

    if not update_data.get("name") and not update_data.get("name_eng"):
        return templates.TemplateResponse(
            request=request,
            name="admin_person_edit.html",
            context={
                "person": {**person, **update_data},
                "display_name": get_person_display_name({**person, **update_data}),
                "person_field_rows": build_person_field_rows({**person, **update_data}),
                "person_edit_fields": build_person_edit_fields({**person, **update_data}),
                "flash": {"level": "error", "message": "Keep at least one employee name: local or English."},
                "active_page": "people",
                "page_title": f"Edit {get_person_display_name(person)}",
                "admin_username": request.session.get("admin_username"),
            },
            status_code=400,
        )

    try:
        supabase.table("persons").update(update_data).eq("person_id", person_id).execute()
        audit_log_field_changes(
            entity_type="Employee",
            entity_id=person_id,
            entity_label=get_person_display_name({**person, **update_data}),
            old_record=person,
            new_record=update_data,
            fields=list(update_data.keys()),
            request=request,
        )
    except Exception as exc:
        return templates.TemplateResponse(
            request=request,
            name="admin_person_edit.html",
            context={
                "person": {**person, **update_data},
                "display_name": get_person_display_name({**person, **update_data}),
                "person_field_rows": build_person_field_rows({**person, **update_data}),
                "person_edit_fields": build_person_edit_fields({**person, **update_data}),
                "flash": {"level": "error", "message": f"Employee could not be updated: {exc}"},
                "active_page": "people",
                "page_title": f"Edit {get_person_display_name(person)}",
                "admin_username": request.session.get("admin_username"),
            },
            status_code=400,
        )

    set_flash(request, "success", f"Employee {get_person_display_name({**person, **update_data})} was updated.")
    return RedirectResponse(url=f"/admin/people/{person_id}", status_code=303)


@app.get("/admin/branding/logo")
def admin_branding_logo(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    _, branding, _ = resolve_branding_for_request(request)
    logo_path = branding.get("logo_path") or ""
    if not logo_path or logo_path.startswith("data:image/") or not os.path.exists(logo_path):
        raise HTTPException(status_code=404, detail="Logo not found")

    return FileResponse(logo_path)


@app.get("/admin/branding", response_class=HTMLResponse)
def admin_branding(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    tenant_key, branding, branding_storage = resolve_branding_for_request(request)
    return templates.TemplateResponse(
        request=request,
        name="admin_branding.html",
        context={
            "branding": branding,
            "branding_storage": branding_storage,
            "tenant_key": tenant_key,
            "branding_logo_url": get_branding_logo_url(branding),
            "flash": pop_flash(request),
            "active_page": "branding",
            "page_title": "Branding",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.post("/admin/branding")
def admin_branding_save(
    request: Request,
    company_name: str = Form(""),
    report_title: str = Form(""),
    report_subtitle: str = Form(""),
    report_theme: str = Form("classic"),
    primary_color: str = Form(""),
    accent_color: str = Form(""),
    footer_note: str = Form(""),
    issuer_label: str = Form(""),
    issuer_signature_label: str = Form(""),
    receiver_label: str = Form(""),
    receiver_signature_label: str = Form(""),
    logo_file: Optional[UploadFile] = File(None),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    tenant_key, branding, _ = resolve_branding_for_request(request)
    branding.update(
        {
            "company_name": company_name.strip() or get_default_branding_settings()["company_name"],
            "report_title": report_title.strip() or get_default_branding_settings()["report_title"],
            "report_subtitle": report_subtitle.strip() or get_default_branding_settings()["report_subtitle"],
            "report_theme": report_theme.strip() if report_theme.strip() in {"classic", "corporate", "compact", "help_standard"} else get_default_branding_settings()["report_theme"],
            "primary_color": primary_color.strip() or get_default_branding_settings()["primary_color"],
            "accent_color": accent_color.strip() or get_default_branding_settings()["accent_color"],
            "footer_note": footer_note.strip() or get_default_branding_settings()["footer_note"],
            "issuer_label": issuer_label.strip() or get_default_branding_settings()["issuer_label"],
            "issuer_signature_label": issuer_signature_label.strip() or get_default_branding_settings()["issuer_signature_label"],
            "receiver_label": receiver_label.strip() or get_default_branding_settings()["receiver_label"],
            "receiver_signature_label": receiver_signature_label.strip() or get_default_branding_settings()["receiver_signature_label"],
        }
    )

    try:
        if logo_file and (logo_file.filename or "").strip():
            branding["logo_path"] = save_branding_logo(tenant_key, logo_file, branding.get("logo_path") or "")
    except ValueError as error:
        set_flash(request, "error", str(error))
        return RedirectResponse(url="/admin/branding", status_code=303)

    storage_backend = save_branding_settings(tenant_key, branding)
    set_flash(request, "success", f"Branding settings were updated for tenant '{tenant_key}' using {storage_backend} storage.")
    return RedirectResponse(url="/admin/branding", status_code=303)


@app.get("/admin/reference-data", response_class=HTMLResponse)
def admin_reference_data(
    request: Request,
    edit_project_id: Optional[int] = None,
    edit_donor_id: Optional[int] = None,
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    projects = list_projects()
    donors = list_donors()
    edit_project = get_project_by_id(edit_project_id) if edit_project_id else None
    edit_donor = get_donor_by_id(edit_donor_id) if edit_donor_id else None

    return templates.TemplateResponse(
        request=request,
        name="admin_reference_data.html",
        context={
            "projects": projects,
            "donors": donors,
            "edit_project": edit_project,
            "edit_donor": edit_donor,
            "project_form": get_project_form_values(edit_project),
            "donor_form": get_donor_form_values(edit_donor),
            "flash": pop_flash(request),
            "active_page": "reference_data",
            "page_title": "Reference Data",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.post("/admin/reference-data/projects")
def admin_reference_data_project_create(
    request: Request,
    project_number: str = Form(""),
    project_name: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
    status: str = Form(""),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    project_number = project_number.strip().upper()
    if not project_number:
        set_flash(request, "error", "Project number is required.")
        return RedirectResponse(url="/admin/reference-data#projects", status_code=303)

    payload = {
        "project_id": get_next_numeric_id("projects", "project_id"),
        "project_number": project_number,
        "project_name": project_name.strip() or None,
        "start_date": start_date.strip() or None,
        "end_date": end_date.strip() or None,
        "status": status.strip() or None,
    }

    try:
        supabase.table("projects").insert(payload).execute()
    except Exception as error:
        set_flash(request, "error", describe_reference_data_error(error, "Project", "project_number"))
        return RedirectResponse(url="/admin/reference-data#projects", status_code=303)

    set_flash(request, "success", f"Project {project_number} was added.")
    return RedirectResponse(url="/admin/reference-data#projects", status_code=303)


@app.post("/admin/reference-data/projects/{project_id}")
def admin_reference_data_project_update(
    request: Request,
    project_id: int,
    project_number: str = Form(""),
    project_name: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
    status: str = Form(""),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    project = get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    project_number = project_number.strip().upper()
    if not project_number:
        set_flash(request, "error", "Project number is required.")
        return RedirectResponse(url=f"/admin/reference-data?edit_project_id={project_id}#projects", status_code=303)

    payload = {
        "project_number": project_number,
        "project_name": project_name.strip() or None,
        "start_date": start_date.strip() or None,
        "end_date": end_date.strip() or None,
        "status": status.strip() or None,
    }

    try:
        supabase.table("projects").update(payload).eq("project_id", project_id).execute()
    except Exception as error:
        set_flash(request, "error", describe_reference_data_error(error, "Project", "project_number"))
        return RedirectResponse(url=f"/admin/reference-data?edit_project_id={project_id}#projects", status_code=303)

    set_flash(request, "success", f"Project {project_number} was updated.")
    return RedirectResponse(url="/admin/reference-data#projects", status_code=303)


@app.post("/admin/reference-data/donors")
def admin_reference_data_donor_create(
    request: Request,
    donor_name: str = Form(""),
    contact_person: str = Form(""),
    contact_email: str = Form(""),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    donor_name = donor_name.strip()
    if not donor_name:
        set_flash(request, "error", "Donor name is required.")
        return RedirectResponse(url="/admin/reference-data#donors", status_code=303)

    payload = {
        "donor_id": get_next_numeric_id("donors", "donor_id"),
        "donor_name": donor_name,
        "contact_person": contact_person.strip() or None,
        "contact_email": contact_email.strip() or None,
    }

    try:
        supabase.table("donors").insert(payload).execute()
    except Exception as error:
        set_flash(request, "error", describe_reference_data_error(error, "Donor", "donor_name"))
        return RedirectResponse(url="/admin/reference-data#donors", status_code=303)

    set_flash(request, "success", f"Donor {donor_name} was added.")
    return RedirectResponse(url="/admin/reference-data#donors", status_code=303)


@app.post("/admin/reference-data/donors/{donor_id}")
def admin_reference_data_donor_update(
    request: Request,
    donor_id: int,
    donor_name: str = Form(""),
    contact_person: str = Form(""),
    contact_email: str = Form(""),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    donor = get_donor_by_id(donor_id)
    if not donor:
        raise HTTPException(status_code=404, detail="Donor not found")

    donor_name = donor_name.strip()
    if not donor_name:
        set_flash(request, "error", "Donor name is required.")
        return RedirectResponse(url=f"/admin/reference-data?edit_donor_id={donor_id}#donors", status_code=303)

    payload = {
        "donor_name": donor_name,
        "contact_person": contact_person.strip() or None,
        "contact_email": contact_email.strip() or None,
    }

    try:
        supabase.table("donors").update(payload).eq("donor_id", donor_id).execute()
    except Exception as error:
        set_flash(request, "error", describe_reference_data_error(error, "Donor", "donor_name"))
        return RedirectResponse(url=f"/admin/reference-data?edit_donor_id={donor_id}#donors", status_code=303)

    set_flash(request, "success", f"Donor {donor_name} was updated.")
    return RedirectResponse(url="/admin/reference-data#donors", status_code=303)


@app.get("/admin/assets/{asset_id}", response_class=HTMLResponse)
def admin_asset_detail(request: Request, asset_id: int):
    redirect = require_admin(request)
    if redirect:
        return redirect

    asset = get_asset_by_id(asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    assignment_history = get_assignment_history(asset_id)

    return templates.TemplateResponse(
        request=request,
        name="admin_asset_detail.html",
        context={
            "asset": asset,
            "assignment_history": assignment_history,
            "flash": pop_flash(request),
            "active_page": "assets",
            "page_title": f"Asset {asset.get('asset_tag_number')}",
            "admin_username": request.session.get("admin_username"),
            "usage_type_options": ASSET_USAGE_TYPE_OPTIONS,
            **get_assignment_form_context(asset),
            **get_asset_project_form_context(asset_id),
            **get_asset_payment_context(asset_id),
            **get_asset_transfer_context(asset_id),
        },
    )


@app.post("/admin/assets/{asset_id}/edit")
def admin_asset_edit(
    request: Request,
    asset_id: int,
    usage_type: str = Form("standard"),
    item_description: str = Form(""),
    brand_make: str = Form(""),
    model: str = Form(""),
    asset_classification: str = Form(""),
    asset_sub_classification: str = Form(""),
    quantity: str = Form(""),
    purchase_price: str = Form(""),
    currency: str = Form(""),
    serial_number: str = Form(""),
    current_status: str = Form(""),
    remarks: str = Form(""),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    asset = get_asset_by_id(asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    try:
        update_data = {
            "usage_type": normalize_asset_usage_type(usage_type, asset.get("asset_tag_number")),
            "item_description": item_description.strip() or None,
            "brand_make": brand_make.strip() or None,
            "model": model.strip() or None,
            "asset_classification": asset_classification.strip() or None,
            "asset_sub_classification": asset_sub_classification.strip() or None,
            "quantity": parse_int_field(quantity),
            "purchase_price": parse_float_field(purchase_price),
            "currency": currency.strip() or None,
            "serial_chassis_number": serial_number.strip() or None,
            "current_status": current_status.strip() or None,
            "remarks": remarks.strip() or None,
        }
    except ValueError:
        set_flash(request, "error", "Quantity must be an integer and purchase price must be a number.")
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    (
        supabase.table("assets")
        .update(update_data)
        .eq("asset_id", asset_id)
        .execute()
    )
    audit_log_field_changes(
        entity_type="Asset",
        entity_id=asset_id,
        entity_label=asset.get("asset_tag_number"),
        old_record=asset,
        new_record=update_data,
        fields=list(update_data.keys()),
        request=request,
    )

    set_flash(request, "success", "Asset details were updated.")
    return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)


@app.post("/admin/assets/{asset_id}/assignment")
def admin_asset_assignment_update(
    request: Request,
    asset_id: int,
    person_id: str = Form(""),
    location_id: str = Form(""),
    assignment_date: str = Form(""),
    status: str = Form(""),
    notes: str = Form(""),
    handover_condition: str = Form(""),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    asset = get_asset_by_id(asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    person_id = person_id.strip()
    location_id = location_id.strip()
    assignment_date = assignment_date.strip()
    status = status.strip()
    notes = notes.strip()
    handover_condition = handover_condition.strip()
    parsed_person_id = int(person_id) if person_id else None
    parsed_location_id = int(location_id) if location_id else None

    if not assignment_date:
        set_flash(request, "error", "Assignment date is required.")
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    if parsed_person_id and not parsed_location_id:
        set_flash(request, "error", "Location is required when a responsible person is selected.")
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    current_assignment = asset.get("current_assignment")
    old_responsible = (current_assignment or {}).get("responsible_person") or "-"

    if not parsed_person_id and not parsed_location_id:
        if current_assignment:
            close_current_assignments(asset_id, assignment_date)
            if status:
                supabase.table("assets").update({"current_status": status}).eq("asset_id", asset_id).execute()
            transfer_id = create_asset_transfer_from_assignment_change(
                asset=asset,
                from_assignment=current_assignment,
                to_person_id=None,
                transfer_date=assignment_date,
                transfer_reason="Assignment closed in web app",
                status=status or current_assignment.get("status") or asset.get("current_status"),
                condition=handover_condition or current_assignment.get("handover_condition"),
            )
            audit_log_event(
                entity_type="Transfer",
                entity_id=transfer_id or asset_id,
                entity_label=asset.get("asset_tag_number"),
                action="transferred",
                old_value=old_responsible,
                new_value="Warehouse",
                summary=f"Assignment closed for {asset.get('asset_tag_number')}: {old_responsible} -> Warehouse",
                request=request,
                event_date=assignment_date,
                event_key=f"asset_transfer:{transfer_id}" if transfer_id else None,
            )
            set_flash(request, "success", "Current assignment was closed and the asset is now unassigned.")
        else:
            if status:
                supabase.table("assets").update({"current_status": status}).eq("asset_id", asset_id).execute()
            set_flash(request, "success", "Asset is already unassigned.")
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    new_assignment = {
        "asset_id": asset_id,
        "person_id": parsed_person_id,
        "location_id": parsed_location_id,
        "assignment_date": assignment_date,
        "return_date": None,
        "status": status or None,
        "notes": notes or None,
        "handover_condition": handover_condition or None,
    }

    if (
        current_assignment
        and current_assignment.get("person_id") == parsed_person_id
        and current_assignment.get("location_id") == parsed_location_id
    ):
        (
            supabase.table("asset_assignments")
            .update({
                "assignment_date": assignment_date,
                "status": status or None,
                "notes": notes or None,
                "handover_condition": handover_condition or None,
            })
            .eq("assignment_id", current_assignment["assignment_id"])
            .execute()
        )
        try:
            if status:
                supabase.table("assets").update({"current_status": status}).eq("asset_id", asset_id).execute()
        except Exception as exc:
            set_flash(request, "error", describe_assignment_update_error(exc))
            return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)
        set_flash(request, "success", "Current assignment was updated.")
        audit_log_event(
            entity_type="Assignment",
            entity_id=asset_id,
            entity_label=asset.get("asset_tag_number"),
            action="updated",
            summary=f"Assignment updated for {asset.get('asset_tag_number')}",
            request=request,
        )
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    try:
        close_current_assignments(asset_id, assignment_date)
        supabase.table("asset_assignments").insert(new_assignment).execute()
        if status:
            supabase.table("assets").update({"current_status": status}).eq("asset_id", asset_id).execute()
    except Exception as exc:
        set_flash(request, "error", describe_assignment_update_error(exc))
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    set_flash(request, "success", "Assignment was updated.")
    target_person = get_person_by_id(parsed_person_id) if parsed_person_id else None
    new_responsible = get_person_display_name(target_person) if target_person else "Warehouse"
    transfer_id = create_asset_transfer_from_assignment_change(
        asset=asset,
        from_assignment=current_assignment,
        to_person_id=parsed_person_id,
        transfer_date=assignment_date,
        transfer_reason="Assignment changed in web app",
        status=status or (current_assignment or {}).get("status") or asset.get("current_status"),
        condition=handover_condition or (current_assignment or {}).get("handover_condition"),
    )
    audit_log_event(
        entity_type="Transfer",
        entity_id=transfer_id or asset_id,
        entity_label=asset.get("asset_tag_number"),
        action="transferred",
        field_name="responsible_person",
        old_value=old_responsible,
        new_value=new_responsible,
        summary=f"{asset.get('asset_tag_number')} reassigned: {old_responsible} -> {new_responsible}",
        request=request,
        event_date=assignment_date,
        event_key=f"asset_transfer:{transfer_id}" if transfer_id else None,
    )
    return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)


@app.post("/admin/assets/{asset_id}/payments")
def admin_asset_payment_create(
    request: Request,
    asset_id: int,
    payment_date: str = Form(""),
    payment_amount: str = Form(""),
    currency: str = Form(""),
    payment_status: str = Form("paid"),
    notes: str = Form(""),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    asset = get_asset_by_id(asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    try:
        payload = build_asset_payment_payload(
            asset_id=asset_id,
            payment_number=None,
            payment_date=payment_date,
            payment_amount=payment_amount,
            currency=currency or asset.get("currency") or "EUR",
            payment_status=payment_status,
            notes=notes,
        )
        supabase.table("asset_payments").insert(payload).execute()
        audit_log_event(
            entity_type="Payment",
            entity_id=asset_id,
            entity_label=asset.get("asset_tag_number"),
            action="created",
            summary=f"Added payment for {asset.get('asset_tag_number')}: {payload.get('payment_amount') or '-'} {payload.get('currency') or ''}",
            request=request,
        )
    except Exception as error:
        set_flash(request, "error", describe_asset_payment_error(error))
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    set_flash(request, "success", "Payment was added.")
    return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)


@app.post("/admin/assets/{asset_id}/payments/{payment_id}")
def admin_asset_payment_update(
    request: Request,
    asset_id: int,
    payment_id: int,
    payment_number: str = Form(""),
    payment_date: str = Form(""),
    payment_amount: str = Form(""),
    currency: str = Form(""),
    payment_status: str = Form("paid"),
    notes: str = Form(""),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    asset = get_asset_by_id(asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    try:
        old_payment_response = (
            supabase.table("asset_payments")
            .select("*")
            .eq("payment_id", payment_id)
            .eq("asset_id", asset_id)
            .limit(1)
            .execute()
        )
        old_payment = (old_payment_response.data or [{}])[0]
        parsed_number = int(payment_number.strip()) if payment_number.strip() else get_next_payment_number(asset_id)
        payload = build_asset_payment_payload(
            asset_id=asset_id,
            payment_number=parsed_number,
            payment_date=payment_date,
            payment_amount=payment_amount,
            currency=currency or asset.get("currency") or "EUR",
            payment_status=payment_status,
            notes=notes,
        )
        payload.pop("asset_id", None)
        supabase.table("asset_payments").update(payload).eq("payment_id", payment_id).eq("asset_id", asset_id).execute()
        audit_log_field_changes(
            entity_type="Payment",
            entity_id=payment_id,
            entity_label=asset.get("asset_tag_number"),
            old_record=old_payment,
            new_record=payload,
            fields=["payment_number", "payment_date", "payment_amount", "currency", "payment_status", "notes"],
            request=request,
        )
    except Exception as error:
        set_flash(request, "error", describe_asset_payment_error(error))
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    set_flash(request, "success", "Payment was updated.")
    return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)


@app.post("/admin/assets/{asset_id}/payments/{payment_id}/delete")
def admin_asset_payment_delete(request: Request, asset_id: int, payment_id: int):
    redirect = require_admin(request)
    if redirect:
        return redirect

    asset = get_asset_by_id(asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    try:
        old_payment_response = (
            supabase.table("asset_payments")
            .select("*")
            .eq("payment_id", payment_id)
            .eq("asset_id", asset_id)
            .limit(1)
            .execute()
        )
        old_payment = (old_payment_response.data or [{}])[0]
        supabase.table("asset_payments").delete().eq("payment_id", payment_id).eq("asset_id", asset_id).execute()
        audit_log_event(
            entity_type="Payment",
            entity_id=payment_id,
            entity_label=asset.get("asset_tag_number"),
            action="deleted",
            summary=f"Removed payment for {asset.get('asset_tag_number')}: {old_payment.get('payment_amount') or '-'} {old_payment.get('currency') or ''}",
            request=request,
        )
    except Exception as error:
        set_flash(request, "error", describe_asset_payment_error(error))
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    set_flash(request, "success", "Payment was removed.")
    return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)


@app.post("/admin/assets/{asset_id}/funding")
def admin_asset_project_create(
    request: Request,
    asset_id: int,
    project_id: str = Form(""),
    donor_id: str = Form(""),
    allocation_percent: str = Form(""),
    allocation_amount: str = Form(""),
    currency: str = Form(""),
    funding_note: str = Form(""),
    is_primary: Optional[str] = Form(None),
    is_current: Optional[str] = Form(None),
    is_purchase_origin: Optional[str] = Form(None),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    asset = get_asset_by_id(asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    project_id = project_id.strip()
    donor_id = donor_id.strip()
    currency = currency.strip()
    funding_note = funding_note.strip()

    if not project_id:
        set_flash(request, "error", "Project is required for funding allocation.")
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    try:
        parsed_percent = safe_parse_percentage(allocation_percent)
        parsed_amount = parse_float_field(allocation_amount)
    except ValueError as error:
        set_flash(request, "error", str(error))
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    if parsed_percent is not None and get_asset_project_total_percent(asset_id) + parsed_percent > 100.001:
        set_flash(request, "error", "Total allocated percent cannot be greater than 100%.")
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    payload = {
        "asset_id": asset_id,
        "project_id": int(project_id),
        "donor_id": int(donor_id) if donor_id else None,
        "allocation_percent": parsed_percent,
        "allocation_amount": parsed_amount,
        "currency": currency or None,
        "funding_note": funding_note or None,
        "is_primary": is_primary == "on",
        "is_current": is_current == "on",
        "is_purchase_origin": is_purchase_origin == "on",
    }
    if not asset_project_purchase_origin_supported():
        payload.pop("is_purchase_origin", None)

    try:
        if payload["is_primary"]:
            supabase.table("asset_projects").update({"is_primary": False}).eq("asset_id", asset_id).execute()
        supabase.table("asset_projects").insert(payload).execute()
        audit_log_event(
            entity_type="Project",
            entity_id=asset_id,
            entity_label=asset.get("asset_tag_number"),
            action="created",
            summary=f"Added project funding for {asset.get('asset_tag_number')}: project #{payload.get('project_id')}",
            request=request,
        )
    except Exception as error:
        set_flash(request, "error", describe_asset_project_error(error))
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    set_flash(request, "success", "Project funding allocation was added.")
    return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)


@app.post("/admin/assets/{asset_id}/funding/{asset_project_id}")
def admin_asset_project_update(
    request: Request,
    asset_id: int,
    asset_project_id: int,
    project_id: str = Form(""),
    donor_id: str = Form(""),
    allocation_percent: str = Form(""),
    allocation_amount: str = Form(""),
    currency: str = Form(""),
    funding_note: str = Form(""),
    is_primary: Optional[str] = Form(None),
    is_current: Optional[str] = Form(None),
    is_purchase_origin: Optional[str] = Form(None),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    asset = get_asset_by_id(asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    project_id = project_id.strip()
    donor_id = donor_id.strip()
    currency = currency.strip()
    funding_note = funding_note.strip()

    if not project_id:
        set_flash(request, "error", "Project is required for funding allocation.")
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    try:
        parsed_percent = safe_parse_percentage(allocation_percent)
        parsed_amount = parse_float_field(allocation_amount)
    except ValueError as error:
        set_flash(request, "error", str(error))
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    if parsed_percent is not None and get_asset_project_total_percent(asset_id, exclude_asset_project_id=asset_project_id) + parsed_percent > 100.001:
        set_flash(request, "error", "Total allocated percent cannot be greater than 100%.")
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    payload = {
        "project_id": int(project_id),
        "donor_id": int(donor_id) if donor_id else None,
        "allocation_percent": parsed_percent,
        "allocation_amount": parsed_amount,
        "currency": currency or None,
        "funding_note": funding_note or None,
        "is_primary": is_primary == "on",
        "is_current": is_current == "on",
        "is_purchase_origin": is_purchase_origin == "on",
    }
    if not asset_project_purchase_origin_supported():
        payload.pop("is_purchase_origin", None)

    try:
        old_project_response = (
            supabase.table("asset_projects")
            .select("*")
            .eq("asset_project_id", asset_project_id)
            .eq("asset_id", asset_id)
            .limit(1)
            .execute()
        )
        old_project = (old_project_response.data or [{}])[0]
        if payload["is_primary"]:
            supabase.table("asset_projects").update({"is_primary": False}).eq("asset_id", asset_id).execute()
        supabase.table("asset_projects").update(payload).eq("asset_project_id", asset_project_id).eq("asset_id", asset_id).execute()
        audit_log_field_changes(
            entity_type="Project",
            entity_id=asset_project_id,
            entity_label=asset.get("asset_tag_number"),
            old_record=old_project,
            new_record=payload,
            fields=list(payload.keys()),
            request=request,
        )
    except Exception as error:
        set_flash(request, "error", describe_asset_project_error(error))
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    set_flash(request, "success", "Project funding allocation was updated.")
    return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)


@app.post("/admin/assets/{asset_id}/funding/{asset_project_id}/delete")
def admin_asset_project_delete(request: Request, asset_id: int, asset_project_id: int):
    redirect = require_admin(request)
    if redirect:
        return redirect

    asset = get_asset_by_id(asset_id)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    try:
        old_project_response = (
            supabase.table("asset_projects")
            .select("*")
            .eq("asset_project_id", asset_project_id)
            .eq("asset_id", asset_id)
            .limit(1)
            .execute()
        )
        old_project = (old_project_response.data or [{}])[0]
        supabase.table("asset_projects").delete().eq("asset_project_id", asset_project_id).eq("asset_id", asset_id).execute()
        audit_log_event(
            entity_type="Project",
            entity_id=asset_project_id,
            entity_label=asset.get("asset_tag_number"),
            action="deleted",
            summary=f"Removed project funding for {asset.get('asset_tag_number')}: project #{old_project.get('project_id') or '-'}",
            request=request,
        )
    except Exception as error:
        set_flash(request, "error", describe_asset_project_error(error))
        return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)

    set_flash(request, "success", "Project funding allocation was removed.")
    return RedirectResponse(url=f"/admin/assets/{asset_id}", status_code=303)


@app.get("/admin/reports", response_class=HTMLResponse)
def admin_reports(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    assets = list_assets()
    summary = build_asset_summary(assets)

    return templates.TemplateResponse(
        request=request,
        name="admin_reports.html",
        context={
            "summary": summary,
            "active_page": "reports",
            "page_title": "Admin Reports",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.get("/admin/reports/export/{report_name}")
def admin_reports_export(request: Request, report_name: str):
    redirect = require_admin(request)
    if redirect:
        return redirect

    assets = list_assets()

    if report_name == "cities":
        rows = summarize_assets_by_field(assets, "city", "Unknown")
        return csv_response(
            "assets-by-city.csv",
            ["label", "count"],
            [{"label": row["label"], "count": row["count"]} for row in rows],
        )

    if report_name == "departments":
        rows = summarize_assets_by_field(assets, "department", "Unknown")
        return csv_response(
            "assets-by-department.csv",
            ["label", "count"],
            [{"label": row["label"], "count": row["count"]} for row in rows],
        )

    if report_name == "people":
        rows = summarize_assets_by_field(assets, "responsible_person", "Unassigned")
        return csv_response(
            "assets-by-responsible-person.csv",
            ["label", "count"],
            [{"label": row["label"], "count": row["count"]} for row in rows],
        )

    if report_name == "assignments":
        return csv_response(
            "assignment-audit.csv",
            [
                "asset_tag_number",
                "item_description",
                "brand_make",
                "model",
                "effective_status",
                "current_status",
                "responsible_person",
                "department",
                "city",
                "location_name",
                "assignment_date",
                "assignment_status",
                "assignment_notes",
            ],
            build_assignment_audit_rows(assets),
        )

    raise HTTPException(status_code=404, detail="Report not found")


@app.get("/admin/sync", response_class=HTMLResponse)
def admin_sync(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    sync_state = load_sync_state()
    sync_rules = [
        "Supabase is the operational working database.",
        "The Excel inventory file remains the official control file.",
        "Imports from Excel should go through controlled sync steps, not direct blind overwrite.",
        "Exports to Excel should use the official workbook as a formatting template.",
        "Before applying sync changes, the system should show a comparison preview.",
    ]

    return templates.TemplateResponse(
        request=request,
        name="admin_sync.html",
        context={
            "sync_rules": sync_rules,
            "excel_file_name": sync_state.get("file_name") or "No workbook uploaded yet",
            "sync_state": sync_state,
            "preview": sync_state.get("preview"),
            "flash": pop_flash(request),
            "active_page": "sync",
            "page_title": "Admin Sync",
            "admin_username": request.session.get("admin_username"),
        },
    )


@app.post("/admin/sync/upload")
async def admin_sync_upload(request: Request, excel_file: UploadFile = File(...)):
    redirect = require_admin(request)
    if redirect:
        return redirect

    filename = (excel_file.filename or "").strip()
    if not filename.lower().endswith(".xlsx"):
        set_flash(request, "error", "Upload an .xlsx workbook for Excel synchronization.")
        return RedirectResponse(url="/admin/sync", status_code=303)

    ensure_sync_storage()
    file_bytes = await excel_file.read()
    if not file_bytes:
        set_flash(request, "error", "The uploaded Excel file is empty.")
        return RedirectResponse(url="/admin/sync", status_code=303)

    with open(SYNC_WORKBOOK_PATH, "wb") as file:
        file.write(file_bytes)

    try:
        excel_records = load_excel_sync_rows(SYNC_WORKBOOK_PATH)
        transfer_records = load_excel_transfer_log_rows(SYNC_WORKBOOK_PATH)
        preview = build_sync_preview(excel_records, list_asset_records(), transfer_records)
    except ValueError as error:
        set_flash(request, "error", str(error))
        return RedirectResponse(url="/admin/sync", status_code=303)
    except Exception as error:
        set_flash(request, "error", f"Excel sync preview failed: {error}")
        return RedirectResponse(url="/admin/sync", status_code=303)

    save_sync_state(
        {
            "file_name": filename,
            "uploaded_at": datetime.now(ZoneInfo("Europe/Kyiv")).strftime("%d.%m.%Y %H:%M"),
            "preview": preview,
        }
    )
    set_flash(
        request,
        "success",
        (
            f"Preview ready: {preview['summary']['new_records']} new, "
            f"{preview['summary']['changed_records']} changed, "
            f"{preview['summary']['unchanged_records']} unchanged, "
            f"{preview.get('transfer_log', {}).get('summary', {}).get('new_records', 0)} transfer records."
        ),
    )
    return RedirectResponse(url="/admin/sync", status_code=303)


@app.post("/admin/sync/apply")
def admin_sync_apply(
    request: Request,
    selected_new_assets: list[str] = Form([]),
    selected_changed_assets: list[str] = Form([]),
    selected_transfers: list[str] = Form([]),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    sync_state = load_sync_state()
    preview = sync_state.get("preview")
    if not preview:
        set_flash(request, "error", "No sync preview is available. Upload an Excel file first.")
        return RedirectResponse(url="/admin/sync", status_code=303)

    selected_preview = filter_sync_preview(preview, selected_new_assets, selected_changed_assets, selected_transfers)
    if not selected_preview["new_records"] and not selected_preview["changed_records"] and not selected_preview.get("transfer_log", {}).get("new_records"):
        set_flash(request, "error", "No sync records were selected for apply.")
        return RedirectResponse(url="/admin/sync", status_code=303)

    try:
        result = apply_sync_preview(selected_preview)
    except Exception as error:
        set_flash(request, "error", f"Could not apply sync changes: {error}")
        return RedirectResponse(url="/admin/sync", status_code=303)

    sync_state["last_applied_at"] = datetime.now(ZoneInfo("Europe/Kyiv")).strftime("%d.%m.%Y %H:%M")
    sync_state["last_apply_result"] = result
    sync_state["preview"] = None
    save_sync_state(sync_state)
    set_flash(
        request,
        "success",
        (
            f"Excel sync applied: {result['inserted']} new assets inserted, "
            f"{result['updated']} assets updated, "
            f"{result['assignment_updated']} assignment changes, "
            f"{result['project_updated']} project changes, "
            f"{result['payment_updated']} payment records, "
            f"{result['transfer_updated']} transfer records."
        ),
    )
    return RedirectResponse(url="/admin/sync", status_code=303)


@app.post("/admin/sync/export")
def admin_sync_export(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    sync_state = load_sync_state()
    try:
        result = export_supabase_to_excel()
    except ValueError as error:
        set_flash(request, "error", str(error))
        return RedirectResponse(url="/admin/sync", status_code=303)
    except Exception as error:
        set_flash(request, "error", f"Could not export Supabase data to Excel: {error}")
        return RedirectResponse(url="/admin/sync", status_code=303)

    sync_state["last_exported_at"] = result["exported_at"]
    sync_state["last_export_result"] = {
        "updated_rows": result["updated_rows"],
        "appended_rows": result["appended_rows"],
        "written_cells": result["written_cells"],
        "exported_records": result["exported_records"],
        "standard_exported_records": result.get("standard_exported_records", 0),
        "low_cost_exported_records": result.get("low_cost_exported_records", 0),
        "transfer_exported_records": result.get("transfer_exported_records", 0),
    }
    save_sync_state(sync_state)

    filename = f"supabase_inventory_export_{datetime.now(ZoneInfo('Europe/Kyiv')).strftime('%Y%m%d_%H%M')}.xlsx"
    return FileResponse(
        result["path"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@app.post("/webhook")
async def telegram_webhook(update: dict = Body(...)):
    try:
        print("UPDATE:", update)

        if "message" not in update:
            return {"ok": True}

        message = update["message"]
        chat_id = message["chat"]["id"]
        telegram_user = message.get("from") or {}
        text = ""

        if "contact" in message:
            contact = message.get("contact") or {}
            if contact.get("user_id") != telegram_user.get("id"):
                send_telegram_message(
                    chat_id,
                    "Please share your own phone number using the Telegram button.",
                    reply_markup=AUTH_KEYBOARD,
                )
                return {"ok": True}

            person = find_person_by_phone(contact.get("phone_number") or "")
            if not person:
                send_telegram_message(
                    chat_id,
                    "Phone number was not found in the employee directory. Please contact the administrator.",
                    reply_markup=AUTH_KEYBOARD,
                )
                return {"ok": True}
            if not is_person_active(person):
                send_telegram_message(
                    chat_id,
                    "Your employee profile is inactive. Please contact the administrator.",
                    reply_markup=AUTH_KEYBOARD,
                )
                return {"ok": True}

            save_person_telegram_identity(person["person_id"], telegram_user, contact.get("phone_number"))
            send_telegram_message(
                chat_id,
                f"Authorization successful. Welcome, {get_person_display_name(person)}.",
                reply_markup=MAIN_KEYBOARD,
            )
            return {"ok": True}

        if "web_app_data" in message and message["web_app_data"].get("data"):
            raw_data = message["web_app_data"]["data"]
            try:
                payload = json.loads(raw_data)
                text = (payload.get("asset_tag") or "").strip()
            except Exception:
                text = raw_data.strip()
        elif message.get("text"):
            text = message["text"].strip()

        if text == "/start":
            person = get_authorized_telegram_person(message)
            if person:
                send_telegram_message(
                    chat_id,
                    f"Welcome back, {get_person_display_name(person)}. Choose an action below.",
                    reply_markup=MAIN_KEYBOARD,
                )
            else:
                send_telegram_auth_prompt(chat_id)
            return {"ok": True}

        if not text:
            send_telegram_auth_prompt(chat_id)
            return {"ok": True}

        person = get_authorized_telegram_person(message)

        if text == "Enter code":
            if not person:
                send_telegram_auth_prompt(chat_id)
                return {"ok": True}
            send_telegram_message(
                chat_id,
                "Enter the asset code, for example: HELP-UKR-0015",
            )
            return {"ok": True}

        if text == "Help":
            send_telegram_message(
                chat_id,
                "Authorize with your phone number, then use 'My assets', 'Scan QR', or enter an asset code manually.",
                reply_markup=MAIN_KEYBOARD if person else AUTH_KEYBOARD,
            )
            return {"ok": True}

        if not person:
            send_telegram_auth_prompt(chat_id)
            return {"ok": True}

        if text == "My assets":
            send_telegram_long_message(
                chat_id,
                format_person_assets_message(person),
                reply_markup={
                    "inline_keyboard": [
                        [{"text": "Open asset list", "url": get_telegram_asset_list_url(person)}]
                    ]
                },
            )
            return {"ok": True}

        try:
            asset = get_asset_by_tag(text)
        except DatabaseConnectionError:
            send_telegram_message(
                chat_id,
                "Database is temporarily unavailable. Please try again later.",
            )
            return {"ok": True}

        if not asset:
            send_telegram_message(chat_id, f"Asset {text} was not found.")
            return {"ok": True}

        message_text = format_asset_message(asset)
        url = f"{PUBLIC_BASE_URL}/view/{text}"

        send_telegram_message(
            chat_id,
            message_text,
            reply_markup={
                "inline_keyboard": [
                    [{"text": "🔍 Open asset card", "url": url}]
                ]
            },
        )

        return {"ok": True}

    except Exception as exc:
        print("WEBHOOK ERROR:", exc)
        return {"ok": False, "error": str(exc)}
